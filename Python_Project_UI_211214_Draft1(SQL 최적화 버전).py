#!/usr/bin/env python
# coding: utf-8

# In[1]:


import os
import sys
import re
import datetime
import time
from io import StringIO
from datetime import date
from datetime import timedelta
from dateutil.relativedelta import relativedelta
import gc
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from pytimekr import pytimekr
import pyodbc
import pandas as pd
import numpy as np
import openpyxl
from threading import Thread


class Communicate(QObject):
    closeApp = pyqtSignal()
    closeApp2 = pyqtSignal(str)


class Calendar(QDialog):
    def __init__(self, parent):
        super(Calendar, self).__init__(parent)
        self.MyApp = MyApp

        self.setGeometry(1050, 400, 400, 200)
        self.setWindowTitle("Calendar")
        self.setWindowIcon(QIcon("./EY_logo.png"))
        self.setWindowModality(Qt.NonModal)
        self.setWindowFlag(Qt.FramelessWindowHint)

        vbox = QVBoxLayout()
        hbox = QHBoxLayout()

        self.calendar = QCalendarWidget()
        self.calendar.setGridVisible(True)
        self.calendar.setVerticalHeaderFormat(QCalendarWidget.NoVerticalHeader)

        self.closebtn = QPushButton("Close")

        hbox.addStretch(3)
        hbox.addWidget(self.closebtn)
        vbox.addWidget(self.calendar)
        vbox.addLayout(hbox)
        self.setLayout(vbox)


class Form(QGroupBox):
    def __init__(self, parent):
        super(Form, self).__init__(parent)

        grid = QGridLayout()
        qh = QHBoxLayout()

        self.setLayout(grid)

        self.btnSelect = QPushButton("Select All")
        self.btnSelect.resize(65, 22)
        self.btnSelect.clicked.connect(self.select_all)
        self.btnSelect.clicked.connect(self.get_selected_leaves)
        self.btnSelect.setStyleSheet('color:white;  background-color : #2E2E38')
        font11 = self.btnSelect.font()
        font11.setBold(True)
        self.btnSelect.setFont(font11)

        self.btnUnselect = QPushButton("Unselect All")
        self.btnUnselect.resize(65, 22)
        self.btnUnselect.clicked.connect(self.unselect_all)
        self.btnUnselect.clicked.connect(self.get_selected_leaves)
        self.btnUnselect.setStyleSheet('color:white;  background-color : #2E2E38')
        font11 = self.btnUnselect.font()
        font11.setBold(True)
        self.btnUnselect.setFont(font11)

        self.setStyleSheet('QGroupBox  {color: white; background-color: white}')

        self.tree = QTreeWidget(self)
        self.tree.setStyleSheet("border-style: outset; border-color : white; background-color:white;")

        headerItem = QTreeWidgetItem()
        item = QTreeWidgetItem()

        qh.addWidget(self.btnSelect)
        qh.addWidget(self.btnUnselect)

        grid.addLayout(qh, 0, 0)
        grid.addWidget(self.tree, 1, 0)

        self.tree.setHeaderHidden(True)
        self.tree.itemClicked.connect(self.get_selected_leaves)

    def unselect_all(self):
        def recurse_unselect(parent):
            for i in range(parent.childCount()):
                child = parent.child(i)
                for j in range(child.childCount()):
                    grandchild = child.child(j)
                    grandgrandchild = grandchild.childCount()
                    if grandgrandchild > 0:
                        recurse_unselect(grandchild)
                    else:
                        if grandchild.checkState(0) == Qt.Checked:
                            grandchild.setCheckState(0, Qt.Unchecked)

        recurse_unselect(self.tree.invisibleRootItem())

    def select_all(self):
        def recurse_select(parent):
            for i in range(parent.childCount()):
                child = parent.child(i)
                for j in range(child.childCount()):
                    grandchild = child.child(j)
                    grandgrandchild = grandchild.childCount()
                    if grandgrandchild > 0:
                        recurse_select(grandchild)
                    else:
                        if grandchild.checkState(0) == Qt.Unchecked:
                            grandchild.setCheckState(0, Qt.Checked)

        recurse_select(self.tree.invisibleRootItem())

    def get_selected_leaves(self):
        checked_items = []

        def recurse(parent):
            for i in range(parent.childCount()):
                child = parent.child(i)
                for j in range(child.childCount()):
                    grandchild = child.child(j)
                    grandgrandchild = grandchild.childCount()
                    if grandgrandchild > 0:
                        recurse(grandchild)
                    else:
                        if grandchild.checkState(0) == Qt.Checked:
                            checked_items.append(grandchild.text(0).split(' ')[0])

        recurse(self.tree.invisibleRootItem())

        global checked_name
        checked_name = ''
        for i in checked_items:
            checked_name = checked_name + ',' + '\'' + i + '\''

        checked_name = checked_name[1:]

        global checked_account
        global checked_account_11
        global checked_account_12

        checked_account = 'AND JournalEntries.GLAccountNumber IN (' + checked_name + ')'
        checked_account_11 = checked_name
        checked_account_12 = 'AND LVL4.GL_Account_Number IN (' + checked_name + ')'


class Form1(QGroupBox):
    def __init__(self, parent):
        super(Form1, self).__init__(parent)

        grid = QGridLayout()
        qh = QHBoxLayout()

        self.setLayout(grid)

        self.btnSelect = QPushButton("Select All")
        self.btnSelect.resize(65, 22)
        self.btnSelect.clicked.connect(self.select_all)
        self.btnSelect.clicked.connect(self.get_selected_leaves_1)
        self.btnSelect.setStyleSheet('color:white;  background-color : #2E2E38')
        font11 = self.btnSelect.font()
        font11.setBold(True)
        self.btnSelect.setFont(font11)

        self.btnUnselect = QPushButton("Unselect All")
        self.btnUnselect.resize(65, 22)
        self.btnUnselect.clicked.connect(self.unselect_all)
        self.btnUnselect.clicked.connect(self.get_selected_leaves_1)
        self.btnUnselect.setStyleSheet('color:white;  background-color : #2E2E38')
        font11 = self.btnUnselect.font()
        font11.setBold(True)
        self.btnUnselect.setFont(font11)

        self.setStyleSheet('QGroupBox  {color: white; background-color: white}')

        self.tree = QTreeWidget(self)
        self.tree.setStyleSheet("border-style: outset; border-color : white; background-color:white;")

        headerItem = QTreeWidgetItem()
        item = QTreeWidgetItem()

        qh.addWidget(self.btnSelect)
        qh.addWidget(self.btnUnselect)

        grid.addLayout(qh, 0, 0)
        grid.addWidget(self.tree, 1, 0)

        self.tree.setHeaderHidden(True)
        self.tree.itemClicked.connect(self.get_selected_leaves_1)

    def unselect_all(self):
        def recurse_unselect(parent):
            for i in range(parent.childCount()):
                child = parent.child(i)
                for j in range(child.childCount()):
                    grandchild = child.child(j)
                    grandgrandchild = grandchild.childCount()
                    if grandgrandchild > 0:
                        recurse_unselect(grandchild)
                    else:
                        if grandchild.checkState(0) == Qt.Checked:
                            grandchild.setCheckState(0, Qt.Unchecked)

        recurse_unselect(self.tree.invisibleRootItem())

    def select_all(self):
        def recurse_select(parent):
            for i in range(parent.childCount()):
                child = parent.child(i)
                for j in range(child.childCount()):
                    grandchild = child.child(j)
                    grandgrandchild = grandchild.childCount()
                    if grandgrandchild > 0:
                        recurse_select(grandchild)
                    else:
                        if grandchild.checkState(0) == Qt.Unchecked:
                            grandchild.setCheckState(0, Qt.Checked)

        recurse_select(self.tree.invisibleRootItem())

    def get_selected_leaves_1(self):
        checked_items = []

        def recurse(parent):
            for i in range(parent.childCount()):
                child = parent.child(i)
                for j in range(child.childCount()):
                    grandchild = child.child(j)
                    grandgrandchild = grandchild.childCount()
                    if grandgrandchild > 0:
                        recurse(grandchild)
                    else:
                        if grandchild.checkState(0) == Qt.Checked:
                            checked_items.append(grandchild.text(0).split(' ')[0])

        recurse(self.tree.invisibleRootItem())

        global checked_name
        checked_name = ''
        for i in checked_items:
            checked_name = checked_name + ',' + '\'' + i + '\''

        checked_name = checked_name[1:]

        global checked_account_11_1
        checked_account_11_1 = checked_name


class Preparer(QGroupBox):
    def __init__(self, parent):
        super(Preparer, self).__init__(parent)

        grid = QGridLayout()
        qh = QHBoxLayout()

        self.setLayout(grid)
        self.setStyleSheet('QGroupBox  {color: white; background-color: white}')

        self.prep = QTreeWidget(self)
        self.prep.setStyleSheet("border-style: outset; border-color : white; background-color:white;")

        headerItem = QTreeWidgetItem()
        item = QTreeWidgetItem()

        self.btnSelectp = QPushButton("Select All")
        self.btnSelectp.resize(65, 22)
        self.btnSelectp.clicked.connect(self.select_all)
        self.btnSelectp.clicked.connect(self.get_selected_leaves)
        self.btnSelectp.setStyleSheet('color:white;  background-color : #2E2E38')
        font11 = self.btnSelectp.font()
        font11.setBold(True)
        self.btnSelectp.setFont(font11)

        self.btnUnselectp = QPushButton("Unselect All")
        self.btnUnselectp.resize(65, 22)
        self.btnUnselectp.clicked.connect(self.unselect_all)
        self.btnUnselectp.clicked.connect(self.get_selected_leaves)
        self.btnUnselectp.setStyleSheet('color:white;  background-color : #2E2E38')
        font11 = self.btnUnselectp.font()
        font11.setBold(True)
        self.btnUnselectp.setFont(font11)

        qh.addWidget(self.btnSelectp)
        qh.addWidget(self.btnUnselectp)

        grid.addLayout(qh, 0, 0)
        grid.addWidget(self.prep, 1, 0)
        self.prep.setHeaderHidden(True)
        self.prep.itemClicked.connect(self.get_selected_leaves)

    def unselect_all(self):
        for i in range(self.prep.topLevelItemCount()):
            self.prep.topLevelItem(i).setCheckState(0, Qt.Unchecked)

    def select_all(self):
        for i in range(self.prep.topLevelItemCount()):
            self.prep.topLevelItem(i).setCheckState(0, Qt.Checked)

    def get_selected_leaves(self):
        checked_items = []
        for i in range(self.prep.topLevelItemCount()):
            if self.prep.topLevelItem(i).checkState(0) == Qt.Checked:
                checked_items.append(self.prep.topLevelItem(i).text(0).split(' ')[0])

        global checked_prep

        checked_prep = ''
        for i in checked_items:
            checked_prep = checked_prep + ',' + '\'' + i + '\''

        checked_prep = checked_prep[1:]

        global checked_preparer

        checked_preparer = 'AND JournalEntries.PreparerID IN (' + checked_prep + ')'


class DataFrameModel(QAbstractTableModel):
    DtypeRole = Qt.UserRole + 1000
    ValueRole = Qt.UserRole + 1001

    def __init__(self, df=pd.DataFrame(), parent=None):
        super(DataFrameModel, self).__init__(parent)
        self._dataframe = df

    def setDataFrame(self, dataframe):
        self.beginResetModel()
        self._dataframe = dataframe.copy()
        self.endResetModel()

    def dataFrame(self):
        return self._dataframe

    dataFrame = pyqtProperty(pd.DataFrame, fget=dataFrame, fset=setDataFrame)

    @pyqtSlot(int, Qt.Orientation, result=str)
    def headerData(self, section: int, orientation: Qt.Orientation, role: int = Qt.DisplayRole):
        if role == Qt.DisplayRole:
            if orientation == Qt.Horizontal:
                return self._dataframe.columns[section]
            else:
                return str(self._dataframe.index[section])
        return QVariant()

    def rowCount(self, parent=QModelIndex()):
        if parent.isValid():
            return 0
        return len(self._dataframe.index)

    def columnCount(self, parent=QModelIndex()):
        if parent.isValid():
            return 0
        return self._dataframe.columns.size

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid() or not (0 <= index.row() < self.rowCount() and 0 <= index.column() < self.columnCount()):
            return QVariant()
        row = self._dataframe.index[index.row()]
        col = self._dataframe.columns[index.column()]
        dt = self._dataframe[col].dtype

        val = self._dataframe.iloc[row][col]
        if role == Qt.DisplayRole:
            return str(val)
        elif role == DataFrameModel.ValueRole:
            return val
        if role == DataFrameModel.DtypeRole:
            return dt
        return QVariant()

    def roleNames(self):
        roles = {
            Qt.DisplayRole: b'display',
            DataFrameModel.DtypeRole: b'dtype',
            DataFrameModel.ValueRole: b'value'
        }
        return roles


class ListBoxWidget(QListWidget):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.resize(600, 600)

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls:
            event.accept()
        else:
            event.ignore()

    def dragMoveEvent(self, event):
        if event.mimeData().hasUrls():
            event.setDropAction(Qt.CopyAction)
            event.accept()
        else:
            event.ignore()

    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            event.setDropAction(Qt.CopyAction)
            event.accept()

            links = []
            for url in event.mimeData().urls():

                if url.isLocalFile():
                    links.append(str(url.toLocalFile()))
                else:
                    links.append(str(url.toString()))
            self.addItems(links)

        else:
            event.ignore()


class MyApp(QWidget):
    # Resource
    def resource_path(self, relative_path):
        try:
            # PyInstaller에 의해 임시폴더에서 실행될 경우 임시폴더로 접근하는 함수
            base_path = sys._MEIPASS
        except Exception:
            base_path = os.path.abspath(".")
        return os.path.join(base_path, relative_path)

    def __init__(self):
        super().__init__()
        self.init_UI()

        ##Initialize Variables
        ####수정사항 시작####
        self.selected_project_id = None
        self.selected_server_name = "--서버 목록--"
        self.dataframe = None
        self.dataframe_refer = None
        self.cnxn = None
        self.my_query = None
        self.selected_scenario_subclass_index = 0
        self.scenario_dic = {}
        self.new_calendar = None
        self.new_tree = None
        self.new_prep = None
        self.dateList = []
        self.string_date_list = []
        self.fianlDate = []
        self.clickCount = 0
        self.dialoglist = set()
        self.timerVar = QTimer()
        self.timerVar.setInterval(1000)
        self.timerVar.timeout.connect(self.printTime)
        ####수정사항 끝####

        ##다이얼로그별 시그널 생성
        self.communicate4 = Communicate()
        self.communicate4.closeApp.connect(self.doneAction4)
        self.communicate5_Non_SAP = Communicate()
        self.communicate5_Non_SAP.closeApp.connect(self.doneAction5_Non_SAP)
        self.communicate5_SAP = Communicate()
        self.communicate5_SAP.closeApp.connect(self.doneAction5_SAP)
        self.communicate6 = Communicate()
        self.communicate6.closeApp.connect(self.doneAction6)
        self.communicate7 = Communicate()
        self.communicate7.closeApp.connect(self.doneAction7)
        self.communicate8 = Communicate()
        self.communicate8.closeApp.connect(self.doneAction8)
        self.communicate9 = Communicate()
        self.communicate9.closeApp.connect(self.doneAction9)
        self.communicate10 = Communicate()
        self.communicate10.closeApp.connect(self.doneAction10)
        self.communicate12 = Communicate()
        self.communicate12.closeApp.connect(self.doneAction12)
        self.communicate13 = Communicate()
        self.communicate13.closeApp.connect(self.doneAction13)
        self.communicate14 = Communicate()
        self.communicate14.closeApp.connect(self.doneAction14)
        self.communicateC = Communicate()
        self.communicateC.closeApp2.connect(self.doneActionC)

    def return_print(self, *message):
        self.io = StringIO()
        print(*message, file=self.io, end="")
        return self.io.getvalue()

    def MessageBox_Open(self, text):
        self.msg = QMessageBox()
        self.msg.setIcon(QMessageBox.Information)
        self.msg.setWindowTitle("Warning")
        self.msg.setWindowIcon(QIcon('./EY_logo.png'))
        self.msg.setText(text)
        self.msg.exec_()

    def MessageBox_Open2(self, text):
        self.msg = QMessageBox()
        self.msg.setIcon(QMessageBox.Information)
        self.msg.setWindowTitle("Project Connected")
        self.msg.setWindowIcon(QIcon('./EY_logo.png'))
        self.msg.setText(text)
        self.msg.exec_()

    def alertbox_open(self):
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        self.alt.setWindowTitle('필수 입력값 누락')
        self.alt.setWindowIcon(QIcon('./EY_logo.png'))
        self.alt.setText('필수 입력값이 누락되었습니다.')
        self.alt.exec_()

    def alertbox_open2(self, state):
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        txt = state
        self.alt.setWindowTitle('필수 입력값 타입 오류')
        self.alt.setWindowIcon(QIcon('./EY_logo.png'))
        self.alt.setText(txt + ' 값을 숫자로만 입력해주시기 바랍니다.')
        self.alt.exec_()

    def alertbox_open3(self):
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        self.alt.setWindowTitle('최대 라인 수 초과 오류')
        self.alt.setWindowIcon(QIcon('./EY_logo.png'))
        self.alt.setText('최대 라인 수가 초과 되었습니다.')
        self.alt.exec_()

    def alertbox_open4(self, state):
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        txt = state
        self.alt.setWindowTitle('입력값 타입 오류')
        self.alt.setWindowIcon(QIcon('./EY_logo.png'))
        self.alt.setText(txt)
        self.alt.exec_()

    def alertbox_open5(self):
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        self.alt.setWindowTitle('시트명 중복')
        self.alt.setWindowIcon(QIcon('./EY_logo.png'))
        self.alt.setText('이미 해당 시트명이 존재합니다.')
        self.alt.exec_()

    def alertbox_open7(self):
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        self.alt.setWindowTitle('차대변 선택 오류')
        self.alt.setWindowIcon(QIcon('./EY_logo.png'))
        self.alt.setText('차변 및 대변이 선택되어 있지 않습니다.')
        self.alt.exec_()

    def alertbox_open10(self):
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        self.alt.setWindowTitle('SKA1 파일 선택 오류')
        self.alt.setWindowIcon(QIcon('./EY_logo.png'))
        self.alt.setText('SKA1 파일만을 드롭하세요.')
        self.alt.exec_()

    def alertbox_open11(self, state):
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        txt = state
        self.alt.setWindowTitle('SKA1 파일 내부 필드값 부재')
        self.alt.setWindowIcon(QIcon('./EY_logo.png'))
        self.alt.setText(txt)
        self.alt.exec_()

    def alertbox_open12(self):
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        self.alt.setWindowTitle('필수 입력값 오류')
        self.alt.setWindowIcon(QIcon("./EY_logo.png"))
        self.alt.setText('T일은 0이상 70만 미만의 정수로만 입력 바랍니다.')
        self.alt.exec_()

    def alertbox_open13(self):
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        self.alt.setWindowTitle('필수 입력값 오류')
        self.alt.setWindowIcon(QIcon("./EY_logo.png"))
        self.alt.setText('N일은 0이상 70만 미만의 정수로만 입력 바랍니다.')
        self.alt.exec_()

    def alertbox_open14(self):
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        self.alt.setWindowTitle('전 후 선택 오류')
        self.alt.setWindowIcon(QIcon("./EY_logo.png"))
        self.alt.setText('T일 이전 및 T일 이후가 선택되어 있지 않습니다.')
        self.alt.exec_()

    def alertbox_open15(self):
        self.alt = QMessageBox()
        self.alt.setIcon(QMessageBox.Information)
        self.alt.setWindowTitle('자릿수 포맷 오류')
        self.alt.setWindowIcon(QIcon("./EY_logo.png"))
        self.alt.setText('연속된 자릿수는 6자리 이상으로 입력해주세요.')
        self.alt.exec_()

    def init_UI(self):

        image = QImage('./dark_gray.png')
        scaledImg = image.scaled(QSize(1000, 900))
        palette = QPalette()
        palette.setBrush(10, QBrush(scaledImg))
        self.setPalette(palette)

        pixmap = QPixmap('./title.png')
        lbl_img = QLabel()
        lbl_img.setPixmap(pixmap)
        lbl_img.setStyleSheet("background-color:#000000")

        widget_layout = QBoxLayout(QBoxLayout.TopToBottom)
        self.splitter_layout = QSplitter(Qt.Vertical)

        self.splitter_layout.addWidget(lbl_img)
        self.splitter_layout.addWidget(self.Connect_ServerInfo_Group())
        self.splitter_layout.addWidget(self.Show_DataFrame_Group())
        self.splitter_layout.addWidget(self.Save_Buttons_Group())
        self.splitter_layout.setHandleWidth(0)
        self.splitter_layout.setStretchFactor(0, 3)
        self.splitter_layout.setStretchFactor(1, 2)
        self.splitter_layout.setStretchFactor(2, 4)
        self.splitter_layout.setStretchFactor(3, 1)

        widget_layout.addWidget(self.splitter_layout)
        self.setLayout(widget_layout)

        self.setWindowIcon(QIcon("./EY_logo.png"))
        self.setWindowTitle('Scenario')

        self.setGeometry(300, 100, 1000, 900)
        self.show()

    def connectButtonClicked(self):

        password = ''
        ecode = self.line_ecode.text().strip()  ##leading/trailing space 포함시 제거
        user = 'guest'
        server = self.selected_server_name
        db = 'master'

        # 예외처리 - 서버 선택
        if server == "--서버 목록--":
            self.MessageBox_Open("서버가 선택되어 있지 않습니다")
            return

        # 예외처리 - Ecode 이상
        elif ecode.isdigit() is False:
            self.MessageBox_Open("Engagement Code가 잘못되었습니다")
            self.ProjectCombobox.clear()
            self.ProjectCombobox.addItem("프로젝트가 없습니다")
            return

        server_path = f"DRIVER={{SQL Server}};SERVER={server};uid={user};pwd={password};DATABASE={db};trusted_connection=yes"

        # 예외처리 - 접속 정보 오류
        try:
            self.cnxn = pyodbc.connect(server_path)
        except:
            QMessageBox.about(self, "Warning", "접속 정보가 잘못되었습니다")
            return

        cursor = self.cnxn.cursor()

        sql_query = f"""
                           SELECT ProjectName
                           From [DataAnalyticsRepository].[dbo].[Projects]
                           WHERE EngagementCode IN ({ecode})
                           AND DeletedBy IS NULL
                     """

        try:
            selected_project_names = pd.read_sql(sql_query, self.cnxn)

        except:
            self.MessageBox_Open("Engagement Code를 입력하세요.")
            self.ProjectCombobox.clear()
            self.ProjectCombobox.addItem("프로젝트가 없습니다")
            return

        # 예외처리 - 해당 ecode에 아무런 프로젝트가 없는 경우
        if len(selected_project_names) == 0:
            self.MessageBox_Open("해당 Engagement Code 내 프로젝트가 존재하지 않습니다.")
            self.ProjectCombobox.clear()
            self.ProjectCombobox.addItem("프로젝트가 없습니다.")
            return
        else:
            self.MessageBox_Open2("프로젝트가 연결되었습니다.")

        ## 서버 연결 시 - 기존 저장 정보를 초기화(메모리 관리)
        del self.selected_project_id, self.dataframe, self.scenario_dic, self.my_query

        self.my_query = pd.DataFrame(columns=["Sheet name", "Scenario number", "Query"])

        self.ProjectCombobox.clear()
        self.ProjectCombobox.addItem("--프로젝트 목록--")
        for i in range(len(selected_project_names)):
            self.ProjectCombobox.addItem(selected_project_names.iloc[i, 0])

        self.combo_sheet.clear()

        self.selected_project_id = None
        self.dataframe = None
        ####수정사항 시작####
        self.dataframe_refer = None
        self.viewtable.setModel(self.dataframe)
        ####수정사항 끝####
        self.scenario_dic = {}
        self.string_date_list = []
        self.fianlDate = []
        self.clickCount = 0

        gc.collect()

    def Server_ComboBox_Selected(self, text):
        self.selected_server_name = text

    def Project_ComboBox_Selected(self, text):
        ####수정사항 시작####
        ## 예외처리 - 서버가 연결되지 않은 상태로 Project name Combo box를 건드리는 경우
        if self.cnxn is None:
            return

        ecode = self.line_ecode.text().strip()  # leading/trailing space 제거
        pname = text
        self.pname_year = "20" + str(pname)[2:4]  # str
        cursor = self.cnxn.cursor()

        sql_query = f"""
                                SELECT Project_ID
                                FROM [DataAnalyticsRepository].[dbo].[Projects]
                                WHERE ProjectName IN (\'{pname}\')
                                AND EngagementCode IN ({ecode})
                                AND DeletedBy is Null
                             """

        ## 예외처리 - 에러 표시인 "프로젝트가 없습니다"가 선택되어 있는 경우
        try:
            self.selected_project_id = pd.read_sql(sql_query, self.cnxn).iloc[0, 0]

        except:
            self.selected_project_id = None
        ####수정사항 끝####

    def Connect_ServerInfo_Group(self):

        groupbox = QGroupBox('접속 정보')
        self.setStyleSheet('QGroupBox  {color: white;}')
        font_groupbox = groupbox.font()
        font_groupbox.setBold(True)
        groupbox.setFont(font_groupbox)

        ##labels 생성 및 스타일 지정
        label1 = QLabel('Server : ', self)
        label2 = QLabel('Engagement Code : ', self)
        label3 = QLabel('Project Name : ', self)
        label4 = QLabel('Scenario : ', self)

        font1 = label1.font()
        font1.setBold(True)
        label1.setFont(font1)
        font2 = label2.font()
        font2.setBold(True)
        label2.setFont(font2)
        font3 = label3.font()
        font3.setBold(True)
        label3.setFont(font3)
        font4 = label4.font()
        font4.setBold(True)
        label4.setFont(font4)

        label1.setStyleSheet("color: white;")
        label2.setStyleSheet("color: white;")
        label3.setStyleSheet("color: white;")
        label4.setStyleSheet("color: white;")

        ##서버 선택 콤보박스
        self.cb_server = QComboBox(self)
        self.cb_server.addItem('--서버 목록--')
        ####수정사항 시작####
        for i in [1, 2, 3, 4, 6, 7, 8]:
            self.cb_server.addItem(f'KRSEOVMPPACSQ0{i}\INST1')
        ####수정사항 끝####

        ### Scenario 유형 콤보박스 - 소분류 수정
        self.comboScenario = QComboBox(self)

        self.comboScenario.addItem('--시나리오 목록--')
        self.comboScenario.addItem('04 : 계정 사용빈도 N번 이하인 계정이 포함된 전표리스트')
        self.comboScenario.addItem('05 : 당기 생성된 계정리스트 추출')
        self.comboScenario.addItem('06 : 결산일 전후 T일 입력 전표')
        self.comboScenario.addItem('07 : 비영업일 전기/입력 전표')
        self.comboScenario.addItem('08 : 효력, 입력 일자 간 차이가 N일 이상인 전표')
        self.comboScenario.addItem('09 : 전표 작성 빈도수가 N회 이하인 작성자에 의한 생성된 전표')
        self.comboScenario.addItem('10 : 특정 전표 입력자(W)에 의해 생성된 전표')
        self.comboScenario.addItem('11-12 : 특정 계정(A) 상대계정 리스트 검토')
        self.comboScenario.addItem('13 : 연속된 숫자로 끝나는 금액 검토')
        self.comboScenario.addItem('14 : 전표 description에 공란 또는 특정단어(key word)가 입력되어 있는 전표 리스트 (중요성 금액 제시 가능)')

        self.ProjectCombobox = QComboBox(self)

        ##Engagement code 입력 line
        self.line_ecode = QLineEdit(self)
        self.line_ecode.setText("")

        ##Project Connect 버튼 생성 및 스타일 지정
        btn_connect = QPushButton('   Project Connect', self)
        font_btn_connect = btn_connect.font()
        font_btn_connect.setBold(True)
        btn_connect.setFont(font_btn_connect)
        btn_connect.setStyleSheet('color:white;  background-image : url(./bar.png)')

        ##Input Conditions 버튼 생성 및 스타일 지정
        btn_condition = QPushButton('   Input Conditions', self)
        font_btn_condition = btn_condition.font()
        font_btn_condition.setBold(True)
        btn_condition.setFont(font_btn_condition)
        btn_condition.setStyleSheet('color:white;  background-image : url(./bar.png)')

        ##Index 버튼 생성 및 스타일 지정
        btn_index = QPushButton('   Sheet Index', self)
        font_btn_index = btn_index.font()
        font_btn_index.setBold(True)
        btn_index.setFont(font_btn_index)
        btn_index.setStyleSheet('color:white;  background-image : url(./bar.png)')

        ### Signal 함수들
        self.comboScenario.activated[str].connect(self.ComboSmall_Selected)
        self.cb_server.activated[str].connect(self.Server_ComboBox_Selected)
        btn_connect.clicked.connect(self.connectButtonClicked)
        self.ProjectCombobox.activated[str].connect(self.Project_ComboBox_Selected)
        btn_condition.clicked.connect(self.connectDialog)
        btn_index.clicked.connect(self.DialogIndex)

        ##layout 쌓기
        grid = QGridLayout()
        grid.addWidget(label1, 0, 0)
        grid.addWidget(label2, 1, 0)
        grid.addWidget(label3, 2, 0)
        grid.addWidget(label4, 3, 0)
        grid.addWidget(self.cb_server, 0, 1)
        grid.addWidget(btn_connect, 1, 2)
        grid.addWidget(self.comboScenario, 3, 1)
        grid.addWidget(btn_condition, 3, 2)
        grid.addWidget(self.line_ecode, 1, 1)
        grid.addWidget(self.ProjectCombobox, 2, 1)
        grid.addWidget(btn_index, 4, 2)

        groupbox.setLayout(grid)
        return groupbox

    def ComboSmall_Selected(self, text):
        self.selected_scenario_subclass_index = self.comboScenario.currentIndex()

    def connectDialog(self):
        if self.cnxn is None:
            self.MessageBox_Open("SQL 서버가 연결되어 있지 않습니다")
            return

        elif self.selected_project_id is None:
            self.MessageBox_Open("프로젝트가 선택되지 않았습니다")
            return

        elif self.selected_scenario_subclass_index == 0:
            self.MessageBox_Open("시나리오가 선택되지 않았습니다")
            return

        if self.selected_scenario_subclass_index == 1:
            self.Dialog4()

        elif self.selected_scenario_subclass_index == 2:
            self.Dialog5()

        elif self.selected_scenario_subclass_index == 3:
            self.Dialog6()

        elif self.selected_scenario_subclass_index == 4:
            self.Dialog7()

        elif self.selected_scenario_subclass_index == 5:
            self.Dialog8()

        elif self.selected_scenario_subclass_index == 6:
            self.Dialog9()

        elif self.selected_scenario_subclass_index == 7:
            self.Dialog10()

        elif self.selected_scenario_subclass_index == 8:
            self.Dialog12()

        elif self.selected_scenario_subclass_index == 9:
            self.Dialog13()

        elif self.selected_scenario_subclass_index == 10:
            self.Dialog14()

        gc.collect()

    def Dialog4(self):
        self.dialoglist.add(4)
        self.dialog4 = QDialog()
        self.dialog4.setStyleSheet('background-color: #2E2E38')
        self.dialog4.setWindowIcon(QIcon('./EY_logo.png'))

        # 트리 작업
        cursor = self.cnxn.cursor()

        sql = '''
                 SELECT 											
                        *
                 FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
            '''.format(field=self.selected_project_id)

        accountsname = pd.read_sql(sql, self.cnxn)

        self.new_tree = Form(self)

        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)

            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Checked)
        self.new_tree.get_selected_leaves()  # 초기값 모두 선택 (추가)

        ### 버튼 1 - Extract Data
        self.btn2 = QPushButton('   Extract Data', self.dialog4)
        self.btn2.setStyleSheet('color:white; background-image : url(./bar.png)')
        self.btn2.clicked.connect(self.Thread4)

        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)

        ### 버튼 2 - Close
        self.btnDialog = QPushButton('   Close', self.dialog4)
        self.btnDialog.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close4)

        font10 = self.btnDialog.font()
        font10.setBold(True)
        self.btnDialog.setFont(font10)

        # JE Line / JE 라디오 버튼
        self.rbtn1 = QRadioButton('JE Line', self.dialog4)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)

        self.rbtn2 = QRadioButton('JE', self.dialog4)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        ### 라벨 1 - 사용빈도
        label_freq = QLabel('사용 빈도(N)* :', self.dialog4)
        label_freq.setStyleSheet('color: white;')

        font1 = label_freq.font()
        font1.setBold(True)
        label_freq.setFont(font1)

        ### LineEdit 1 - 사용 빈도
        self.D4_N = QLineEdit(self.dialog4)
        self.D4_N.setStyleSheet('background-color: white;')
        self.D4_N.setPlaceholderText('사용빈도를 입력하세요')

        ### 라벨 2 - 중요성 금액
        label_TE = QLabel('중요성 금액: ', self.dialog4)
        label_TE.setStyleSheet('color: white;')

        font2 = label_TE.font()
        font2.setBold(True)
        label_TE.setFont(font2)

        ### LineEdit 2 - 중요성 금액
        self.D4_TE = QLineEdit(self.dialog4)
        self.D4_TE.setStyleSheet('background-color: white;')
        self.D4_TE.setPlaceholderText('중요성 금액을 입력하세요')

        ### 라벨 3 - 시트명
        labelSheet = QLabel('시나리오 번호* : ', self.dialog4)
        labelSheet.setStyleSheet("color: white;")

        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)

        ### LineEdit 3 - 시트명
        self.D4_Sheet = QLineEdit(self.dialog4)
        self.D4_Sheet.setStyleSheet("background-color: white;")
        self.D4_Sheet.setPlaceholderText('※ 입력 예시 : F01')

        label_tree = QLabel('특정 계정명 : ', self.dialog4)
        label_tree.setStyleSheet("color: white;")
        font4 = label_tree.font()
        font4.setBold(True)
        label_tree.setFont(font4)

        # 차변/대변 체크박스로 구현
        labelDC = QLabel('차변/대변* : ', self.dialog4)
        labelDC.setStyleSheet("color: white;")
        font1 = labelDC.font()
        font1.setBold(True)
        labelDC.setFont(font1)
        self.checkC = QCheckBox('Credit', self.dialog4)
        self.checkD = QCheckBox('Debit', self.dialog4)
        self.checkC.setStyleSheet("color: white;")
        self.checkD.setStyleSheet("color: white;")
        self.checkC.setChecked(True)
        self.checkD.setChecked(True)

        self.D4_N.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소
        self.D4_TE.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소
        self.D4_Sheet.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소

        ### 요소 배치
        layout1 = QGridLayout()
        layout1.addWidget(self.rbtn1, 0, 0)
        layout1.addWidget(self.rbtn2, 0, 1)
        layout1.addWidget(label_freq, 1, 0)
        layout1.addWidget(self.D4_N, 1, 1)
        layout1.addWidget(label_TE, 2, 0)
        layout1.addWidget(self.D4_TE, 2, 1)
        layout1.addWidget(label_tree, 3, 0)
        layout1.addWidget(self.new_tree, 3, 1)
        layout1.addWidget(labelSheet, 4, 0)
        layout1.addWidget(self.D4_Sheet, 4, 1)

        layout2 = QHBoxLayout()
        layout2.addStretch(2)
        layout2.addWidget(self.btn2)
        layout2.addWidget(self.btnDialog)

        layout2.setContentsMargins(-1, 10, -1, -1)

        layout_dc = QHBoxLayout()
        layout_dc.addWidget(labelDC)
        layout_dc.addWidget(self.checkC)
        layout_dc.addWidget(self.checkD)

        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop)
        main_layout.addLayout(layout1)
        main_layout.addLayout(layout_dc)
        main_layout.addLayout(layout2)
        self.dialog4.setLayout(main_layout)
        self.dialog4.setGeometry(300, 300, 600, 400)

        # ? 제거
        self.dialog4.setWindowFlags(Qt.WindowCloseButtonHint)

        self.dialog4.setWindowTitle('Scenario4')
        self.dialog4.setWindowModality(Qt.NonModal)
        self.dialog4.show()

    def Dialog5(self):
        ### 공통 elements================================================================
        self.dialoglist.add(5)
        self.dialog5 = QDialog()
        self.dialog5.setStyleSheet('background-color: #2E2E38')
        self.dialog5.setWindowIcon(QIcon('./EY_logo.png'))

        ### JE Line / JE 라디오 버튼
        self.rbtn1 = QRadioButton('JE Line             ', self.dialog5)  # 간격 유지를 위한 공백 13개
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)

        self.rbtn2 = QRadioButton('JE', self.dialog5)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        ### Non-SAP======================================================================
        ### 버튼 1 - Extract Data (Non-SAP)
        self.btn2 = QPushButton('   Extract Data', self.dialog5)
        self.btn2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btn2.clicked.connect(self.Thread5_Non_SAP)

        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)
        self.btn2.resize(110, 30)

        ### 버튼 2 - Close (Non-SAP)
        self.btnDialog1 = QPushButton('Close', self.dialog5)
        self.btnDialog1.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDialog1.clicked.connect(self.dialog_close5)

        font11 = self.btnDialog1.font()
        font11.setBold(True)
        self.btnDialog1.setFont(font11)
        self.btnDialog1.resize(110, 30)

        ### 계정 트리
        cursor2 = self.cnxn.cursor()
        sql2 = '''
                         SELECT
                                *
                         FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA
                    '''.format(field=self.selected_project_id)
        accountsname2 = pd.read_sql(sql2, self.cnxn)

        self.new_tree2 = Form(self)
        self.new_tree2.tree.clear()
        accountType2 = accountsname2.AccountType.unique()
        accountType2.sort()
        for n, i in enumerate(accountType2):
            self.new_tree2.parent = QTreeWidgetItem(self.new_tree2.tree)

            self.new_tree2.parent.setText(0, "{}".format(i))
            self.new_tree2.parent.setFlags(self.new_tree2.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname2.AccountSubType[
                accountsname2.AccountType == accountType2[n]].unique()
            child_items.sort()
            for m, x in enumerate(child_items):
                self.new_tree2.child = QTreeWidgetItem(self.new_tree2.parent)

                self.new_tree2.child.setText(0, "{}".format(x))
                self.new_tree2.child.setFlags(self.new_tree2.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname2.AccountClass[accountsname2.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree2.grandchild = QTreeWidgetItem(self.new_tree2.child)

                    self.new_tree2.grandchild.setText(0, "{}".format(y))
                    self.new_tree2.grandchild.setFlags(
                        self.new_tree2.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname2[accountsname2.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree2.grandgrandchild = QTreeWidgetItem(self.new_tree2.grandchild)

                        self.new_tree2.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree2.grandgrandchild.setFlags(
                            self.new_tree2.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree2.grandgrandchild.setCheckState(0, Qt.Checked)
        self.new_tree2.get_selected_leaves()  # 초기값 모두 선택 (추가)

        # 차변/대변 체크박스로 구현
        labelDC1 = QLabel('차변/대변* : ', self.dialog5)
        labelDC1.setStyleSheet("color: white;")
        font1 = labelDC1.font()
        font1.setBold(True)
        labelDC1.setFont(font1)
        self.checkC1 = QCheckBox('Credit', self.dialog5)
        self.checkD1 = QCheckBox('Debit', self.dialog5)
        self.checkC1.setStyleSheet("color: white;")
        self.checkD1.setStyleSheet("color: white;")
        self.checkC1.setChecked(True)
        self.checkD1.setChecked(True)

        ### 라벨1 - 계정코드 입력
        label_AccCode = QLabel('계정코드* : ', self.dialog5)
        label_AccCode.setStyleSheet('color: white;')

        font1 = label_AccCode.font()
        font1.setBold(True)
        label_AccCode.setFont(font1)

        ### 라벨 2 - 시트명
        labelSheet2 = QLabel('시나리오 번호* : ', self.dialog5)
        labelSheet2.setStyleSheet("color: white;")
        font5 = labelSheet2.font()
        font5.setBold(True)
        labelSheet2.setFont(font5)

        ### 라벨 3 - 계정 트리
        label_tree2 = QLabel('특정 계정명 : ', self.dialog5)
        label_tree2.setStyleSheet("color: white;")
        font40 = label_tree2.font()
        font40.setBold(True)
        label_tree2.setFont(font40)

        ### 라벨 4 - 계정 코드 입력 예시
        label_ex = QLabel('※ 입력 예시 : OO', self.dialog5)
        label_ex.setStyleSheet('color: red;')
        font30 = label_ex.font()
        font30.setBold(False)
        label_ex.setFont(font30)

        ### TextEdit 1 - 계정코드 Paste
        self.MyInput = QTextEdit(self.dialog5)
        self.MyInput.setAcceptRichText(False)
        self.MyInput.setStyleSheet('background-color: white;')
        self.MyInput.setPlaceholderText('계정코드를 입력하세요')

        ### LineEdit 1 - 시트명
        self.D5_Sheet2 = QLineEdit(self.dialog5)
        self.D5_Sheet2.setStyleSheet("background-color: white;")
        self.D5_Sheet2.setPlaceholderText('※ 입력 예시 : F01')

        ### SAP=============================================================================
        ### 버튼 1 - Clear Files
        self.gbtn = QPushButton('Clear Files', self.dialog5)
        self.gbtn.setStyleSheet('color:white; background-image: url(./bar.png)')
        self.gbtn.clicked.connect(self.dropFiles)

        font90 = self.gbtn.font()
        font90.setBold(True)
        self.gbtn.setFont(font90)
        self.gbtn.resize(110, 30)

        ### 버튼 2 - Extract Data
        self.btn3 = QPushButton('   Extract Data', self.dialog5)
        self.btn3.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btn3.clicked.connect(self.Thread5_SAP)

        font11 = self.btn3.font()
        font11.setBold(True)
        self.btn3.setFont(font11)
        self.btn3.resize(110, 30)

        ### 버튼 3 - Close
        self.btnDialog = QPushButton('Close', self.dialog5)
        self.btnDialog.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close5)

        font9 = self.btnDialog.font()
        font9.setBold(True)
        self.btnDialog.setFont(font9)
        self.btnDialog.resize(110, 30)

        ### 계정 트리
        cursor = self.cnxn.cursor()
        sql = '''
                         SELECT
                                *
                         FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA
                    '''.format(field=self.selected_project_id)

        accountsname = pd.read_sql(sql, self.cnxn)

        self.new_tree = Form(self)
        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)

            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Checked)
        self.new_tree.get_selected_leaves()  # 초기값 모두 선택 (추가)

        ### 라벨 0 - 차/대변
        labelDC2 = QLabel('차변/대변* : ', self.dialog5)
        labelDC2.setStyleSheet("color: white;")
        font1 = labelDC2.font()
        font1.setBold(True)
        labelDC2.setFont(font1)

        ### 체크박스 - 차/대변
        self.checkC2 = QCheckBox('Credit', self.dialog5)
        self.checkD2 = QCheckBox('Debit', self.dialog5)
        self.checkC2.setStyleSheet("color: white;")
        self.checkD2.setStyleSheet("color: white;")

        self.checkC2.setChecked(True)
        self.checkD2.setChecked(True)

        ### 라벨 1-1 - SKA1
        label_SKA1_text = QLabel('SKA1* : ', self.dialog5)
        label_SKA1_text.setStyleSheet('color:white;')

        font18 = label_SKA1_text.font()
        font18.setBold(True)
        label_SKA1_text.setFont(font18)

        ### 라벨 1-2 - SKA1 파일 드롭하기
        label_SKA1 = QLabel('※ SKA1 파일을 Drop 하세요', self.dialog5)
        label_SKA1.setStyleSheet('color: red;')

        font12 = label_SKA1.font()
        font12.setBold(False)
        label_SKA1.setFont(font12)

        ### 라벨 2 - 시트명
        labelSheet = QLabel('시나리오 번호* : ', self.dialog5)
        labelSheet.setStyleSheet("color: white;")
        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)

        ### 라벨 3 - 계정 트리
        label_tree = QLabel('특정 계정명 : ', self.dialog5)
        label_tree.setStyleSheet("color: white;")
        font40 = label_tree.font()
        font40.setBold(True)
        label_tree.setFont(font40)

        ### ListBox Widget 1 - SKA1
        self.listbox_drops = ListBoxWidget()
        self.listbox_drops.setStyleSheet('background-color: white;')

        ### LineEdit 1 - 시트명
        self.D5_Sheet = QLineEdit(self.dialog5)
        self.D5_Sheet.setStyleSheet("background-color: white;")
        self.D5_Sheet.setPlaceholderText('※ 입력 예시 : F01')

        ### Layout 구성=====================================================================

        layout = QVBoxLayout()

        layout1 = QVBoxLayout()
        sublayout2 = QHBoxLayout()
        sublayout5 = QGridLayout()

        layout2 = QVBoxLayout()
        sublayout4 = QHBoxLayout()
        sublayout6 = QGridLayout()

        ##간격 유지를 위한 임시 lineedit
        temp_lineedit = QLineEdit(self.dialog5)
        temp_lineedit.setStyleSheet('background-color: #2E2E38;')
        temp_lineedit.setDisabled(True)
        temp_lineedit.setFrame(False)

        layout0 = QGridLayout()
        layout0.addWidget(self.rbtn1, 0, 0)
        layout0.addWidget(self.rbtn2, 0, 1)
        layout0.addWidget(temp_lineedit, 0, 2)

        layout_dc1 = QHBoxLayout()
        layout_dc1.addWidget(labelDC1)
        layout_dc1.addWidget(self.checkC1)
        layout_dc1.addWidget(self.checkD1)

        layout_dc2 = QHBoxLayout()
        layout_dc2.addWidget(labelDC2)
        layout_dc2.addWidget(self.checkC2)
        layout_dc2.addWidget(self.checkD2)

        ### 탭
        tabs = QTabWidget()
        tab1 = QWidget()
        tab2 = QWidget()

        tab1.setLayout(layout1)
        tab2.setLayout(layout2)

        tabs.addTab(tab1, "Non SAP")
        tabs.addTab(tab2, "SAP")

        layout.addLayout(layout0)
        layout.addWidget(tabs)

        ### 배치 - 탭 1 (Non SAP)======================================================
        sublayout5.addWidget(label_ex, 0, 1)
        sublayout5.addWidget(label_AccCode, 1, 0)
        sublayout5.addWidget(self.MyInput, 1, 1)
        sublayout5.addWidget(label_tree2, 2, 0)
        sublayout5.addWidget(self.new_tree2, 2, 1)
        sublayout5.addWidget(labelSheet2, 3, 0)
        sublayout5.addWidget(self.D5_Sheet2, 3, 1)

        layout1.addLayout(sublayout5, stretch=4)
        layout1.addLayout(layout_dc1, stretch=1)
        layout1.addLayout(sublayout2, stretch=1)

        sublayout2.addStretch(2)
        sublayout2.addWidget(self.btn2)
        sublayout2.addWidget(self.btnDialog)

        ### 배치 - 탭 2 (SAP)============================================================
        sublayout6.addWidget(label_SKA1, 0, 1)
        sublayout6.addWidget(label_SKA1_text, 1, 0)
        sublayout6.addWidget(self.listbox_drops, 1, 1)
        sublayout6.addWidget(self.gbtn, 2, 1)

        sublayout6.addWidget(label_tree, 3, 0)
        sublayout6.addWidget(self.new_tree, 3, 1)

        sublayout6.addWidget(labelSheet, 4, 0)
        sublayout6.addWidget(self.D5_Sheet, 4, 1)

        layout2.addLayout(sublayout6, stretch=4)
        layout2.addLayout(layout_dc2, stretch=4)
        layout2.addLayout(sublayout4, stretch=1)

        sublayout4.addStretch(2)
        sublayout4.addWidget(self.btn3)
        sublayout4.addWidget(self.btnDialog1)

        # ? 제거
        self.dialog5.setWindowFlags(Qt.WindowCloseButtonHint)

        ### 공통 지정
        self.dialog5.setLayout(layout)
        self.dialog5.setGeometry(300, 300, 700, 570)
        self.dialog5.setWindowTitle('Scenario5')
        self.dialog5.setWindowModality(Qt.NonModal)
        self.dialog5.show()

    def Dialog6(self):
        self.dialoglist.add(6)
        self.dialog6 = QDialog()
        self.dialog6.setStyleSheet('background-color: #2E2E38')
        self.dialog6.setWindowIcon(QIcon("./EY_logo.png"))

        groupbox0 = QGroupBox('')
        groupbox0.setStyleSheet('QGroupBox  {border:0;}')
        font_groupbox0 = groupbox0.font()
        font_groupbox0.setBold(True)
        groupbox0.setFont(font_groupbox0)

        groupbox1 = QGroupBox('')
        groupbox1.setStyleSheet('QGroupBox  {border:0;}')
        font_groupbox1 = groupbox1.font()
        font_groupbox1.setBold(True)
        groupbox1.setFont(font_groupbox1)

        groupbox2 = QGroupBox('')
        groupbox2.setStyleSheet('QGroupBox  {border:0;}')
        font_groupbox2 = groupbox2.font()
        font_groupbox2.setBold(True)
        groupbox2.setFont(font_groupbox2)

        groupbox3 = QGroupBox('')
        groupbox3.setStyleSheet('QGroupBox  {border:0;}')
        font_groupbox3 = groupbox3.font()
        font_groupbox3.setBold(True)
        groupbox3.setFont(font_groupbox3)

        groupbox4 = QGroupBox('')
        groupbox4.setStyleSheet('QGroupBox  {border:0;}')
        font_groupbox4 = groupbox4.font()
        font_groupbox4.setBold(True)
        groupbox4.setFont(font_groupbox4)

        groupbox5 = QGroupBox('')
        groupbox5.setStyleSheet('QGroupBox  {border:0;}')
        font_groupbox5 = groupbox5.font()
        font_groupbox5.setBold(True)
        groupbox5.setFont(font_groupbox5)

        cursor = self.cnxn.cursor()
        sql = '''
                 SELECT 											
                        *
                 FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
            '''.format(field=self.selected_project_id)
        accountsname = pd.read_sql(sql, self.cnxn)
        self.new_tree = Form(self)
        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)

            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Checked)

        self.new_tree.get_selected_leaves()  # 초기값 모두 선택 (추가)

        cursor2 = self.cnxn.cursor()
        sql2 = '''
                    SELECT
                        UserName,
                        FullName
                    FROM [{field}_Import_Dim].[dbo].[pbcUser];
            '''.format(field=self.selected_project_id)

        pID = pd.read_sql(sql2, self.cnxn)
        self.new_prep = Preparer(self)
        self.new_prep.prep.clear()

        rc = pID.shape[0]
        real_PID = []

        for i in range(0, rc):
            a = ''
            b = ''
            s = ''
            a = str(pID['UserName'][i])
            b = str(pID['FullName'][i])
            s = a + ' | ' + b
            real_PID.append(s)

        pID['real_PID'] = real_PID
        pID.sort_values(by='real_PID', inplace=True)

        for n, i in enumerate(pID.real_PID.unique()):
            self.new_prep.parent = QTreeWidgetItem(self.new_prep.prep)
            self.new_prep.parent.setText(0, "{}".format(i))
            self.new_prep.parent.setFlags(self.new_prep.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            self.new_prep.parent.setCheckState(0, Qt.Checked)

        self.new_prep.get_selected_leaves()

        self.btn2 = QPushButton('   Extract Data', self.dialog6)
        self.btn2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btn2.clicked.connect(self.Thread6)

        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)
        self.btnDialog = QPushButton("   Close", self.dialog6)
        self.btnDialog.setStyleSheet(
            'color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close6)
        font10 = self.btnDialog.font()
        font10.setBold(True)
        self.btnDialog.setFont(font10)
        self.btn2.resize(110, 30)
        self.btnDialog.resize(110, 30)

        self.rbtn1 = QRadioButton('JE Line             ', self.dialog6)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)
        self.rbtn2 = QRadioButton('JE             ', self.dialog6)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        labelDC = QLabel('차변/대변* : ', self.dialog6)
        labelDC.setStyleSheet("color: white;")
        font1 = labelDC.font()
        font1.setBold(True)
        labelDC.setFont(font1)

        self.checkC = QCheckBox('Credit', self.dialog6)
        self.checkD = QCheckBox('Debit', self.dialog6)
        self.checkC.setStyleSheet("color: white;")
        self.checkD.setStyleSheet("color: white;")
        self.checkC.setChecked(True)
        self.checkD.setChecked(True)

        self.checkB = QCheckBox('T일 이전', self.dialog6)
        self.checkF = QCheckBox('T일 이후', self.dialog6)
        self.checkB.setStyleSheet("color: white;")
        self.checkF.setStyleSheet("color: white;")
        self.checkB.setChecked(True)
        self.checkF.setChecked(True)

        labelDate = QLabel('결산일* : ', self.dialog6)
        labelDate.setStyleSheet("color: white;")

        font1 = labelDate.font()
        font1.setBold(True)
        labelDate.setFont(font1)

        self.D6_Date = QLineEdit(self.dialog6)
        self.D6_Date.setReadOnly(True)
        self.D6_Date.setStyleSheet("background-color: white;")
        self.D6_Date.setPlaceholderText('날짜를 선택하세요')

        self.btnDate = QPushButton("Date", self.dialog6)
        self.btnDate.resize(65, 22)
        self.new_calendar = Calendar(self)
        self.new_calendar.calendar.clicked.connect(self.handle_date_clicked)
        self.new_calendar.closebtn.clicked.connect(self.closeCalendar6)
        self.btnDate.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDate.clicked.connect(self.calendar6)
        font11 = self.btnDate.font()
        font11.setBold(True)
        self.btnDate.setFont(font11)

        self.btnDelete = QPushButton("Delete", self.dialog6)
        self.btnDelete.resize(65, 22)
        self.btnDelete.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDelete.clicked.connect(self.delete_date6)
        font12 = self.btnDelete.font()
        font12.setBold(True)
        self.btnDelete.setFont(font12)

        labelDate2 = QLabel('T일 : ', self.dialog6)
        labelDate2.setStyleSheet("color: white;")

        font2 = labelDate2.font()
        font2.setBold(True)
        labelDate2.setFont(font2)

        self.D6_Date2 = QLineEdit(self.dialog6)
        self.D6_Date2.setStyleSheet("background-color: white;")
        self.D6_Date2.setPlaceholderText('T 값을 입력하세요')

        label_tree = QLabel('특정 계정명 : ', self.dialog6)
        label_tree.setStyleSheet("color: white;")
        font4 = label_tree.font()
        font4.setBold(True)
        label_tree.setFont(font4)

        labelJE = QLabel('전표입력자 : ', self.dialog6)
        labelJE.setStyleSheet("color: white;")

        font4 = labelJE.font()
        font4.setBold(True)
        labelJE.setFont(font4)

        labelCost = QLabel('중요성금액 : ', self.dialog6)
        labelCost.setStyleSheet("color: white;")

        font5 = labelCost.font()
        font5.setBold(True)
        labelCost.setFont(font5)

        self.D6_Cost = QLineEdit(self.dialog6)
        self.D6_Cost.setStyleSheet("background-color: white;")
        self.D6_Cost.setPlaceholderText('중요성 금액을 입력하세요')

        labelSheet = QLabel('시나리오 번호* : ', self.dialog6)
        labelSheet.setStyleSheet("color: white;")

        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)

        self.D6_Sheet = QLineEdit(self.dialog6)
        self.D6_Sheet.setStyleSheet("background-color: white;")
        self.D6_Sheet.setPlaceholderText('※ 입력 예시 : F01')

        temp_lineedit = QLineEdit(self.dialog6)
        temp_lineedit.setStyleSheet('background-color: #2E2E38;')
        temp_lineedit.setDisabled(True)
        temp_lineedit.setFrame(False)

        self.D6_Date.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소
        self.D6_Date2.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소
        self.D6_Cost.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소
        self.D6_Sheet.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소

        layout_dc = QHBoxLayout()
        layout_dc.addWidget(labelDC)
        layout_dc.addWidget(self.checkC)
        layout_dc.addWidget(self.checkD)
        groupbox2.setLayout(layout_dc)

        layout0 = QGridLayout()
        layout0.addWidget(self.rbtn1, 0, 0)
        layout0.addWidget(self.rbtn2, 0, 1)
        layout0.addWidget(temp_lineedit, 0, 2)
        groupbox0.setLayout(layout0)

        layout1 = QGridLayout()
        layout1.addWidget(labelDate, 0, 0)
        layout1.addWidget(self.D6_Date, 0, 1)
        layout1.addWidget(self.btnDate, 0, 2)
        layout1.addWidget(self.btnDelete, 0, 3)
        layout1.addWidget(labelDate2, 1, 0)
        layout1.addWidget(self.D6_Date2, 1, 1)
        layout1.addWidget(self.checkB, 1, 2)
        layout1.addWidget(self.checkF, 1, 3)
        layout1.addWidget(label_tree, 2, 0)
        layout1.addWidget(self.new_tree, 2, 1)
        layout1.addWidget(labelJE, 3, 0)
        layout1.addWidget(self.new_prep, 3, 1)
        layout1.addWidget(labelCost, 4, 0)
        layout1.addWidget(self.D6_Cost, 4, 1)
        layout1.addWidget(labelSheet, 5, 0)
        layout1.addWidget(self.D6_Sheet, 5, 1)
        groupbox1.setLayout(layout1)

        layout_btn = QHBoxLayout()
        layout_btn.addStretch()
        layout_btn.addStretch()
        layout_btn.addWidget(self.btn2)
        layout_btn.addWidget(self.btnDialog)
        layout_btn.setContentsMargins(-1, 10, -1, -1)
        groupbox3.setLayout(layout_btn)

        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop)
        main_layout.addWidget(groupbox0)
        main_layout.addWidget(groupbox1)
        main_layout.addWidget(groupbox2)
        main_layout.addWidget(groupbox3)

        self.dialog6.setLayout(main_layout)
        self.dialog6.setGeometry(300, 300, 750, 500)

        self.dialog6.setWindowFlags(Qt.WindowCloseButtonHint)

        self.dialog6.setWindowTitle("Scenario6")
        self.dialog6.setWindowModality(Qt.NonModal)
        self.dialog6.show()

    def Dialog7(self):
        self.dialoglist.add(7)
        self.dialog7 = QDialog()
        self.dialog7.setStyleSheet('background-color: #2E2E38')
        self.dialog7.setWindowIcon(QIcon("./EY_logo.png"))

        groupbox1 = QGroupBox('')
        groupbox1.setStyleSheet('QGroupBox  {border:0;}')
        font_groupbox1 = groupbox1.font()
        font_groupbox1.setBold(True)
        groupbox1.setFont(font_groupbox1)

        groupbox2 = QGroupBox('')
        groupbox2.setStyleSheet('QGroupBox  {border:0;}')
        font_groupbox2 = groupbox2.font()
        font_groupbox2.setBold(True)
        groupbox2.setFont(font_groupbox2)

        groupbox3 = QGroupBox('')
        groupbox3.setStyleSheet('QGroupBox  {border:0;}')
        font_groupbox3 = groupbox3.font()
        font_groupbox3.setBold(True)
        groupbox3.setFont(font_groupbox3)

        groupbox4 = QGroupBox('')
        groupbox4.setStyleSheet('QGroupBox  {border:0;}')
        font_groupbox4 = groupbox4.font()
        font_groupbox4.setBold(True)
        groupbox4.setFont(font_groupbox4)

        groupbox5 = QGroupBox('')
        groupbox5.setStyleSheet('QGroupBox  {border:0;}')
        font_groupbox5 = groupbox5.font()
        font_groupbox5.setBold(True)
        groupbox5.setFont(font_groupbox5)

        cursor = self.cnxn.cursor()

        sql = '''
                 SELECT 											
                        *
                 FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
            '''.format(field=self.selected_project_id)

        accountsname = pd.read_sql(sql, self.cnxn)

        self.new_tree = Form(self)

        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)

            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Checked)
        self.new_tree.get_selected_leaves()  # 초기값 모두 선택 (추가)

        cursor2 = self.cnxn.cursor()
        sql2 = '''
                    SELECT
                        UserName,
                        FullName
                    FROM [{field}_Import_Dim].[dbo].[pbcUser]
            '''.format(field=self.selected_project_id)

        pID = pd.read_sql(sql2, self.cnxn)
        self.new_prep = Preparer(self)
        self.new_prep.prep.clear()

        rc = pID.shape[0]
        real_PID = []

        for i in range(0, rc):
            a = ''
            b = ''
            s = ''
            a = str(pID['UserName'][i])
            b = str(pID['FullName'][i])
            s = a + ' | ' + b
            real_PID.append(s)

        pID['real_PID'] = real_PID
        pID.sort_values(by='real_PID', inplace=True)

        for n, i in enumerate(pID.real_PID.unique()):
            self.new_prep.parent = QTreeWidgetItem(self.new_prep.prep)
            self.new_prep.parent.setText(0, "{}".format(i))
            self.new_prep.parent.setFlags(self.new_prep.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            self.new_prep.parent.setCheckState(0, Qt.Checked)

        self.new_prep.get_selected_leaves()

        self.checkC = QCheckBox('Credit', self.dialog7)
        self.checkD = QCheckBox('Debit', self.dialog7)
        self.checkC.setStyleSheet("color: white;")
        self.checkD.setStyleSheet("color: white;")
        self.checkC.setChecked(True)
        self.checkD.setChecked(True)

        labelDC = QLabel('차변/대변* : ', self.dialog7)
        labelDC.setStyleSheet("color: white;")
        font1 = labelDC.font()
        font1.setBold(True)
        labelDC.setFont(font1)

        self.btn2 = QPushButton('   Extract Data', self.dialog7)
        self.btn2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btn2.clicked.connect(self.Thread7)

        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)

        self.btnDialog = QPushButton("   Close", self.dialog7)
        self.btnDialog.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close7)

        font10 = self.btnDialog.font()
        font10.setBold(True)
        self.btnDialog.setFont(font10)

        self.btn2.resize(110, 30)
        self.btnDialog.resize(110, 30)

        self.rbtn3 = QRadioButton('JE Line             ', self.dialog7)
        self.rbtn3.setStyleSheet("color: white;")
        font11 = self.rbtn3.font()
        font11.setBold(True)
        self.rbtn3.setFont(font11)

        self.rbtn4 = QRadioButton('JE             ', self.dialog7)
        self.rbtn4.setStyleSheet("color: white;")
        font12 = self.rbtn4.font()
        font12.setBold(True)
        self.rbtn4.setFont(font12)

        self.rbtn1 = QRadioButton('Effective Date ', self.dialog7)
        self.rbtn1.setStyleSheet("color: white;")
        font1 = self.rbtn1.font()
        font1.setBold(True)
        self.rbtn1.setFont(font1)

        self.rbtn2 = QRadioButton('Entry Date             ', self.dialog7)
        self.rbtn2.setStyleSheet("color: white;")
        font2 = self.rbtn2.font()
        font2.setBold(True)
        self.rbtn2.setFont(font2)

        labelDate = QLabel('비영업일 : ', self.dialog7)
        labelDate.setStyleSheet("color: white;")

        font3 = labelDate.font()
        font3.setBold(True)
        labelDate.setFont(font3)

        self.D7_Date = QTextEdit(self.dialog7)
        self.D7_Date.setReadOnly(True)
        self.D7_Date.setStyleSheet("background-color: white;")
        self.D7_Date.setPlaceholderText('날짜를 추가해주세요\n법정 공휴일 및 주말은 포함되어 있습니다')

        self.btnDate = QPushButton("Add Date", self.dialog7)
        self.btnDate.resize(65, 22)
        self.new_calendar = Calendar(self)
        self.new_calendar.closebtn.clicked.connect(self.closeCalendar7)
        self.new_calendar.calendar.clicked.connect(self.handle_date_clicked2)
        self.btnDate.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDate.clicked.connect(self.calendar7)

        self.btnDelete = QPushButton("Delete All", self.dialog7)
        self.btnDelete.resize(65, 22)
        self.btnDelete.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDelete.clicked.connect(self.delete_date7)

        font11 = self.btnDate.font()
        font11.setBold(True)
        self.btnDate.setFont(font11)

        font12 = self.btnDelete.font()
        font12.setBold(True)
        self.btnDelete.setFont(font12)

        label_tree = QLabel('특정 계정명 : ', self.dialog7)
        label_tree.setStyleSheet("color: white;")
        font4 = label_tree.font()
        font4.setBold(True)
        label_tree.setFont(font4)

        labelJE = QLabel('전표입력자 : ', self.dialog7)
        labelJE.setStyleSheet("color: white;")

        font5 = labelJE.font()
        font5.setBold(True)
        labelJE.setFont(font5)

        labelCost = QLabel('중요성금액 : ', self.dialog7)
        labelCost.setStyleSheet("color: white;")

        font6 = labelCost.font()
        font6.setBold(True)
        labelCost.setFont(font6)

        self.D7_Cost = QLineEdit(self.dialog7)
        self.D7_Cost.setStyleSheet("background-color: white;")
        self.D7_Cost.setPlaceholderText('중요성 금액을 입력하세요')

        labelSheet = QLabel('시나리오 번호* : ', self.dialog7)
        labelSheet.setStyleSheet("color: white;")

        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)

        self.D7_Sheet = QLineEdit(self.dialog7)
        self.D7_Sheet.setStyleSheet("background-color: white;")
        self.D7_Sheet.setPlaceholderText('※ 입력 예시 : F01')

        self.D7_Date.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소
        self.D7_Cost.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소
        self.D7_Sheet.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소

        temp_lineedit = QLineEdit(self.dialog7)
        temp_lineedit.setStyleSheet('background-color: #2E2E38;')
        temp_lineedit.setDisabled(True)
        temp_lineedit.setFrame(False)

        temp_lineedit2 = QLineEdit(self.dialog7)
        temp_lineedit2.setStyleSheet('background-color: #2E2E38;')
        temp_lineedit2.setDisabled(True)
        temp_lineedit2.setFrame(False)

        layoutr = QGridLayout()
        self.rbtn1.setChecked(True)
        layoutr.addWidget(self.rbtn1, 0, 0)
        layoutr.addWidget(self.rbtn2, 0, 1)
        layoutr.addWidget(temp_lineedit, 0, 2)
        groupbox1.setLayout(layoutr)

        layout1 = QGridLayout()
        layout1.addWidget(labelDate, 0, 0)
        layout1.addWidget(self.D7_Date, 0, 1)
        layout1.addWidget(self.btnDate, 0, 2)
        layout1.addWidget(self.btnDelete, 0, 3)
        layout1.addWidget(label_tree, 1, 0)
        layout1.addWidget(self.new_tree, 1, 1)
        layout1.addWidget(labelJE, 2, 0)
        layout1.addWidget(self.new_prep, 2, 1)
        layout1.addWidget(labelCost, 3, 0)
        layout1.addWidget(self.D7_Cost, 3, 1)
        layout1.addWidget(labelSheet, 4, 0)
        layout1.addWidget(self.D7_Sheet, 4, 1)
        groupbox3.setLayout(layout1)

        layout2 = QHBoxLayout()
        layout2.addStretch()
        layout2.addStretch()
        layout2.addWidget(self.btn2)
        layout2.addWidget(self.btnDialog)
        layout2.setContentsMargins(-1, 10, -1, -1)
        groupbox4.setLayout(layout2)

        layout3 = QGridLayout()
        self.rbtn3.setChecked(True)
        layout3.addWidget(self.rbtn3, 0, 0)
        layout3.addWidget(self.rbtn4, 0, 1)
        layout3.addWidget(temp_lineedit2, 0, 2)
        groupbox2.setLayout(layout3)

        layout_dc = QHBoxLayout()
        layout_dc.addWidget(labelDC)
        layout_dc.addWidget(self.checkC)
        layout_dc.addWidget(self.checkD)
        groupbox5.setLayout(layout_dc)

        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop)
        main_layout.addWidget(groupbox2)
        main_layout.addWidget(groupbox1)
        main_layout.addWidget(groupbox3)
        main_layout.addWidget(groupbox5)
        main_layout.addWidget(groupbox4)

        self.dialog7.setLayout(main_layout)
        self.dialog7.setGeometry(300, 300, 750, 500)

        self.dialog7.setWindowFlags(Qt.WindowCloseButtonHint)  # ? 제거

        self.dialog7.setWindowTitle("Scenario7")
        self.dialog7.setWindowModality(Qt.NonModal)
        self.dialog7.show()

    def Dialog8(self):
        self.dialoglist.add(8)
        self.dialog8 = QDialog()
        self.dialog8.setStyleSheet('background-color: #2E2E38')
        self.dialog8.setWindowIcon(QIcon("./EY_logo.png"))

        groupbox0 = QGroupBox('')
        groupbox0.setStyleSheet('QGroupBox  {border:0;}')
        font_groupbox0 = groupbox0.font()
        font_groupbox0.setBold(True)
        groupbox0.setFont(font_groupbox0)

        groupbox1 = QGroupBox('')
        groupbox1.setStyleSheet('QGroupBox  {border:0;}')
        font_groupbox1 = groupbox1.font()
        font_groupbox1.setBold(True)
        groupbox1.setFont(font_groupbox1)

        groupbox2 = QGroupBox('')
        groupbox2.setStyleSheet('QGroupBox  {border:0;}')
        font_groupbox2 = groupbox2.font()
        font_groupbox2.setBold(True)
        groupbox2.setFont(font_groupbox2)

        groupbox3 = QGroupBox('')
        groupbox3.setStyleSheet('QGroupBox  {border:0;}')
        font_groupbox3 = groupbox3.font()
        font_groupbox3.setBold(True)
        groupbox3.setFont(font_groupbox3)

        cursor = self.cnxn.cursor()

        sql = '''
                 SELECT 											
                        *
                 FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
            '''.format(field=self.selected_project_id)

        accountsname = pd.read_sql(sql, self.cnxn)
        self.new_tree = Form(self)
        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)

            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Checked)
        self.new_tree.get_selected_leaves()  # 초기값 모두 선택 (추가)

        cursor2 = self.cnxn.cursor()
        sql2 = '''
                    SELECT
                        UserName,
                        FullName
                    FROM [{field}_Import_Dim].[dbo].[pbcUser]
            '''.format(field=self.selected_project_id)

        pID = pd.read_sql(sql2, self.cnxn)
        self.new_prep = Preparer(self)
        self.new_prep.prep.clear()

        rc = pID.shape[0]
        real_PID = []

        for i in range(0, rc):
            a = ''
            b = ''
            s = ''
            a = str(pID['UserName'][i])
            b = str(pID['FullName'][i])
            s = a + ' | ' + b
            real_PID.append(s)

        pID['real_PID'] = real_PID
        pID.sort_values(by='real_PID', inplace=True)

        for n, i in enumerate(pID.real_PID.unique()):
            self.new_prep.parent = QTreeWidgetItem(self.new_prep.prep)
            self.new_prep.parent.setText(0, "{}".format(i))
            self.new_prep.parent.setFlags(self.new_prep.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            self.new_prep.parent.setCheckState(0, Qt.Checked)

        self.new_prep.get_selected_leaves()

        self.checkC = QCheckBox('Credit', self.dialog8)
        self.checkD = QCheckBox('Debit', self.dialog8)
        self.checkC.setStyleSheet("color: white;")
        self.checkD.setStyleSheet("color: white;")
        self.checkC.setChecked(True)
        self.checkD.setChecked(True)

        labelDC = QLabel('차변/대변* : ', self.dialog8)
        labelDC.setStyleSheet("color: white;")
        font1 = labelDC.font()
        font1.setBold(True)
        labelDC.setFont(font1)

        self.btn2 = QPushButton('   Extract Data', self.dialog8)
        self.btn2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btn2.clicked.connect(self.Thread8)

        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)

        self.btnDialog = QPushButton("   Close", self.dialog8)
        self.btnDialog.setStyleSheet(
            'color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close8)

        font10 = self.btnDialog.font()
        font10.setBold(True)
        self.btnDialog.setFont(font10)

        self.btn2.resize(110, 30)
        self.btnDialog.resize(110, 30)

        self.rbtn1 = QRadioButton('JE Line             ', self.dialog8)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)
        self.rbtn2 = QRadioButton('JE             ', self.dialog8)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        labelDate = QLabel('N일* : ', self.dialog8)
        labelDate.setStyleSheet("color: white;")

        font1 = labelDate.font()
        font1.setBold(True)
        labelDate.setFont(font1)

        self.D8_N = QLineEdit(self.dialog8)
        self.D8_N.setStyleSheet("background-color: white;")
        self.D8_N.setPlaceholderText('N 값을 입력하세요')

        label_tree = QLabel('특정 계정명 : ', self.dialog8)
        label_tree.setStyleSheet("color: white;")
        font4 = label_tree.font()
        font4.setBold(True)
        label_tree.setFont(font4)

        labelJE = QLabel('전표입력자 : ', self.dialog8)
        labelJE.setStyleSheet("color: white;")

        font3 = labelJE.font()
        font3.setBold(True)
        labelJE.setFont(font3)

        labelCost = QLabel('중요성금액 : ', self.dialog8)
        labelCost.setStyleSheet("color: white;")

        font4 = labelCost.font()
        font4.setBold(True)
        labelCost.setFont(font4)

        self.D8_Cost = QLineEdit(self.dialog8)
        self.D8_Cost.setStyleSheet("background-color: white;")
        self.D8_Cost.setPlaceholderText('중요성 금액을 입력하세요')

        labelSheet = QLabel('시나리오 번호* : ', self.dialog8)
        labelSheet.setStyleSheet("color: white;")

        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)

        self.D8_Sheet = QLineEdit(self.dialog8)
        self.D8_Sheet.setStyleSheet("background-color: white;")
        self.D8_Sheet.setPlaceholderText('※ 입력 예시 : F01')

        self.D8_N.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소
        self.D8_Cost.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소
        self.D8_Sheet.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소

        temp_lineedit = QLineEdit(self.dialog8)
        temp_lineedit.setStyleSheet('background-color: #2E2E38;')
        temp_lineedit.setDisabled(True)
        temp_lineedit.setFrame(False)

        layout_dc = QHBoxLayout()
        layout_dc.addWidget(labelDC)
        layout_dc.addWidget(self.checkC)
        layout_dc.addWidget(self.checkD)
        groupbox2.setLayout(layout_dc)

        layout0 = QGridLayout()
        layout0.addWidget(self.rbtn1, 0, 0)
        layout0.addWidget(self.rbtn2, 0, 1)
        layout0.addWidget(temp_lineedit, 0, 2)
        groupbox0.setLayout(layout0)

        layout1 = QGridLayout()
        layout1.addWidget(labelDate, 0, 0)
        layout1.addWidget(self.D8_N, 0, 1)
        layout1.addWidget(label_tree, 1, 0)
        layout1.addWidget(self.new_tree, 1, 1)
        layout1.addWidget(labelJE, 2, 0)
        layout1.addWidget(self.new_prep, 2, 1)
        layout1.addWidget(labelCost, 3, 0)
        layout1.addWidget(self.D8_Cost, 3, 1)
        layout1.addWidget(labelSheet, 4, 0)
        layout1.addWidget(self.D8_Sheet, 4, 1)
        groupbox1.setLayout(layout1)

        layout2 = QHBoxLayout()
        layout2.addStretch()
        layout2.addStretch()
        layout2.addWidget(self.btn2)
        layout2.addWidget(self.btnDialog)
        layout2.setContentsMargins(-1, 10, -1, -1)
        groupbox3.setLayout(layout2)

        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop)
        main_layout.addWidget(groupbox0)
        main_layout.addWidget(groupbox1)
        main_layout.addWidget(groupbox2)
        main_layout.addWidget(groupbox3)

        self.dialog8.setLayout(main_layout)
        self.dialog8.setGeometry(300, 300, 750, 500)

        self.dialog8.setWindowFlags(Qt.WindowCloseButtonHint)

        self.dialog8.setWindowTitle("Scenario8")
        self.dialog8.setWindowModality(Qt.NonModal)
        self.dialog8.show()

    def Dialog9(self):
        self.dialoglist.add(9)
        self.dialog9 = QDialog()
        groupbox = QGroupBox('접속 정보')
        cursor = self.cnxn.cursor()

        sql = '''
                 SELECT 											
                        *
                 FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
            '''.format(field=self.selected_project_id)

        accountsname = pd.read_sql(sql, self.cnxn)

        self.new_tree = Form(self)

        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)

            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Checked)
        self.new_tree.get_selected_leaves()  # 초기값 모두 선택 (추가)

        self.dialog9.setStyleSheet('background-color: #2E2E38')
        self.dialog9.setWindowIcon(QIcon("./EY_logo.png"))

        self.btn2 = QPushButton('   Extract Data', self.dialog9)
        self.btn2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btn2.clicked.connect(self.Thread9)

        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)

        self.btnDialog = QPushButton("  Close", self.dialog9)
        self.btnDialog.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close9)

        font10 = self.btnDialog.font()
        font10.setBold(True)
        self.btnDialog.setFont(font10)

        # JE Line / JE 라디오 버튼
        self.rbtn1 = QRadioButton('JE Line', self.dialog9)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)
        self.rbtn2 = QRadioButton('JE', self.dialog9)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        labelKeyword = QLabel('작성빈도수* : ', self.dialog9)
        labelKeyword.setStyleSheet("color: white;")

        font1 = labelKeyword.font()
        font1.setBold(True)
        labelKeyword.setFont(font1)

        self.D9_N = QLineEdit(self.dialog9)
        self.D9_N.setStyleSheet("background-color: white;")
        self.D9_N.setPlaceholderText('작성 빈도수를 입력하세요')

        labelTE = QLabel('중요성 금액 : ', self.dialog9)
        labelTE.setStyleSheet("color: white;")

        font2 = labelTE.font()
        font2.setBold(True)
        labelTE.setFont(font2)

        self.D9_TE = QLineEdit(self.dialog9)
        self.D9_TE.setStyleSheet("background-color: white;")
        self.D9_TE.setPlaceholderText('중요성 금액을 입력하세요')

        self.btn2.resize(110, 30)
        self.btnDialog.resize(110, 30)

        label_tree = QLabel('특정 계정명 : ', self.dialog9)
        label_tree.setStyleSheet("color: white;")
        font4 = label_tree.font()
        font4.setBold(True)
        label_tree.setFont(font4)

        labelSheet = QLabel('시나리오 번호* : ', self.dialog9)
        labelSheet.setStyleSheet("color: white;")

        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)

        self.D9_Sheet = QLineEdit(self.dialog9)
        self.D9_Sheet.setStyleSheet("background-color: white;")
        self.D9_Sheet.setPlaceholderText('※ 입력 예시 : F01')

        labelDC = QLabel('차변/대변* : ', self.dialog9)
        labelDC.setStyleSheet("color: white;")
        font1 = labelDC.font()
        font1.setBold(True)
        labelDC.setFont(font1)

        # 차변/대변 체크박스로 구현
        self.checkC = QCheckBox('Credit', self.dialog9)
        self.checkD = QCheckBox('Debit', self.dialog9)
        self.checkC.setStyleSheet("color: white;")
        self.checkD.setStyleSheet("color: white;")
        self.checkC.setChecked(True)
        self.checkD.setChecked(True)

        self.D9_N.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소
        self.D9_TE.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소
        self.D9_Sheet.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소

        layout1 = QGridLayout()
        layout1.addWidget(self.rbtn1, 0, 0)
        layout1.addWidget(self.rbtn2, 0, 1)
        layout1.addWidget(labelKeyword, 1, 0)
        layout1.addWidget(self.D9_N, 1, 1)
        layout1.addWidget(labelTE, 2, 0)
        layout1.addWidget(self.D9_TE, 2, 1)
        layout1.addWidget(label_tree, 3, 0)
        layout1.addWidget(self.new_tree, 3, 1)
        layout1.addWidget(labelSheet, 4, 0)
        layout1.addWidget(self.D9_Sheet, 4, 1)

        layout_dc = QHBoxLayout()
        layout_dc.addWidget(labelDC)
        layout_dc.addWidget(self.checkC)
        layout_dc.addWidget(self.checkD)

        layout2 = QHBoxLayout()
        layout2.addStretch()
        layout2.addStretch()
        layout2.addWidget(self.btn2)
        layout2.addWidget(self.btnDialog)

        layout2.setContentsMargins(-1, 10, -1, -1)

        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop)
        main_layout.addLayout(layout1)
        main_layout.addLayout(layout_dc)
        main_layout.addLayout(layout2)

        self.dialog9.setLayout(main_layout)
        self.dialog9.setGeometry(300, 300, 600, 300)

        # ? 제거
        self.dialog9.setWindowFlags(Qt.WindowCloseButtonHint)

        self.dialog9.setWindowTitle("Scenario9")
        self.dialog9.setWindowModality(Qt.NonModal)
        self.dialog9.show()

    def Dialog10(self):
        self.dialoglist.add(10)
        self.dialog10 = QDialog()
        self.dialog10.setStyleSheet('background-color: #2E2E38')
        self.dialog10.setWindowIcon(QIcon("./EY_logo.png"))

        cursor = self.cnxn.cursor()

        sql = '''
                 SELECT 											
                        *
                 FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
            '''.format(field=self.selected_project_id)

        accountsname = pd.read_sql(sql, self.cnxn)

        self.new_tree = Form(self)

        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)

            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Checked)
        self.new_tree.get_selected_leaves()  # 초기값 모두 선택 (추가)

        # 전표입력자 체크리스트
        cursor2 = self.cnxn.cursor()
        sql2 = '''
                    SELECT
                        UserName,
                        FullName
                    FROM [{field}_Import_Dim].[dbo].[pbcUser]
            '''.format(field=self.selected_project_id)

        pID = pd.read_sql(sql2, self.cnxn)
        self.new_prep = Preparer(self)
        self.new_prep.prep.clear()

        rc = pID.shape[0]
        real_PID = []

        for i in range(0, rc):
            a = ''
            b = ''
            s = ''
            a = str(pID['UserName'][i])
            b = str(pID['FullName'][i])
            s = a + ' | ' + b
            real_PID.append(s)

        pID['real_PID'] = real_PID
        pID.sort_values(by='real_PID', inplace=True)

        for n, i in enumerate(pID.real_PID.unique()):
            self.new_prep.parent = QTreeWidgetItem(self.new_prep.prep)
            self.new_prep.parent.setText(0, "{}".format(i))
            self.new_prep.parent.setFlags(self.new_prep.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            self.new_prep.parent.setCheckState(0, Qt.Checked)

        self.new_prep.get_selected_leaves()  # 초기값 모두 선택 (추가)

        labelDC = QLabel('차변/대변* : ', self.dialog10)
        labelDC.setStyleSheet("color: white;")
        font1 = labelDC.font()
        font1.setBold(True)
        labelDC.setFont(font1)

        # 차변/대변 체크박스로 구현
        self.checkC = QCheckBox('Credit', self.dialog10)
        self.checkD = QCheckBox('Debit', self.dialog10)
        self.checkC.setStyleSheet("color: white;")
        self.checkD.setStyleSheet("color: white;")
        self.checkC.setChecked(True)
        self.checkD.setChecked(True)

        self.btn2 = QPushButton('   Extract Data', self.dialog10)
        self.btn2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btn2.clicked.connect(self.Thread10)

        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)

        self.btnDialog = QPushButton("   Close", self.dialog10)
        self.btnDialog.setStyleSheet(
            'color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close10)

        font10 = self.btnDialog.font()
        font10.setBold(True)
        self.btnDialog.setFont(font10)

        self.btn2.resize(110, 30)
        self.btnDialog.resize(110, 30)

        # JE Line / JE 라디오 버튼
        self.rbtn1 = QRadioButton('JE Line', self.dialog10)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)
        self.rbtn2 = QRadioButton('JE', self.dialog10)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        labelKeyword = QLabel('전표입력자* : ', self.dialog10)
        labelKeyword.setStyleSheet("color: white;")

        font1 = labelKeyword.font()
        font1.setBold(True)
        labelKeyword.setFont(font1)

        ## 캘린더 버튼 추가
        self.btnDate1 = QPushButton("Date", self.dialog10)
        self.btnDate1.resize(65, 22)
        self.new_calendar1 = Calendar(self)
        self.new_calendar1.closebtn.clicked.connect(self.closeCalendar10_1)
        self.new_calendar1.calendar.clicked.connect(self.handle_date_clicked3)
        self.btnDate1.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDate1.clicked.connect(self.calendar10_1)
        font11 = self.btnDate1.font()
        font11.setBold(True)
        self.btnDate1.setFont(font11)

        self.btnDate2 = QPushButton("Date", self.dialog10)
        self.btnDate2.resize(65, 22)
        self.new_calendar2 = Calendar(self)
        self.new_calendar2.closebtn.clicked.connect(self.closeCalendar10_2)
        self.new_calendar2.calendar.clicked.connect(self.handle_date_clicked4)
        self.btnDate2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDate2.clicked.connect(self.calendar10_2)
        font11 = self.btnDate1.font()
        font11.setBold(True)
        self.btnDate2.setFont(font11)

        labelPoint1 = QLabel('시작시점 : ', self.dialog10)
        labelPoint2 = QLabel('종료시점 : ', self.dialog10)
        labelPoint1.setStyleSheet("color: white;")
        labelPoint2.setStyleSheet("color: white;")

        font2_1 = labelPoint1.font()
        font2_1.setBold(True)
        labelPoint1.setFont(font2_1)

        font2_2 = labelPoint2.font()
        font2_2.setBold(True)
        labelPoint2.setFont(font2_2)

        self.D10_Point1 = QLineEdit(self.dialog10)
        self.D10_Point1.setReadOnly(True)
        self.D10_Point1.setStyleSheet("background-color: white;")
        self.D10_Point1.setPlaceholderText('시작시점을 선택하세요')

        self.D10_Point2 = QLineEdit(self.dialog10)
        self.D10_Point2.setReadOnly(True)
        self.D10_Point2.setStyleSheet("background-color: white;")
        self.D10_Point2.setPlaceholderText('종료시점을 선택하세요')

        label_tree = QLabel('특정 계정명 : ', self.dialog10)
        label_tree.setStyleSheet("color: white;")
        font4 = label_tree.font()
        font4.setBold(True)
        label_tree.setFont(font4)

        labelTE = QLabel('중요성 금액 : ', self.dialog10)
        labelTE.setStyleSheet("color: white;")

        font4 = labelTE.font()
        font4.setBold(True)
        labelTE.setFont(font4)

        self.D10_TE = QLineEdit(self.dialog10)
        self.D10_TE.setStyleSheet("background-color: white;")
        self.D10_TE.setPlaceholderText('중요성 금액을 입력하세요')

        labelSheet = QLabel('시나리오 번호* : ', self.dialog10)
        labelSheet.setStyleSheet("color: white;")

        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)

        self.D10_Sheet = QLineEdit(self.dialog10)
        self.D10_Sheet.setStyleSheet("background-color: white;")
        self.D10_Sheet.setPlaceholderText('※ 입력 예시 : F01')

        self.btnDelete1 = QPushButton("Delete", self.dialog10)
        self.btnDelete1.resize(65, 22)
        self.btnDelete1.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDelete1.clicked.connect(self.delete_date101)
        font12 = self.btnDelete1.font()
        font12.setBold(True)
        self.btnDelete1.setFont(font12)

        self.btnDelete2 = QPushButton("Delete", self.dialog10)
        self.btnDelete2.resize(65, 22)
        self.btnDelete2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDelete2.clicked.connect(self.delete_date102)
        font12 = self.btnDelete2.font()
        font12.setBold(True)
        self.btnDelete2.setFont(font12)

        self.D10_Point1.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소
        self.D10_Point2.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소
        self.D10_TE.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소
        self.D10_Sheet.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소

        layout_dc = QHBoxLayout()
        layout_dc.addWidget(labelDC)
        layout_dc.addWidget(self.checkC)
        layout_dc.addWidget(self.checkD)

        layout1 = QGridLayout()
        layout1.addWidget(self.rbtn1, 0, 0)
        layout1.addWidget(self.rbtn2, 0, 1)
        layout1.addWidget(labelKeyword, 1, 0)
        layout1.addWidget(self.new_prep, 1, 1)
        layout1.addWidget(labelPoint1, 2, 0)
        layout1.addWidget(self.D10_Point1, 2, 1)
        layout1.addWidget(self.btnDate1, 2, 2)
        layout1.addWidget(self.btnDelete1, 2, 3)
        layout1.addWidget(labelPoint2, 3, 0)
        layout1.addWidget(self.D10_Point2, 3, 1)
        layout1.addWidget(self.btnDate2, 3, 2)
        layout1.addWidget(self.btnDelete2, 3, 3)
        layout1.addWidget(label_tree, 4, 0)
        layout1.addWidget(self.new_tree, 4, 1)
        layout1.addWidget(labelTE, 5, 0)
        layout1.addWidget(self.D10_TE, 5, 1)
        layout1.addWidget(labelSheet, 6, 0)
        layout1.addWidget(self.D10_Sheet, 6, 1)

        layout2 = QHBoxLayout()
        layout2.addStretch()
        layout2.addStretch()
        layout2.addWidget(self.btn2)
        layout2.addWidget(self.btnDialog)

        layout2.setContentsMargins(-1, 10, -1, -1)

        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop)
        main_layout.addLayout(layout1)
        main_layout.addLayout(layout_dc)
        main_layout.addLayout(layout2)

        self.dialog10.setLayout(main_layout)
        self.dialog10.setGeometry(300, 300, 700, 300)

        # ? 제거
        self.dialog10.setWindowFlags(Qt.WindowCloseButtonHint)

        self.dialog10.setWindowTitle("Scenario10")
        self.dialog10.setWindowModality(Qt.NonModal)
        self.dialog10.show()

    def delete_date7(self):
        self.string_date_list = []
        self.finalDate = []
        self.D7_Date.setText('')

    def delete_date6(self):
        self.D6_Date.setText('')

    def delete_date101(self):
        self.D10_Point1.setText('')

    def delete_date102(self):
        self.D10_Point2.setText('')

    def closeCalendar6(self):
        self.new_calendar.close()
        self.dialog6.activateWindow()

    def closeCalendar7(self):
        self.new_calendar.close()
        self.dialog7.activateWindow()

    def closeCalendar10_1(self):
        self.new_calendar1.close()
        self.dialog10.activateWindow()

    def closeCalendar10_2(self):
        self.new_calendar2.close()
        self.dialog10.activateWindow()

    def handle_date_clicked(self, date):
        self.D6_Date.setText(date.toString("yyyy-MM-dd"))
        self.dialog6.activateWindow()

    def handle_date_clicked2(self, date):
        self.dialog7.activateWindow()

        self.dateList = []
        self.dateList.append(date)

        self.string_date_list = [date_obj.toString("yyyy-MM-dd") for date_obj in self.dateList]

        for self.string_date in self.string_date_list:
            self.D7_Date.append(self.string_date)
            self.fianlDate.append(self.string_date)

        self.dialog7.activateWindow()

    def handle_date_clicked3(self, date):
        self.D10_Point1.setText(date.toString("yyyy-MM-dd"))
        self.dialog10.activateWindow()

    def handle_date_clicked4(self, date):
        self.D10_Point2.setText(date.toString("yyyy-MM-dd"))
        self.dialog10.activateWindow()

    def Dialog12(self):
        self.dialoglist.add(12)
        self.dialog12 = QDialog()
        self.dialog12.setStyleSheet('background-color: #2E2E38')
        self.dialog12.setWindowIcon(QIcon('./EY_logo.png'))

        # 시나리오12
        cursor = self.cnxn.cursor()
        sql = '''
                         SELECT 											
                                *
                         FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
                    '''.format(field=self.selected_project_id)

        accountsname = pd.read_sql(sql, self.cnxn)
        self.new_tree = Form(self)
        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)

            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Checked)
        self.new_tree.get_selected_leaves()

        self.btn = QPushButton('   Extract Data', self.dialog12)
        self.btn.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btn.clicked.connect(self.Thread12)
        font9 = self.btn.font()
        font9.setBold(True)
        self.btn.setFont(font9)

        self.btnDialog = QPushButton("   Close", self.dialog12)
        self.btnDialog.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close12)
        font10 = self.btnDialog.font()
        font10.setBold(True)
        self.btnDialog.setFont(font10)
        self.btn.resize(110, 30)
        self.btnDialog.resize(110, 30)

        labelAccount = QLabel('특정 계정명/계정 코드* : ', self.dialog12)
        labelAccount.setStyleSheet("color: white;")
        font3 = labelAccount.font()
        font3.setBold(True)
        labelAccount.setFont(font3)

        labelCost = QLabel('중요성 금액 : ', self.dialog12)
        labelCost.setStyleSheet("color: white;")
        font3 = labelCost.font()
        font3.setBold(True)
        labelCost.setFont(font3)

        self.D12_Cost = QLineEdit(self.dialog12)
        self.D12_Cost.setStyleSheet("background-color: white;")
        self.D12_Cost.setPlaceholderText('중요성 금액을 입력하세요')
        self.D12_Cost.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소

        labelSheet12 = QLabel('시나리오 번호* : ', self.dialog12)
        labelSheet12.setStyleSheet("color: white;")
        font5 = labelSheet12.font()
        font5.setBold(True)
        labelSheet12.setFont(font5)
        self.D12_Sheet12 = QLineEdit(self.dialog12)
        self.D12_Sheet12.setStyleSheet("background-color: white;")
        self.D12_Sheet12.setPlaceholderText('※ 입력 예시 : F01')
        self.D12_Sheet12.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소

        self.checkC1 = QCheckBox('Credit', self.dialog12)
        self.checkD1 = QCheckBox('Debit', self.dialog12)
        self.checkC1.setStyleSheet("color: white;")
        self.checkD1.setStyleSheet("color: white;")

        labelDC1 = QLabel('차변/대변* : ', self.dialog12)
        labelDC1.setStyleSheet("color: white;")
        font1 = labelDC1.font()
        font1.setBold(True)
        labelDC1.setFont(font1)

        self.checkC1.setChecked(True)
        self.checkD1.setChecked(True)

        sublayout0 = QHBoxLayout()
        sublayout0.addWidget(labelDC1)
        sublayout0.addWidget(self.checkC1)
        sublayout0.addWidget(self.checkD1)

        temp_lineedit = QLineEdit(self.dialog12)
        temp_lineedit.setDisabled(True)
        temp_lineedit.setFrame(False)
        temp_lineedit.setStyleSheet('background-color: #2E2E38;')

        sublayout1 = QGridLayout()
        sublayout1.addWidget(temp_lineedit, 0, 1)
        sublayout1.addWidget(labelAccount, 1, 0)
        sublayout1.addWidget(self.new_tree, 1, 1)
        sublayout1.addWidget(labelCost, 2, 0)
        sublayout1.addWidget(self.D12_Cost, 2, 1)
        sublayout1.addWidget(labelSheet12, 3, 0)
        sublayout1.addWidget(self.D12_Sheet12, 3, 1)

        sublayout2 = QHBoxLayout()
        sublayout2.addStretch()
        sublayout2.addStretch()
        sublayout2.addWidget(self.btn)
        sublayout2.addWidget(self.btnDialog)

        main_layout1 = QVBoxLayout()
        main_layout1.addLayout(sublayout1)
        main_layout1.addLayout(sublayout0)
        main_layout1.addLayout(sublayout2)

        ### Cursor문
        self.btn2 = QPushButton('   Extract Data', self.dialog12)
        self.btn2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btn2.clicked.connect(self.ThreadC)
        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)

        self.btnDialog2 = QPushButton("   Close", self.dialog12)
        self.btnDialog2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDialog2.clicked.connect(self.dialog_close12)
        font10 = self.btnDialog2.font()
        font10.setBold(True)
        self.btnDialog2.setFont(font10)
        self.btn2.resize(110, 30)
        self.btnDialog2.resize(110, 30)

        # JE Line / JE 라디오 버튼
        self.rbtn1 = QRadioButton('JE Line', self.dialog12)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)
        self.rbtn2 = QRadioButton('JE', self.dialog12)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        # 입력된 Cursor문
        labelCursortext = QLabel('입력된 Cursor : ', self.dialog12)
        labelCursortext.setStyleSheet("color: white;")
        font17 = labelCursortext.font()
        font17.setBold(True)
        labelCursortext.setFont(font17)
        self.Cursortext = QTextEdit(self.dialog12)
        self.Cursortext.setPlaceholderText('추출된 Cursor 조건이 표시됩니다')
        self.Cursortext.setReadOnly(True)
        self.Cursortext.setStyleSheet("background-color: white;")

        labelCursor = QLabel('Cursor 파일 위치* : ', self.dialog12)
        labelCursor.setStyleSheet("color: white;")
        font3 = labelCursor.font()
        font3.setBold(True)
        labelCursor.setFont(font3)

        listCursor = QLabel('Cursor Sheet 위치* : ', self.dialog12)
        listCursor.setStyleSheet("color: white;")
        font13 = listCursor.font()
        font13.setBold(True)
        listCursor.setFont(font13)
        self.listCursor = QComboBox(self)
        self.listCursor.setStyleSheet("background-color: white;")

        self.cursorCondition = QLineEdit(self.dialog12)
        self.cursorCondition.setStyleSheet("background-color: white;")
        self.cursorCondition.setPlaceholderText('Cursor 파일을 넣어주세요')

        self.cursorFile = QPushButton('File Open')
        self.cursorFile.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.cursorFile.clicked.connect(self.CursorFileOpen)
        font10 = self.cursorFile.font()
        font10.setBold(True)
        self.cursorFile.setFont(font10)

        labelSheetc = QLabel('시나리오 번호* : ', self.dialog12)
        labelSheetc.setStyleSheet("color: white;")
        font5 = labelSheetc.font()
        font5.setBold(True)
        labelSheetc.setFont(font5)
        self.D12_Sheetc = QLineEdit(self.dialog12)
        self.D12_Sheetc.setStyleSheet("background-color: white;")
        self.D12_Sheetc.setPlaceholderText('※ 입력 예시 : F01')

        sublayout5 = QGridLayout()
        sublayout5.addWidget(self.rbtn1, 0, 0)
        sublayout5.addWidget(self.rbtn2, 0, 1)
        sublayout5.addWidget(labelCursor, 1, 0)
        sublayout5.addWidget(self.cursorCondition, 1, 1)
        sublayout5.addWidget(self.cursorFile, 1, 2)
        sublayout5.addWidget(listCursor, 2, 0)
        sublayout5.addWidget(self.listCursor, 2, 1)
        sublayout5.addWidget(labelCursortext, 3, 0)
        sublayout5.addWidget(self.Cursortext, 3, 1)
        sublayout5.addWidget(labelSheetc, 4, 0)
        sublayout5.addWidget(self.D12_Sheetc, 4, 1)

        sublayout6 = QHBoxLayout()
        sublayout6.addStretch(2)
        sublayout6.addWidget(self.btn2)
        sublayout6.addWidget(self.btnDialog2)

        main_layout2 = QVBoxLayout()
        main_layout2.addLayout(sublayout5)
        main_layout2.addLayout(sublayout6)

        ### 탭 지정
        layout = QVBoxLayout()
        tabs = QTabWidget()
        tab1 = QWidget()  # 시나리오12
        tab2 = QWidget()  # cursor문
        tab1.setLayout(main_layout1)
        tab2.setLayout(main_layout2)
        tabs.addTab(tab1, "Step1")
        tabs.addTab(tab2, "Step2")
        layout.addWidget(tabs)

        self.dialog12.setLayout(layout)
        self.dialog12.setGeometry(300, 300, 650, 450)

        # ? 제거
        self.dialog12.setWindowFlags(Qt.WindowCloseButtonHint)

        self.dialog12.setWindowTitle('Scenario')
        self.dialog12.setWindowModality(Qt.NonModal)
        self.dialog12.show()

    def CursorFileOpen(self):
        self.listCursor.clear()
        fname = QFileDialog.getOpenFileName(self)
        self.cursorCondition.setText(fname[0])
        self.wb2 = pd.ExcelFile(fname[0])
        wbname = self.wb2.sheet_names
        for name in wbname:
            self.listCursor.addItem(str(name))
        self.dialog12.activateWindow()

    def Dialog13(self):
        self.dialoglist.add(13)
        self.dialog13 = QDialog()
        self.dialog13.setStyleSheet('background-color: #2E2E38')
        self.dialog13.setWindowIcon(QIcon('./EY_logo.png'))

        cursor = self.cnxn.cursor()

        sql = '''
                 SELECT 											
                        *
                 FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
            '''.format(field=self.selected_project_id)

        accountsname = pd.read_sql(sql, self.cnxn)
        ### 계정 트리
        self.new_tree = Form(self)
        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)

            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Checked)
        self.new_tree.get_selected_leaves()  # 초기값 모두 선택 (추가)

        ### JE Line / JE 라디오 버튼
        self.rbtn1 = QRadioButton('JE Line', self.dialog13)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)

        self.rbtn2 = QRadioButton('JE', self.dialog13)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        ### 버튼 1 - Extract Data
        self.btn2 = QPushButton('   Extract Data', self.dialog13)
        self.btn2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btn2.clicked.connect(self.Thread13)

        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)

        ### 버튼 2 - Close
        self.btnDialog = QPushButton('Close', self.dialog13)
        self.btnDialog.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close13)

        font9 = self.btnDialog.font()
        font9.setBold(True)
        self.btnDialog.setFont(font9)

        ### 라벨 1 - 연속된 자릿수
        label_Continuous = QLabel('연속된 자릿수* : ', self.dialog13)
        label_Continuous.setStyleSheet("color: white;")

        font1 = label_Continuous.font()
        font1.setBold(True)
        label_Continuous.setFont(font1)

        ### Text Edit - 연속된 자릿수
        self.text_continuous = QTextEdit(self.dialog13)
        self.text_continuous.setAcceptRichText(False)
        self.text_continuous.setStyleSheet("background-color: white;")
        self.text_continuous.setPlaceholderText('연속된 자릿수를 입력하세요 (입력 예시: 3333, 666666)')

        ### 라벨 2 - 중요성 금액
        label_amount = QLabel('중요성 금액 : ', self.dialog13)
        label_amount.setStyleSheet("color: white;")

        font3 = label_amount.font()
        font3.setBold(True)
        label_amount.setFont(font3)

        ### Line Edit - 중요성 금액
        self.line_amount = QLineEdit(self.dialog13)
        self.line_amount.setStyleSheet("background-color: white;")
        self.line_amount.setPlaceholderText('중요성 금액을 입력하세요')

        ### 라벨 3 - 계정 트리
        label_tree = QLabel('특정 계정명 : ', self.dialog13)
        label_tree.setStyleSheet("color: white;")
        font4 = label_tree.font()
        font4.setBold(True)
        label_tree.setFont(font4)

        labelSheet = QLabel('시나리오 번호* : ', self.dialog13)
        labelSheet.setStyleSheet("color: white;")

        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)

        ### Line Edit - 시트명
        self.D13_Sheet = QLineEdit(self.dialog13)
        self.D13_Sheet.setStyleSheet("background-color: white;")
        self.D13_Sheet.setPlaceholderText('※ 입력 예시 : F01')

        labelDC = QLabel('차변/대변* : ', self.dialog13)
        labelDC.setStyleSheet("color: white;")
        font1 = labelDC.font()
        font1.setBold(True)
        labelDC.setFont(font1)

        ### 차변/대변 체크박스로 구현
        self.checkC = QCheckBox('Credit', self.dialog13)
        self.checkD = QCheckBox('Debit', self.dialog13)
        self.checkC.setStyleSheet("color: white;")
        self.checkD.setStyleSheet("color: white;")
        self.checkC.setChecked(True)
        self.checkD.setChecked(True)

        layout_dc = QHBoxLayout()
        layout_dc.addWidget(labelDC)
        layout_dc.addWidget(self.checkC)
        layout_dc.addWidget(self.checkD)

        ### Layout - 다이얼로그 UI

        main_layout = QVBoxLayout()
        sublayout1 = QGridLayout()
        sublayout2 = QHBoxLayout()

        ### sublayout 배치 - 탭 삭제
        sublayout1.addWidget(self.rbtn1, 0, 0)
        sublayout1.addWidget(self.rbtn2, 0, 1)
        sublayout1.addWidget(label_Continuous, 1, 0)
        sublayout1.addWidget(self.text_continuous, 1, 1)
        sublayout1.addWidget(label_amount, 2, 0)
        sublayout1.addWidget(self.line_amount, 2, 1)
        sublayout1.addWidget(label_tree, 3, 0)
        sublayout1.addWidget(self.new_tree, 3, 1)
        sublayout1.addWidget(labelSheet, 4, 0)
        sublayout1.addWidget(self.D13_Sheet, 4, 1)

        sublayout2.addStretch(2)
        sublayout2.addWidget(self.btn2)
        sublayout2.addWidget(self.btnDialog)

        main_layout.addLayout(sublayout1, stretch=4)
        main_layout.addLayout(layout_dc, stretch=4)
        main_layout.addLayout(sublayout2, stretch=1)

        ### 공통 지정
        self.dialog13.setLayout(main_layout)
        self.dialog13.setGeometry(300, 300, 600, 450)

        # ? 제거
        self.dialog13.setWindowFlags(Qt.WindowCloseButtonHint)
        self.dialog13.setWindowTitle('Scenario13')
        self.dialog13.setWindowModality(Qt.NonModal)
        self.dialog13.show()

    def Dialog14(self):
        self.dialoglist.add(14)
        self.dialog14 = QDialog()
        self.dialog14.setStyleSheet('background-color: #2E2E38')
        self.dialog14.setWindowIcon(QIcon("./EY_logo.png"))

        cursor = self.cnxn.cursor()

        sql = '''
                 SELECT 											
                        *
                 FROM  [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] COA											
            '''.format(field=self.selected_project_id)

        accountsname = pd.read_sql(sql, self.cnxn)

        self.new_tree = Form(self)

        self.new_tree.tree.clear()
        accountType = accountsname.AccountType.unique()
        accountType.sort()
        for n, i in enumerate(accountType):
            self.new_tree.parent = QTreeWidgetItem(self.new_tree.tree)

            self.new_tree.parent.setText(0, "{}".format(i))
            self.new_tree.parent.setFlags(self.new_tree.parent.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
            child_items = accountsname.AccountSubType[
                accountsname.AccountType == accountType[n]].unique()
            child_items.sort()

            for m, x in enumerate(child_items):

                self.new_tree.child = QTreeWidgetItem(self.new_tree.parent)

                self.new_tree.child.setText(0, "{}".format(x))
                self.new_tree.child.setFlags(self.new_tree.child.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                grandchild_items = accountsname.AccountClass[accountsname.AccountSubType == child_items[m]].unique()
                grandchild_items.sort()
                for o, y in enumerate(grandchild_items):
                    self.new_tree.grandchild = QTreeWidgetItem(self.new_tree.child)

                    self.new_tree.grandchild.setText(0, "{}".format(y))
                    self.new_tree.grandchild.setFlags(
                        self.new_tree.grandchild.flags() | Qt.ItemIsTristate | Qt.ItemIsUserCheckable)
                    num_name = accountsname[accountsname.AccountClass == grandchild_items[o]].iloc[:, 2:4]
                    full_name = num_name["GLAccountNumber"].map(str) + ' ' + num_name["GLAccountName"]
                    full_name.sort_values(inplace=True)
                    for z in full_name:
                        self.new_tree.grandgrandchild = QTreeWidgetItem(self.new_tree.grandchild)

                        self.new_tree.grandgrandchild.setText(0, "{}".format(z))
                        self.new_tree.grandgrandchild.setFlags(
                            self.new_tree.grandgrandchild.flags() | Qt.ItemIsUserCheckable)
                        self.new_tree.grandgrandchild.setCheckState(0, Qt.Checked)
        self.new_tree.get_selected_leaves()  # 초기값 모두 선택 (추가)

        self.btn2 = QPushButton('   Extract Data', self.dialog14)
        self.btn2.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.btn2.clicked.connect(self.Thread14)

        font9 = self.btn2.font()
        font9.setBold(True)
        self.btn2.setFont(font9)

        self.btnDialog = QPushButton("   Close", self.dialog14)
        self.btnDialog.setStyleSheet(
            'color:white;  background-image : url(./bar.png)')
        self.btnDialog.clicked.connect(self.dialog_close14)

        font10 = self.btnDialog.font()
        font10.setBold(True)
        self.btnDialog.setFont(font10)

        self.btn2.resize(110, 30)
        self.btnDialog.resize(110, 30)

        # JE Line / JE 라디오 버튼
        self.rbtn1 = QRadioButton('JE Line', self.dialog14)
        self.rbtn1.setStyleSheet("color: white;")
        font11 = self.rbtn1.font()
        font11.setBold(True)
        self.rbtn1.setFont(font11)
        self.rbtn1.setChecked(True)
        self.rbtn2 = QRadioButton('JE', self.dialog14)
        self.rbtn2.setStyleSheet("color: white;")
        font12 = self.rbtn2.font()
        font12.setBold(True)
        self.rbtn2.setFont(font12)

        labelKeyword = QLabel('Key Words* : ', self.dialog14)
        labelKeyword.setStyleSheet("color: white;")

        font1 = labelKeyword.font()
        font1.setBold(True)
        labelKeyword.setFont(font1)

        self.D14_Key = QLineEdit(self.dialog14)
        self.D14_Key.setStyleSheet("background-color: white;")
        self.D14_Key.setPlaceholderText('검색할 단어를 입력하세요')

        labelTE = QLabel('중요성 금액 : ', self.dialog14)
        labelTE.setStyleSheet("color: white;")

        font2 = labelTE.font()
        font2.setBold(True)
        labelTE.setFont(font2)

        self.D14_TE = QLineEdit(self.dialog14)
        self.D14_TE.setStyleSheet("background-color: white;")
        self.D14_TE.setPlaceholderText('중요성 금액을 입력하세요')

        label_tree = QLabel('특정 계정 : ', self.dialog14)
        label_tree.setStyleSheet("color: white;")
        font4 = label_tree.font()
        font4.setBold(True)
        label_tree.setFont(font4)

        labelSheet = QLabel('시나리오 번호* : ', self.dialog14)
        labelSheet.setStyleSheet("color: white;")

        font5 = labelSheet.font()
        font5.setBold(True)
        labelSheet.setFont(font5)

        self.D14_Sheet = QLineEdit(self.dialog14)
        self.D14_Sheet.setStyleSheet("background-color: white;")
        self.D14_Sheet.setPlaceholderText('※ 입력 예시 : F01')

        labelDC = QLabel('차변/대변* : ', self.dialog14)
        labelDC.setStyleSheet("color: white;")
        font1 = labelDC.font()
        font1.setBold(True)
        labelDC.setFont(font1)

        # 차변/대변 체크박스로 구현
        self.checkC = QCheckBox('Credit', self.dialog14)
        self.checkD = QCheckBox('Debit', self.dialog14)
        self.checkC.setStyleSheet("color: white;")
        self.checkD.setStyleSheet("color: white;")
        self.checkC.setChecked(True)
        self.checkD.setChecked(True)

        self.D14_Key.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소
        self.D14_TE.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소
        self.D14_Sheet.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)  # LineEdit만 창 크기에 따라 확대/축소

        layout1 = QGridLayout()
        layout1.addWidget(self.rbtn1, 0, 0)
        layout1.addWidget(self.rbtn2, 0, 1)
        layout1.addWidget(labelKeyword, 1, 0)
        layout1.addWidget(self.D14_Key, 1, 1)
        layout1.addWidget(labelTE, 2, 0)
        layout1.addWidget(self.D14_TE, 2, 1)
        layout1.addWidget(label_tree, 3, 0)
        layout1.addWidget(self.new_tree, 3, 1)
        layout1.addWidget(labelSheet, 4, 0)
        layout1.addWidget(self.D14_Sheet, 4, 1)

        layout_dc = QHBoxLayout()
        layout_dc.addWidget(labelDC)
        layout_dc.addWidget(self.checkC)
        layout_dc.addWidget(self.checkD)

        layout2 = QHBoxLayout()
        layout2.addStretch()
        layout2.addStretch()
        layout2.addWidget(self.btn2)
        layout2.addWidget(self.btnDialog)

        layout2.setContentsMargins(-1, 10, -1, -1)

        main_layout = QVBoxLayout()
        main_layout.setAlignment(Qt.AlignTop)
        main_layout.addLayout(layout1)
        main_layout.addLayout(layout_dc)
        main_layout.addLayout(layout2)

        self.dialog14.setLayout(main_layout)
        self.dialog14.setGeometry(300, 300, 550, 350)

        # ? 제거
        self.dialog14.setWindowFlags(Qt.WindowCloseButtonHint)

        self.dialog14.setWindowTitle("Scenario14")
        self.dialog14.setWindowModality(Qt.NonModal)
        self.dialog14.show()

    def DialogIndex(self):
        self.dialogIndex = QDialog()
        self.dialogIndex.setStyleSheet('background-color: #2E2E38')
        self.dialogIndex.setWindowIcon(QIcon("./EY_logo.png"))
        self.dialogIndex.setWindowFlags(Qt.WindowCloseButtonHint)

        ##파일 불러오기
        filelabel = QLabel('파일 위치 : ', self.dialogIndex)
        filelabel.setStyleSheet("color: white;")
        font2 = filelabel.font()
        font2.setBold(True)
        filelabel.setFont(font2)

        self.loadCondition = QLineEdit(self.dialogIndex)
        self.loadCondition.setStyleSheet("background-color: white;")
        self.loadCondition.setReadOnly(True)
        self.loadCondition.setPlaceholderText('파일을 넣어주세요')

        self.loadFile = QPushButton('File Open')
        self.loadFile.setStyleSheet('color:white;  background-image : url(./bar.png)')
        self.loadFile.clicked.connect(self.IndexFileOpen)
        font3 = self.loadFile.font()
        font3.setBold(True)
        self.loadFile.setFont(font3)

        ## sheet 목록
        loadlabel = QLabel('Sheet List : ', self.dialogIndex)
        loadlabel.setStyleSheet("color: white;")
        font1 = loadlabel.font()
        font1.setBold(True)
        loadlabel.setFont(font1)

        self.loadtext = QTextEdit(self.dialogIndex)
        self.loadtext.setPlaceholderText('Sheet 목록이 표시됩니다\n메모장으로 사용하실 수 있습니다')
        self.loadtext.setStyleSheet("background-color: white;")

        main_layout = QGridLayout()
        main_layout.addWidget(filelabel, 0, 0)
        main_layout.addWidget(self.loadCondition, 0, 1)
        main_layout.addWidget(self.loadFile, 0, 2)
        main_layout.addWidget(loadlabel, 1, 0)
        main_layout.addWidget(self.loadtext, 1, 1)

        self.dialogIndex.setLayout(main_layout)
        self.dialogIndex.setGeometry(300, 300, 550, 350)

        self.dialogIndex.setWindowTitle("Dialog Index")
        self.dialogIndex.setWindowModality(Qt.NonModal)
        self.dialogIndex.show()

    def IndexFileOpen(self):
        fname = QFileDialog.getOpenFileName(self)
        self.loadCondition.setText(fname[0])
        wb = pd.ExcelFile(fname[0])
        sheetnames = wb.sheet_names
        namelist = 'Sheet List\n-------------------------------------------------------------'
        for name in sheetnames:
            namelist = namelist + '\n' + name
        self.loadtext.setText(namelist)
        self.dialogIndex.activateWindow()

    def dialog_close4(self):
        self.dialog4.close()

    def dialog_close5(self):
        self.dialog5.close()

    def dialog_close6(self):
        self.dialog6.close()

    def dialog_close7(self):
        self.dialog7.close()

    def dialog_close8(self):
        self.dialog8.close()

    def dialog_close9(self):
        self.dialog9.close()

    def dialog_close10(self):
        self.dialog10.close()

    def dialog_close12(self):
        self.dialog12.close()

    def dialog_close13(self):
        self.dialog13.close()

    def dialog_close14(self):
        self.dialog14.close()

    def dropFiles(self):
        self.listbox_drops.clear()

    def getSelectedItem(self):
        myItem = QListWidgetItem(self.listbox_drops.currentItem())
        return myItem.text()

    def Show_DataFrame_Group(self):
        tables = QGroupBox('데이터')
        self.setStyleSheet('QGroupBox  {color: white;}')
        font6 = tables.font()
        font6.setBold(True)
        tables.setFont(font6)
        box = QBoxLayout(QBoxLayout.TopToBottom)

        self.viewtable = QTableView(self)

        box.addWidget(self.viewtable)
        tables.setLayout(box)

        return tables

    ###########################
    ####추가하셔야 하는 함수들#####
    ###########################
    def Timer(self):
        self.secondTimer = 0
        self.timerVar.start()

    def printTime(self):
        self.secondTimer += 1
        elapsetime = "Elapsed time : " + str(int(self.secondTimer / 3600)) + "h " + str(
            int(self.secondTimer / 60)) + "m " + str(
            self.secondTimer) + "s"
        self.progressLabel.setText(elapsetime)

    def pClose(self):
        self.close()
        self.Action.close()
        for a in self.dialoglist:
            if a == 4:
                self.dialog4.close()
            elif a == 5:
                self.dialog5.close()
            elif a == 6:
                self.dialog6.close()
            elif a == 7:
                self.dialog7.close()
            elif a == 8:
                self.dialog8.close()
            elif a == 9:
                self.dialog9.close()
            elif a == 10:
                self.dialog10.close()
            elif a == 12:
                self.dialog12.close()
            elif a == 13:
                self.dialog13.close()
            elif a == 14:
                self.dialog14.close()

    def doAction(self):
        self.Timer()
        self.Action = QDialog()
        self.Action.setStyleSheet('background-color : black;')
        lbl_img = QLabel()
        label = QLabel('Now Loading')
        label.setStyleSheet("font : bold 14pt; color: white;")

        self.progressLabel = QLabel("Elapsed time : 0h 0m 0s")
        self.progressLabel.setStyleSheet("font : bold 8pt; color: grey;")
        pixmap = QPixmap('./Loading.png')
        lbl_img.setPixmap(pixmap)

        self.pclosebtn = QPushButton("Close", self.Action)
        self.pclosebtn.setStyleSheet('color:white; background-image : url(./bar.png) ;')
        self.pclosebtn.clicked.connect(self.pClose)

        self.progressBar = QProgressBar()
        self.progressBar.setRange(0, 0)
        self.progressBar.setStyleSheet("QProgressBar::chunk "
                                       "{"
                                       "background-color: yellow;"
                                       "}")
        sub_layout = QHBoxLayout()
        sub_layout.addWidget(lbl_img)

        sub_layout2 = QVBoxLayout()
        sub_layout2.addWidget(label)
        sub_layout2.addWidget(self.progressLabel)
        sub_layout2.setAlignment(Qt.AlignCenter)

        sub_layout3 = QHBoxLayout()
        sub_layout3.addStretch(2)
        sub_layout3.addWidget(self.pclosebtn)

        sub_layout.addLayout(sub_layout2)
        main_layout = QVBoxLayout()
        main_layout.addLayout(sub_layout)
        main_layout.addWidget(self.progressBar)
        main_layout.addLayout(sub_layout3)

        self.Action.setLayout(main_layout)
        self.Action.setGeometry(700, 400, 400, 220)
        self.Action.setWindowFlags(Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        self.Action.setWindowModality(Qt.NonModal)
        self.Action.show()

    ###Thread End시 로딩창 종료 및 조건 팝업
    def doneAction4(self):
        self.Action.close()
        self.timerVar.stop()

        if len(self.dataframe) > 1048576:
            self.alertbox_open3()

        elif self.rbtn1.isChecked():
            if len(self.dataframe) == 0:
                buttonReply = QMessageBox.information(self, '라인수 추출', '- 계정사용 빈도수가 ' + str(self.temp_N)
                                                      + '회 이하인 작성자에 의해 생성된 전표가 '
                                                      + str(len(self.dataframe) - 1) + '건 추출되었습니다. <br> - TE 금액('
                                                      + str(self.temp_TE) + ')을 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)

            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, '라인수 추출', '- 계정사용 빈도수가 ' + str(self.temp_N)
                                                      + '회 이하인 작성자에 의해 생성된 전표가 '
                                                      + str(len(self.dataframe) - 1) + '건 추출되었습니다. <br> - TE 금액('
                                                      + str(self.temp_TE) + ')을 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)

            else:
                buttonReply = QMessageBox.information(self, '라인수 추출', '- 계정사용 빈도수가 ' + str(self.temp_N) + '회 이하인 전표가 '
                                                      + str(len(self.dataframe)) + '건 추출되었습니다. <br> - TE 금액('
                                                      + str(self.temp_TE) + ')을 적용하였습니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog4.activateWindow()

        elif self.rbtn2.isChecked():
            if len(self.dataframe) == 0:
                buttonReply = QMessageBox.information(self, '라인수 추출', '- 계정사용 빈도수가' + str(self.temp_N)
                                                      + '회 이하인 작성자에 의해 생성된 전표가 '
                                                      + str(len(self.dataframe) - 1) + '건 추출되었습니다. <br> - TE 금액('
                                                      + str(self.temp_TE) + ')을 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]'
                                                      , QMessageBox.Ok)
            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, '라인수 추출', '- 계정사용 빈도수가' + str(self.temp_N)
                                                      + '회 이하인 작성자에 의해 생성된 전표가 '
                                                      + str(len(self.dataframe) - 1) + '건 추출되었습니다. <br> - TE 금액('
                                                      + str(self.temp_TE) + ')을 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]'
                                                      , QMessageBox.Ok)

            else:
                buttonReply = QMessageBox.information(self, '라인수 추출', '- 계정사용 빈도수가' + str(self.temp_N)
                                                      + '회 이하인 작성자에 의해 생성된 전표가 '
                                                      + str(len(self.dataframe) - 1) + '건 추출되었습니다. <br> - TE 금액('
                                                      + str(self.temp_TE) + ')을 적용하였습니다. <br> [전표번호 기준]'
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog4.activateWindow()

        self.th4.join()

    def doneAction5_SAP(self):
        self.Action.close()
        self.timerVar.stop()

        ### 예외처리 5 - 최대 라인 수 초과
        if len(self.dataframe) > 1048576:
            self.alertbox_open3()

        elif self.rbtn1.isChecked():
            if len(self.dataframe) == 0:
                buttonReply = QMessageBox.information(self, '라인수 추출', '-당기('
                                                      + str(self.tempYear_SAP) + ')에 생성된 계정 리스트가 '
                                                      + str(len(self.dataframe) - 1)
                                                      + '건 추출되었습니다. <br> - SKA1(' + str(self.dropped_items)
                                                      + ')를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)

            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, '라인수 추출', '-당기('
                                                      + str(self.tempYear_SAP) + ')에 생성된 계정 리스트가 '
                                                      + str(len(self.dataframe) - 1)
                                                      + '건 추출되었습니다. <br> - SKA1(' + str(self.dropped_items)
                                                      + ')를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)
            else:
                buttonReply = QMessageBox.information(self, '라인수 추출', '-당기('
                                                      + str(self.tempYear_SAP) + ')에 생성된 계정 리스트가 '
                                                      + str(len(self.dataframe) - 1)
                                                      + '건 추출되었습니다. <br> - SKA1(' + str(self.dropped_items)
                                                      + ')를 적용하였습니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog5.activateWindow()

        elif self.rbtn2.isChecked():
            if len(self.dataframe) == 0:
                buttonReply = QMessageBox.information(self, '라인수 추출', '-당기('
                                                      + str(self.tempYear_SAP) + ')에 생성된 계정 리스트가 '
                                                      + str(len(self.dataframe) - 1)
                                                      + '건 추출되었습니다. <br> - SKA1(' + str(self.dropped_items)
                                                      + ')를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]'
                                                      , QMessageBox.Ok)

            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, '라인수 추출', '-당기('
                                                      + str(self.tempYear_SAP) + ')에 생성된 계정 리스트가 '
                                                      + str(len(self.dataframe) - 1)
                                                      + '건 추출되었습니다. <br> - SKA1(' + str(self.dropped_items)
                                                      + ')를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]'
                                                      , QMessageBox.Ok)

            else:
                buttonReply = QMessageBox.information(self, '라인수 추출', '-당기('
                                                      + str(self.tempYear_SAP) + ')에 생성된 계정 리스트가 '
                                                      + str(len(self.dataframe) - 1)
                                                      + '건 추출되었습니다. <br> - SKA1(' + str(self.dropped_items)
                                                      + ')를 적용하였습니다. <br> [전표번호 기준]'
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog5.activateWindow()

        self.th5_SAP.join()

    def doneAction5_Non_SAP(self):
        self.Action.close()
        self.timerVar.stop()

        if len(self.dataframe) > 1048576:
            self.alertbox_open3()

        elif self.rbtn1.isChecked():
            if len(self.dataframe) == 0:
                buttonReply = QMessageBox.information(self, '라인수 추출', '-당기('
                                                      + str(self.tempYear_NonSAP) + ')에 생성된 계정 리스트가 '
                                                      + str(len(self.dataframe) - 1)
                                                      + '건 추출되었습니다. <br> - 계정코드(' + str(self.temp_Code_Non_SAP)
                                                      + ')를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)

            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, '라인수 추출', '-당기('
                                                      + str(self.tempYear_NonSAP) + ')에 생성된 계정 리스트가 '
                                                      + str(len(self.dataframe) - 1)
                                                      + '건 추출되었습니다. <br> - 계정코드(' + str(self.temp_Code_Non_SAP)
                                                      + ')를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)

            else:
                buttonReply = QMessageBox.information(self, '라인수 추출', '-당기('
                                                      + str(self.tempYear_NonSAP) + ')에 생성된 계정 리스트가 '
                                                      + str(len(self.dataframe) - 1)
                                                      + '건 추출되었습니다. <br> - 계정코드(' + str(self.temp_Code_Non_SAP)
                                                      + ')를 적용하였습니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog5.activateWindow()

        elif self.rbtn2.isChecked():
            if len(self.dataframe) == 0:
                buttonReply = QMessageBox.information(self, '라인수 추출', '-당기('
                                                      + str(self.tempYear_NonSAP) + ')에 생성된 계정 리스트가 '
                                                      + str(len(self.dataframe) - 1)
                                                      + '건 추출되었습니다. <br> - 계정코드(' + str(self.temp_Code_Non_SAP)
                                                      + ')를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]'
                                                      , QMessageBox.Ok)
            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, '라인수 추출', '-당기('
                                                      + str(self.tempYear_NonSAP) + ')에 생성된 계정 리스트가 '
                                                      + str(len(self.dataframe) - 1)
                                                      + '건 추출되었습니다. <br> - 계정코드(' + str(self.temp_Code_Non_SAP)
                                                      + ')를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]'
                                                      , QMessageBox.Ok)

            else:
                buttonReply = QMessageBox.information(self, '라인수 추출', '-당기('
                                                      + str(self.tempYear_NonSAP) + ')에 생성된 계정 리스트가 '
                                                      + str(len(self.dataframe) - 1)
                                                      + '건 추출되었습니다. <br> - 계정코드(' + str(self.temp_Code_Non_SAP)
                                                      + ')를 적용하였습니다. <br> [전표번호 기준]'
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog5.activateWindow()

        self.th5_Non_SAP.join()

    def doneAction6(self):
        self.Action.close()
        self.timerVar.stop()

        if len(self.dataframe) > 1048576:
            self.alertbox_open3()

        elif self.rbtn1.isChecked():
            if len(self.dataframe) == 0:
                buttonReply = QMessageBox.information(self, "라인수 추출",
                                                      "- 결산일(" + str(self.tempDate) + ") 전후" + str(int(self.tempTDate))
                                                      + "일에 입력된 전표가 " + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.tempCost)
                                                      + ")를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)

            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, "라인수 추출",
                                                      "- 결산일(" + str(self.tempDate) + ") 전후" + str(int(self.tempTDate))
                                                      + "일에 입력된 전표가 " + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.tempCost)
                                                      + ")를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)
            else:
                buttonReply = QMessageBox.information(self, "라인수 추출",
                                                      "- 결산일(" + str(self.tempDate) + ") 전후" + str(int(self.tempTDate))
                                                      + "일에 입력된 전표가 " + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.tempCost)
                                                      + ")를 적용하였습니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog6.activateWindow()

        elif self.rbtn2.isChecked():
            if len(self.dataframe) == 0:
                buttonReply = QMessageBox.information(self, "라인수 추출",
                                                      "- 결산일(" + str(self.tempDate) + ") 전후" + str(int(self.tempTDate))
                                                      + "일에 입력된 전표가 " + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.tempCost)
                                                      + ")를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)

            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, "라인수 추출",
                                                      "- 결산일(" + str(self.tempDate) + ") 전후" + str(int(self.tempTDate))
                                                      + "일에 입력된 전표가 " + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.tempCost)
                                                      + ")를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)
            else:
                buttonReply = QMessageBox.information(self, "라인수 추출",
                                                      "- 결산일(" + str(self.tempDate) + ") 전후" + str(int(self.tempTDate))
                                                      + "일에 입력된 전표가 " + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.tempCost)
                                                      + ")를 적용하였습니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog6.activateWindow()

        self.th6.join()

    def doneAction7(self):
        self.Action.close()
        self.timerVar.stop()

        if len(self.dataframe) > 1048576:
            self.alertbox_open3()

        elif self.rbtn3.isChecked():
            if len(self.dataframe) == 0:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 비영업일에 전기된 or 입력된 전표가 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.tempCost)
                                                      + ")를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)

            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 비영업일에 전기된 or 입력된 전표가 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.tempCost)
                                                      + ")를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)
            else:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 비영업일에 전기된 or 입력된 전표가 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.tempCost)
                                                      + ")를 적용하였습니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog7.activateWindow()

        elif self.rbtn4.isChecked():
            if len(self.dataframe) == 0:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 비영업일에 전기된 or 입력된 전표가 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.tempCost)
                                                      + ")를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)

            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 비영업일에 전기된 or 입력된 전표가 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.tempCost)
                                                      + ")를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)
            else:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 비영업일에 전기된 or 입력된 전표가 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.tempCost)
                                                      + ")를 적용하였습니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog7.activateWindow()

        self.th7.join()

    def doneAction8(self):
        self.Action.close()
        self.timerVar.stop()

        if len(self.dataframe) > 1048576:
            self.alertbox_open3()

        elif self.rbtn1.isChecked():
            if len(self.dataframe) == 0:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- Effective Date와 Entry Date 간 차이가 "
                                                      + str(int(self.realNDate)) + "인 전표가 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.tempCost)
                                                      + ")를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)

            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- Effective Date와 Entry Date 간 차이가 "
                                                      + str(int(self.realNDate)) + "인 전표가 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.tempCost)
                                                      + ")를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)
            else:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- Effective Date와 Entry Date 간 차이가 "
                                                      + str(int(self.realNDate)) + "인 전표가 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.tempCost)
                                                      + ")를 적용하였습니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog8.activateWindow()

        elif self.rbtn2.isChecked():
            if len(self.dataframe) == 0:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- Effective Date와 Entry Date 간 차이가 "
                                                      + str(int(self.realNDate)) + "인 전표가 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.tempCost)
                                                      + ")를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)

            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- Effective Date와 Entry Date 간 차이가 "
                                                      + str(int(self.realNDate)) + "인 전표가 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.tempCost)
                                                      + ")를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)
            else:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- Effective Date와 Entry Date 간 차이가 "
                                                      + str(int(self.realNDate)) + "인 전표가 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성 금액(" + str(self.tempCost)
                                                      + ")를 적용하였습니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog8.activateWindow()

        self.th8.join()

    def doneAction12(self):
        self.Action.close()
        self.timerVar.stop()

        if len(self.dataframe) > 1048576:
            self.alertbox_open3()

        elif len(self.dataframe) == 0:
            buttonReply = QMessageBox.information(self, "라인수 추출", "[중요성 금액: " + str(
                self.tempCost) + "] 라인수 " + str(len(self.dataframe) - 1) + "개입니다", QMessageBox.Ok)
            
        elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information( self, "라인수 추출", "[중요성 금액: " + str(
                self.tempCost) + "] 라인수 " + str(len(self.dataframe) - 1) + "개입니다 <br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]", QMessageBox.Ok)

        else:
            buttonReply = QMessageBox.information(self, "라인수 추출", "[중요성 금액: " + str(
                self.tempCost) + "] 라인수 " + str(len(self.dataframe) - 1) + "개입니다", QMessageBox.Ok)

        if buttonReply == QMessageBox.Ok: self.dialog12.activateWindow()

        self.th12.join()

    @pyqtSlot(str)
    def doneActionC(self, cursortext):
        self.Action.close()
        self.timerVar.stop()
        self.Cursortext.setText(cursortext)
        if len(self.dataframe) > 1048576:
            self.alertbox_open3()
        if len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame({'No Data': ['No Cursor']})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                buttonReply = QMessageBox.information(self, "라인수 추출", "총 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)
            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                buttonReply = QMessageBox.information(self, "라인수 추출", "총 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok:
                self.dialog12.activateWindow()
        
        elif len(self.dataframe) > 300:
            if self.rbtn1.isChecked():
                buttonReply = QMessageBox.information(self, "라인수 추출", "총 "
                                                      + str(len(self.dataframe))
                                                      + "건 추출되었습니다. <br> 추가 필터링이 필요해보입니다.<br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)
            elif self.rbtn2.isChecked():
                buttonReply = QMessageBox.information(self, "라인수 추출", "총 "
                                                      + str(len(self.dataframe))
                                                      + "건 추출되었습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)
            if buttonReply == QMessageBox.Ok:
                self.dialog12.activateWindow()
        
        else:
            if self.rbtn1.isChecked():
                buttonReply = QMessageBox.information(self, "라인수 추출", "총 "
                                                      + str(len(self.dataframe))
                                                      + "건 추출되었습니다.<br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)
            elif self.rbtn2.isChecked():
                buttonReply = QMessageBox.information(self, "라인수 추출", "총 "
                                                      + str(len(self.dataframe))
                                                      + "건 추출되었습니다.<br> [전표번호 기준]"
                                                      , QMessageBox.Ok)
            if buttonReply == QMessageBox.Ok:
                self.dialog12.activateWindow()
        self.thC.join()

    def doneAction13(self):
        self.Action.close()
        self.timerVar.stop()

        ### 예외처리 3 - 최대 추출 라인수
        if len(self.dataframe) > 1048576:
            self.alertbox_open3()

        elif self.rbtn1.isChecked():
            if len(self.dataframe) == 0:
                buttonReply = QMessageBox.information(self, '라인수 추출',
                                                      '- 연속된 숫자(' + str(self.temp_Continuous) + ')로 끝나는 금액을 검토한 결과 '
                                                      + str(len(self.dataframe) - 1) + '건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(self.temp_TE_13) + ')를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)
            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, '라인수 추출',
                                                      '- 연속된 숫자(' + str(self.temp_Continuous) + ')로 끝나는 금액을 검토한 결과 '
                                                      + str(len(self.dataframe) - 1) + '건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(self.temp_TE_13) + ')를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)

            else:
                buttonReply = QMessageBox.information(self, '라인수 추출',
                                                      '- 연속된 숫자(' + str(self.temp_Continuous) + ')로 끝나는 금액을 검토한 결과 '
                                                      + str(len(self.dataframe) - 1) + '건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(self.temp_TE_13) + ')를 적용하였습니다. <br> [전표라인번호 기준]'
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog13.activateWindow()

        elif self.rbtn2.isChecked():
            if len(self.dataframe) == 0:
                buttonReply = QMessageBox.information(self, '라인수 추출',
                                                      '- 연속된 숫자(' + str(self.temp_Continuous) + ')로 끝나는 금액을 검토한 결과 '
                                                      + str(len(self.dataframe) - 1) + '건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(self.temp_TE_13)+ ')를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]'
                                                      , QMessageBox.Ok)
            elif len(self.dataframe) > 300:
                buttonReply = QMessageBox.information(self, '라인수 추출',
                                                      '- 연속된 숫자(' + str(self.temp_Continuous) + ')로 끝나는 금액을 검토한 결과 '
                                                      + str(len(self.dataframe) - 1) + '건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(self.temp_TE_13) + ')를 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]'
                                                      , QMessageBox.Ok)

            else:
                buttonReply = QMessageBox.information(self, '라인수 추출',
                                                      '- 연속된 숫자(' + str(self.temp_Continuous) + ')로 끝나는 금액을 검토한 결과 '
                                                      + str(len(self.dataframe) - 1) + '건 추출되었습니다. <br> - 중요성 금액('
                                                      + str(self.temp_TE_13) + ')를 적용하였습니다. <br> [전표번호 기준]'
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok: self.dialog13.activateWindow()

        self.th13.join()

    def doneAction9(self):
        self.Action.close()
        self.timerVar.stop()

        if len(self.dataframe) > 1048576:
            self.alertbox_open3()

        if len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame({'No Data': ["[전표작성 빈도수: " + str(self.tempN) + "," + " 중요성금액: " + str(
                self.tempTE) + "] 라인수 " + str(len(self.dataframe)) + "개입니다"]})
            model = DataFrameModel(self.dataframe)
            model_refer = DataFrameModel(self.dataframe_refer)
            self.viewtable.setModel(model)
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.scenario_dic[self.tempSheet + '_Reference'] = self.dataframe_refer
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.addItem(self.tempSheet + '_Reference')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 2)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            if self.rbtn1.isChecked():

                buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표작성 빈도수가 " + str(self.tempN)
                                                      + "회 이하인 작성자에 의해 생성된 전표가 "
                                                      + str(len(self.dataframe) - 1) + "건 추출되었습니다. <br> - 중요성금액("
                                                      + str(
                    self.tempTE) + ")을 적용하였습니다.<br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)
            else:

                buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표작성 빈도수가 " + str(self.tempN)
                                                      + "회 이하인 작성자에 의해 생성된 전표가 "
                                                      + str(len(self.dataframe) - 1) + "건 추출되었습니다. <br> - 중요성금액("
                                                      + str(
                    self.tempTE) + ")을 적용하였습니다.<br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)
            if buttonReply == QMessageBox.Ok:
                self.dialog9.activateWindow()

        else:
            if self.rbtn1.isChecked():

                if len(self.dataframe) > 300:
                    buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표작성 빈도수가 " + str(self.tempN)
                                                          + "회 이하인 작성자에 의해 생성된 전표가 "
                                                          + str(
                        len(self.dataframe)) + "건 추출되었습니다. <br> - 중요성금액(" + str(self.tempTE)
                                                          + ")을 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]"
                                                          , QMessageBox.Ok)
                else:
                    buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표작성 빈도수가 " + str(self.tempN)
                                                          + "회 이하인 작성자에 의해 생성된 전표가 "
                                                          + str(
                        len(self.dataframe)) + "건 추출되었습니다. <br> - 중요성금액(" + str(self.tempTE)
                                                          + ")을 적용하였습니다. <br> [전표라인번호 기준]"
                                                          , QMessageBox.Ok)
                if buttonReply == QMessageBox.Ok:
                    self.dialog9.activateWindow()

            else:

                if len(self.dataframe) > 300:
                    buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표작성 빈도수가 " + str(self.tempN)
                                                          + "회 이하인 작성자에 의해 생성된 전표가 "
                                                          + str(
                        len(self.dataframe)) + "건 추출되었습니다. <br> - 중요성금액("
                                                          + str(self.tempTE) + ")을 적용하였습니다. <br> 추가 필터링이 필요해보입니다.<br> [전표번호 기준]"
                                                          , QMessageBox.Ok)
                else:
                    buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표작성 빈도수가 " + str(self.tempN)
                                                          + "회 이하인 작성자에 의해 생성된 전표가 "
                                                          + str(
                        len(self.dataframe)) + "건 추출되었습니다. <br> - 중요성금액("
                                                          + str(
                        self.tempTE) + ")을 적용하였습니다. <br> [전표번호 기준]"
                                                          , QMessageBox.Ok)
                if buttonReply == QMessageBox.Ok:
                    self.dialog9.activateWindow()
        self.th9.join()

    def doneAction10(self):
        self.Action.close()
        self.timerVar.stop()

        if len(self.dataframe) > 1048576:
            self.alertbox_open3()

        if len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame(
                {'No Data': ["[시작시점: " + str(
                    self.tempPoint1) + " 종료시점: " + str(self.tempPoint2) + " 중요성금액: " + str(
                    self.tempTE) + "] 라인수 " + str(
                    len(self.dataframe) - 1) + "개입니다"]})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            if self.rbtn1.isChecked():

                buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표가 "
                                                      + str(
                    len(self.dataframe) - 1) + "건 추출되었습니다. <br> - 중요성금액: "
                                                      + str(self.tempTE) + "<br> - 시작시점: " + str(
                    self.tempPoint1) + "<br> - 종료시점: " + str(
                    self.tempPoint2) + "을 적용하였습니다.<br> 추가 필터링이 필요해보입니다.<br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)
            else:

                buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표가 "
                                                      + str(
                    len(self.dataframe) - 1) + "건 추출되었습니다. <br> - 중요성금액: "
                                                      + str(self.tempTE) + "<br> - 시작시점: " + str(
                    self.tempPoint1) + "<br> - 종료시점: " + str(
                    self.tempPoint2) + "을 적용하였습니다.<br> 추가 필터링이 필요해보입니다.<br> [전표번호 기준]"
                                                      , QMessageBox.Ok)
            if buttonReply == QMessageBox.Ok:
                self.dialog10.activateWindow()

        else:
            if self.rbtn1.isChecked():

                if len(self.dataframe) > 300:
                    buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표가 "
                                                          + str(
                        len(self.dataframe)) + "건 추출되었습니다. <br> - 중요성금액: "
                                                          + str(self.tempTE) + "<br> - 시작시점: " + str(
                        self.tempPoint1) + "<br> - 종료시점: " + str(self.tempPoint2) + "을 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]"
                                                          , QMessageBox.Ok)
                else:
                    buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표가 "
                                                          + str(
                        len(self.dataframe)) + "건 추출되었습니다. <br> - 중요성금액: "
                                                          + str(self.tempTE) + "<br> - 시작시점: " + str(
                        self.tempPoint1) + "<br> - 종료시점: " + str(
                        self.tempPoint2) + "을 적용하였습니다.<br> [전표라인번호 기준]"
                                                          , QMessageBox.Ok)
                if buttonReply == QMessageBox.Ok:
                    self.dialog10.activateWindow()
            else:
                if len(self.dataframe) > 300:
                    buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표가 "
                                                          + str(
                        len(self.dataframe)) + "건 추출되었습니다. <br> - 중요성금액: "
                                                          + str(self.tempTE) + "<br> - 시작시점: " + str(
                        self.tempPoint1) + "<br> - 종료시점: " + str(self.tempPoint2) + "을 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]"
                                                          , QMessageBox.Ok)
                else:
                    buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표가 "
                                                          + str(
                        len(self.dataframe)) + "건 추출되었습니다. <br> - 중요성금액: "
                                                          + str(self.tempTE) + "<br> - 시작시점: " + str(
                        self.tempPoint1) + "<br> - 종료시점: " + str(
                        self.tempPoint2) + "을 적용하였습니다. <br> [전표번호 기준]"
                                                          , QMessageBox.Ok)
                if buttonReply == QMessageBox.Ok:
                    self.dialog10.activateWindow()
        self.th10.join()

    def doneAction14(self):
        self.Action.close()
        self.timerVar.stop()

        if len(self.dataframe) > 1048576:
            self.alertbox_open3()

        if len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame({'No Data': ["[전표 적요 특정단어: " + str(self.baseKey) + "," + " 중요성금액: " + str(
                self.tempTE) + "] 라인수 " + str(len(self.dataframe)) + "개입니다"]})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
            if self.rbtn1.isChecked():
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표 적요에 "
                                                      + str(self.baseKey) + "이/가 포함된 전표가 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성금액(" + str(self.tempTE)
                                                      + ")을 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표라인번호 기준]"
                                                      , QMessageBox.Ok)
            else:
                buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표 적요에 "
                                                      + str(self.baseKey) + "이/가 포함된 전표가 "
                                                      + str(len(self.dataframe) - 1)
                                                      + "건 추출되었습니다. <br> - 중요성금액(" + str(self.tempTE)
                                                      + ")을 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]"
                                                      , QMessageBox.Ok)

            if buttonReply == QMessageBox.Ok:
                self.dialog14.activateWindow()

        else:
            if self.rbtn1.isChecked():

                if len(self.dataframe) > 300:
                    buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표 적요에 "
                                                          + str(self.baseKey) + "이/가 포함된 전표가 "
                                                          + str(len(self.dataframe))
                                                          + "건 추출되었습니다. <br> - 중요성금액(" + str(self.tempTE)
                                                          + ")을 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br>  [전표라인번호 기준]"
                                                          , QMessageBox.Ok)
                else:
                    buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표 적요에 "
                                                          + str(self.baseKey) + "이/가 포함된 전표가 "
                                                          + str(len(self.dataframe))
                                                          + "건 추출되었습니다. <br> - 중요성금액(" + str(self.tempTE)
                                                          + ")을 적용하였습니다. <br> [전표라인번호 기준]"
                                                          , QMessageBox.Ok)
                if buttonReply == QMessageBox.Ok:
                    self.dialog14.activateWindow()
            else:
                if len(self.dataframe) > 300:
                    buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표 적요에 "
                                                          + str(self.baseKey) + "이/가 포함된 전표가 "
                                                          + str(len(self.dataframe))
                                                          + "건 추출되었습니다. <br> - 중요성금액(" + str(self.tempTE)
                                                          + ")을 적용하였습니다. <br> 추가 필터링이 필요해보입니다. <br> [전표번호 기준]"
                                                          , QMessageBox.Ok)
                else:
                    buttonReply = QMessageBox.information(self, "라인수 추출", "- 전표 적요에 "
                                                          + str(self.baseKey) + "이/가 포함된 전표가 "
                                                          + str(len(self.dataframe))
                                                          + "건 추출되었습니다. <br> - 중요성금액(" + str(self.tempTE)
                                                          + ")을 적용하였습니다. <br> [전표번호 기준]"
                                                          , QMessageBox.Ok)
                if buttonReply == QMessageBox.Ok:
                    self.dialog14.activateWindow()
        self.th14.join()

    def Thread4(self):
        self.temp_N = self.D4_N.text()  # 필수값
        self.temp_TE = self.D4_TE.text()
        self.tempSheet = self.D4_Sheet.text()

        ### 예외처리 1 - 필수값 입력 누락
        if self.temp_N == '' or self.tempSheet == '':
            self.alertbox_open()

        ### 예외처리 2 - 시트명 중복 확인 (JE Line)
        elif self.rbtn1.isChecked() and (
                self.combo_sheet.findText(self.tempSheet + '_Result') != -1 or self.combo_sheet.findText(
            self.tempSheet + '_Reference') != -1):
            self.alertbox_open5()

        ### 예외처리 3 - 시트명 중복 확인 (JE)
        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        elif not (self.checkC.isChecked()) and not (self.checkD.isChecked()):
            self.alertbox_open7()

        ### 쿼리 연동
        else:
            if self.temp_TE == '': self.temp_TE = 0
            try:
                int(self.temp_N)
                int(self.temp_TE)

                if checked_account == 'AND JournalEntries.GLAccountNumber IN ()':
                    self.checked_account4 = ''
                else:
                    self.checked_account4 = checked_account

                self.doAction()
                self.th4 = Thread(target=self.extButtonClicked4)
                self.th4.daemon = True
                self.th4.start()

            ### 예외처리 5 - 필수 입력값 타입 오류
            except ValueError:
                try:
                    int(self.temp_N)
                    try:
                        int(self.temp_TE)
                    except:
                        self.alertbox_open2('중요성금액')
                except:
                    try:
                        int(self.temp_TE)
                        self.alertbox_open2('계정사용 빈도수')
                    except:
                        self.alertbox_open2('계정사용 빈도수와 중요성금액')

    ####수정사항 시작####
    def ChangeInt(self, row):
        try:
            return str(int(row))
        except:
            return str(row)

    def Thread5_SAP(self):
        ### 입력값 - 시트명, 연도
        self.tempSheet_SAP = self.D5_Sheet.text()
        self.tempYear_SAP = self.pname_year

        ### 예외처리0 - 차/대 체크 오류
        if not (self.checkC2.isChecked()) and not (self.checkD2.isChecked()):
            self.alertbox_open7()

        ## 예외처리1 - SKA1 드롭박스에 아무 것도 없는 경우
        elif self.listbox_drops.count() == 0:
            self.MessageBox_Open("파일이 Drop되지 않았습니다")

        ### 예외처리 2 - 필수값 누락
        elif self.tempSheet_SAP == '':
            self.alertbox_open()

        ### 예외처리 3 - 시트명 중복 확인
        elif self.rbtn1.isChecked() and self.combo_sheet.findText(self.tempSheet_SAP + '_Result') != -1:
            self.alertbox_open5()

        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet_SAP + '_Journals') != -1:
            self.alertbox_open5()

        else:
            self.dropped_items = []  ### ListBox 인풋값 append

            for i in range(self.listbox_drops.count()):
                myItem = QListWidgetItem(self.listbox_drops.item(i))
                myItem = str(myItem.text())
                self.dropped_items.append(myItem)
                # self.listbox_drops.item(i) = re.sub(r'\'', '/', self.listbox_drops.item(i))

            ### 예외처리 2 - SKA1 파일인지 확인 후 df로 변환
            df = pd.DataFrame()  ### dataframe으로 저장
            count = 0
            for file in self.dropped_items:
                if 'SKA1' in file:
                    df = df.append(pd.read_csv(file, sep='|'))
                else:
                    count += 1
                    break

            ## 예외처리 - SKA1 파일 이외의 파일이 드롭된 경우
            if count > 0:
                self.alertbox_open10()
                return

            ## 예외처리 - SKA1 파일을 읽어들인 결과 빈 파일인 경우
            if len(df) == 0:
                self.MessageBox_Open("SKA1 파일이 비어있습니다.")
                return

            ## 예외처리 - 생성일자 혹은 계정코드 필드가 없을 경우
            if not 'ERDAT' in df.columns.tolist() or not 'SAKNR' in df.columns.tolist():
                self.alertbox_open11('"ERDAT", "SAKNR" 필드를 찾을 수 없습니다')
                return

            if checked_account == 'AND JournalEntries.GLAccountNumber IN ()':
                self.checked_account5_SAP = ''
            else:
                self.checked_account5_SAP = checked_account

            ### 당기 생성된 계정 코드 반환
            df['ERDAT'] = df['ERDAT'].apply(lambda row: str(row))
            df = df[df['ERDAT'].str.slice(start=0, stop=4) == self.tempYear_SAP]
            self.AccCode_SKA1 = df['SAKNR'].apply(lambda row : self.ChangeInt(row))

            temp_query = """SELECT GLAccountNumber
                            FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts]""".format(field=self.selected_project_id)
            AccCode_GL = pd.read_sql(temp_query, self.cnxn)

            mask = AccCode_GL['GLAccountNumber'].apply(lambda row : self.ChangeInt(row) in self.AccCode_SKA1.values)
            AccCode_GL = AccCode_GL[mask]['GLAccountNumber'].apply(lambda row : f"\'{row}\'")

            self.real_Code = ','.join(AccCode_GL)

            ## 예외처리 - 당기 생선된 계정이 없을 시
            if self.real_Code == '':
                self.readl_Code = "\'\'"

            self.doAction()
            self.th5_SAP = Thread(target=self.extButtonClicked5_SAP)
            self.th5_SAP.daemon = True
            self.th5_SAP.start()
        ####수정사항 끝####


    def Thread5_Non_SAP(self):
        ### 예외처리 0 - 차/대 체크 오류
        if not (self.checkC1.isChecked()) and not (self.checkD1.isChecked()):
            self.alertbox_open7()

        else:
            if checked_account == 'AND JournalEntries.GLAccountNumber IN ()':
                self.checked_account5_Non = ''
            else:
                self.checked_account5_Non = checked_account

            self.tempSheet_NonSAP = self.D5_Sheet2.text()  # 필수값
            self.tempYear_NonSAP = int(self.pname_year)  # 필수값
            self.temp_Code_Non_SAP_1 = self.MyInput.toPlainText()  # 필수값 (계정코드)

            self.temp_Code_Non_SAP_1 = re.sub(r"[:,|\s]", ",", self.temp_Code_Non_SAP_1)
            self.temp_Code_Non_SAP_1 = re.split(",", self.temp_Code_Non_SAP_1)

            self.temp_Code_Non_SAP = ''
            for code in self.temp_Code_Non_SAP_1:
                self.temp_Code_Non_SAP += "'" + str(code) + "', "

            self.temp_Code_Non_SAP = str(self.temp_Code_Non_SAP)
            self.temp_Code_Non_SAP = self.temp_Code_Non_SAP[:-2]

            ### 예외처리 1 - 필수값 입력 누락
            if self.temp_Code_Non_SAP == '' or self.tempSheet_NonSAP == '':
                self.alertbox_open()

            ### 예외처리 2 - 시트명 중복 확인
            elif self.rbtn1.isChecked() and self.combo_sheet.findText(self.tempSheet_NonSAP + '_Result') != -1:
                self.alertbox_open5()

            elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet_NonSAP + '_Journals') != -1:
                self.alertbox_open5()

            else:
                self.doAction()
                self.th5_Non_SAP = Thread(target=self.extButtonClicked5_Non_SAP)
                self.th5_Non_SAP.daemon = True
                self.th5_Non_SAP.start()

    ###extraction버튼 클릭 시 유효성 확인 및 Thread 시작
    def Thread6(self):
        self.tempDate = self.D6_Date.text()
        self.tempTDate = self.D6_Date2.text()
        self.tempCost = self.D6_Cost.text()
        self.tempSheet = self.D6_Sheet.text()

        if self.tempDate == '' or self.tempSheet == '':
            self.alertbox_open()

        elif self.rbtn1.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Result') != -1:
            self.alertbox_open5()

        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        elif not (self.checkC.isChecked()) and not (self.checkD.isChecked()):
            self.alertbox_open7()

        elif not (self.checkB.isChecked()) and not (self.checkF.isChecked()):
            self.alertbox_open14()

        else:
            if self.tempCost == '': self.tempCost = 0
            if self.tempTDate == '': self.tempTDate = 0

            if checked_account == 'AND JournalEntries.GLAccountNumber IN ()':
                self.checked_account6 = ''
            else:
                self.checked_account6 = checked_account

            if checked_preparer == 'AND JournalEntries.PreparerID IN ()':
                self.checked_preparer6 = ''
            else:
                self.checked_preparer6 = checked_preparer

            try:
                int(self.tempTDate)
                int(self.tempCost)

                self.tempTDate = int(self.tempTDate)

                if self.tempTDate < 0 or self.tempTDate > 700000:
                    self.alertbox_open12()
                    int('False')

                else:
                    self.realDate = date.fromisoformat(self.tempDate)

                    self.first_origin = self.realDate - timedelta(days=int(self.tempTDate))
                    self.second_origin = self.realDate + timedelta(days=int(self.tempTDate))

                    self.first_mid = str(self.first_origin).split('-')
                    self.second_mid = str(self.second_origin).split('-')
                    self.base_mid = str(self.realDate).split('-')

                    self.first_final = "'" + self.first_mid[0] + self.first_mid[1] + self.first_mid[2] + "'"  # 이전
                    self.second_final = "'" + self.second_mid[0] + self.second_mid[1] + self.second_mid[2] + "'"  # 이후
                    self.base_final = "'" + self.base_mid[0] + self.base_mid[1] + self.base_mid[2] + "'"  # 기준

                    if self.checkB.isChecked() and self.checkF.isChecked():
                        self.first = self.first_final
                        self.second = self.second_final

                    elif self.checkB.isChecked():
                        self.first = self.first_final
                        self.second = self.base_final

                    elif self.checkF.isChecked():
                        self.first = self.base_final
                        self.second = self.second_final

                    self.doAction()
                    self.th6 = Thread(target=self.extButtonClicked6)
                    self.th6.daemon = True
                    self.th6.start()

            except ValueError:
                try:
                    int(self.tempTDate)
                    try:
                        int(self.tempCost)
                    except:
                        self.alertbox_open2('중요성금액')
                except:
                    try:
                        int(self.tempCost)
                        self.alertbox_open2('T')
                    except:
                        self.alertbox_open2('T값과 중요성금액')

    def Thread7(self):
        self.holiday = []
        for i in range(2020, 2031):
            self.holiday.append(pytimekr.holidays(i))

        self.holiday_str = []

        for i in range(len(self.holiday)):
            for d in range(0, len(self.holiday[i])):
                self.date_str = self.holiday[i][d].strftime('%Y-%m-%d')
                self.holiday_str.append(self.date_str)

        for i in self.fianlDate:
            self.holiday_str.append(i)

        self.start_date = date(2021, 1, 1)
        self.end_date = date(2023, 12, 31)
        self.delta = timedelta(days=1)
        while self.start_date <= self.end_date:
            if self.start_date.weekday() == 5 or self.start_date.weekday() == 6:  # 주말 추가
                self.a = self.start_date.strftime('%Y-%m-%d')
                self.holiday_str.append(self.a)
            self.start_date += self.delta

        self.fianlDate = []  # 초기화 작업

        self.realDate_List = []  # SQL 쿼리에 들어갈 리스트

        for i in range(0, len(self.holiday_str)):
            self.tempDate = []
            self.tempDate = str(self.holiday_str[i]).split('-')
            self.realDate = self.tempDate[0] + self.tempDate[1] + self.tempDate[2]
            self.realDate_List.append(self.realDate)

        self.checked_date = ''
        for i in self.realDate_List:
            self.checked_date = self.checked_date + ',' + '\'' + i + '\''

        self.checked_date = self.checked_date[1:]

        self.checked_effective = 'AND JournalEntries.EffectiveDate IN (' + self.checked_date + ')'
        self.checked_entry = 'AND JournalEntries.EntryDate IN (' + self.checked_date + ')'

        self.tempCost = self.D7_Cost.text()
        self.tempSheet = self.D7_Sheet.text()

        if self.rbtn1.isChecked():  # Effective Date 일 때
            self.tempState = self.checked_effective

        elif self.rbtn2.isChecked():  # Entry Date 일 때
            self.tempState = self.checked_entry

        if self.tempCost == '':
            self.tempCost = 0

        if self.tempDate == '' or self.tempSheet == '':
            ####수정사항 시작####
            self.alertbox_open()
            ####수정사항 끝####

        elif not (self.checkC.isChecked()) and not (self.checkD.isChecked()):
            self.alertbox_open7()

        elif self.combo_sheet.findText(self.tempSheet) != -1:
            self.alertbox_open5()

        elif self.rbtn3.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Result') != -1:
            self.alertbox_open5()

        elif self.rbtn4.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        else:

            try:
                int(self.tempCost)

                if checked_account == 'AND JournalEntries.GLAccountNumber IN ()':
                    self.checked_account7 = ''
                else:
                    self.checked_account7 = checked_account

                if checked_preparer == 'AND JournalEntries.PreparerID IN ()':
                    self.checked_preparer7 = ''
                else:
                    self.checked_preparer7 = checked_preparer

                self.doAction()
                self.th7 = Thread(target=self.extButtonClicked7)
                self.th7.daemon = True
                self.th7.start()

            except ValueError:
                self.alertbox_open2('중요성 금액')

    def Thread8(self):
        self.tempN = self.D8_N.text()
        self.tempCost = self.D8_Cost.text()
        self.tempSheet = self.D8_Sheet.text()

        if self.tempN == '' or self.tempSheet == '':
            self.alertbox_open()

        elif not (self.checkC.isChecked()) and not (self.checkD.isChecked()):
            self.alertbox_open7()

        elif self.rbtn1.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Result') != -1:
            self.alertbox_open5()

        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        else:

            if self.tempCost == '': self.tempCost = 0
            try:
                int(self.tempN)
                int(self.tempCost)

                if int(self.tempN) < 0 or int(self.tempN) > 700000:
                    self.alertbox_open13()
                    int('False')

                else:

                    if checked_account == 'AND JournalEntries.GLAccountNumber IN ()':
                        self.checked_account8 = ''
                    else:
                        self.checked_account8 = checked_account

                    if checked_preparer == 'AND JournalEntries.PreparerID IN ()':
                        self.checked_preparer8 = ''
                    else:
                        self.checked_preparer8 = checked_preparer

                    self.realNDate = int(self.tempN)

                    self.doAction()
                    self.th8 = Thread(target=self.extButtonClicked8)
                    self.th8.daemon = True
                    self.th8.start()

            except ValueError:
                try:
                    int(self.tempN)
                    try:
                        int(self.tempCost)
                    except:
                        self.alertbox_open2('중요성금액')
                except:
                    try:
                        int(self.tempCost)
                        self.alertbox_open2('N')
                    except:
                        self.alertbox_open2('N값과 중요성금액')

    def Thread12(self):
        self.tempCost = self.D12_Cost.text()
        self.tempSheet = self.D12_Sheet12.text()
        self.tempState12 = ''

        if self.tempCost == '':
            self.tempCost = 0

        if self.tempSheet == '':
            self.alertbox_open()

        elif self.combo_sheet.findText(self.tempSheet + '_Reference') != -1:
            self.alertbox_open5()

        elif not (self.checkC1.isChecked()) and not (self.checkD1.isChecked()):
            self.alertbox_open7()

        else:
            try:
                int(self.tempCost)
                if (self.checkC1.isChecked()) and (self.checkD1.isChecked()):
                    self.tempState12 = 'LVL4.GL_Account_Position IN (' + "'" + 'Credit' + "'" + "," + "'" + 'Debit' + "'" + ')'
                elif self.checkC1.isChecked():
                    self.tempState12 = 'LVL4.GL_Account_Position =' + "'" + 'Credit' + "'"
                elif self.checkD1.isChecked():
                    self.tempState12 = 'LVL4.GL_Account_Position =' + "'" + 'Debit' + "'"

                if checked_account_12 == 'AND LVL4.GL_Account_Number IN ()':
                    self.checked_account12 = ''
                else:
                    self.checked_account12 = checked_account_12

                self.doAction()
                self.th12 = Thread(target=self.extButtonClicked12)
                self.th12.daemon = True
                self.th12.start()

            except ValueError:
                self.alertbox_open2('중요성 금액')

    def Thread13(self):
        self.temp_Continuous = self.text_continuous.toPlainText()  # 필수
        self.temp_Continuous = str(self.temp_Continuous)
        self.temp_TE_13 = self.line_amount.text()
        self.tempSheet = self.D13_Sheet.text()  # 필수

        ###예외처리 0 - TE 금액 누락시
        if self.temp_TE_13 == '':
            self.temp_TE_13 = 0
        ####수정사항 시작####
        if self.temp_Continuous != '' and len(str(self.temp_Continuous)) < 6:
            self.alertbox_open15()
        ####수정사항 끝####
        ### 예외처리 1 - 필수값 누락
        elif self.temp_Continuous == '' or self.tempSheet == '':
            self.alertbox_open()

        ### 예외처리 2 - 시트명 중복 확인
        elif self.rbtn1.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Result') != -1:
            self.alertbox_open5()

        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        elif not (self.checkC.isChecked()) and not (self.checkD.isChecked()):
            self.alertbox_open7()

        else:
            try:
                int(self.temp_TE_13)
                int(self.temp_Continuous)

                if checked_account == 'AND JournalEntries.GLAccountNumber IN ()':
                    self.checked_account13 = ''
                else:
                    self.checked_account13 = checked_account

                self.doAction()
                self.th13 = Thread(target=self.extButtonClicked13)
                self.th13.daemon = True
                self.th13.start()

            ### 예외처리 4 - 필수값 타입 오류
            except ValueError:
                try:
                    int(self.temp_Continuous)
                    try:
                        int(self.temp_TE_13)
                    except:
                        self.alertbox_open2('중요성금액')
                except:
                    try:
                        int(self.temp_TE_13)
                        self.alertbox_open2('연속된 자릿수')
                    except:
                        self.alertbox_open2('연속된 자릿수와 중요성금액')

    def ThreadC(self):
        self.tempSheet = self.D12_Sheetc.text()
        self.cursorpath = self.cursorCondition.text()
        self.wbC = self.wb2.parse(self.listCursor.currentText())
        if self.tempSheet == '' or self.cursorpath == '':
            self.alertbox_open()
            # 시트명 중복 확인
        elif self.rbtn1.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Result') != -1:
            self.alertbox_open5()

        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        elif os.path.isfile(self.cursorpath) == False:
            self.MessageBox_Open("경로에 해당 파일이 존재하지 않습니다")
        elif len(self.wbC.columns) <= 12:
            self.alertbox_open4('Cursor 필드가 존재하지 않습니다')
        elif self.wbC.iloc[:, 12].empty == True:
            self.alertbox_open4('Check된 조건이 없습니다')
        elif self.wbC.iloc[:, [0, 3, 5, 8]].isnull().any().any() == True:
            self.alertbox_open4('필요 조건 필드를 충족하지 않습니다')
        else:
            try:
                self.wbC.astype({'GL_Account_Number': 'int64', 'Analysis_GL_Account_Number': 'int64'})
                self.doAction()
                self.thC = Thread(target=self.extButtonClickedC)
                self.thC.start()
            except ValueError:
                self.alertbox_open4('필요 조건필드의 데이터 타입을 확인 바랍니다')

    def Thread9(self):
        self.tempN = self.D9_N.text()  # 필수값
        self.tempTE = self.D9_TE.text()
        self.tempSheet = self.D9_Sheet.text()
        if self.tempN == '' or self.tempSheet == '':
            self.alertbox_open()
        # 시트명 중복 확인
        elif self.rbtn1.isChecked() and (
                self.combo_sheet.findText(self.tempSheet + '_Result') != -1 or self.combo_sheet.findText(
            self.tempSheet + '_Reference') != -1):
            self.alertbox_open5()

        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()


        elif not (self.checkC.isChecked()) and not (self.checkD.isChecked()):
            self.alertbox_open7()
        else:
            if self.tempTE == '': self.tempTE = 0
            try:
                int(self.tempN)
                int(self.tempTE)

                if checked_account == 'AND JournalEntries.GLAccountNumber IN ()':
                    self.checked_account9 = ''
                else:
                    self.checked_account9 = checked_account

                self.doAction()
                self.th9 = Thread(target=self.extButtonClicked9)
                self.th9.daemon = True
                self.th9.start()
            except ValueError:
                try:
                    int(self.tempN)
                    try:
                        int(self.tempTE)
                    except:
                        self.alertbox_open4('중요성금액을 숫자로만 입력해주시기 바랍니다.')
                except:
                    try:
                        int(self.tempTE)
                        self.alertbox_open4('작성빈도수를 숫자로만 입력해주시기 바랍니다.')
                    except:
                        self.alertbox_open4('작성빈도수와 중요성금액을 숫자로만 입력해주시기 바랍니다.')

    def Thread10(self):
        self.tempTE = self.D10_TE.text()
        self.tempSheet = self.D10_Sheet.text()  # 필수값
        self.tempPoint1 = self.D10_Point1.text()
        self.tempPoint2 = self.D10_Point2.text()

        if self.tempSheet == '':
            self.alertbox_open()

        # 시트명 중복 확인
        elif self.rbtn1.isChecked() and (
                self.combo_sheet.findText(self.tempSheet + '_Result') != -1 or self.combo_sheet.findText(
            self.tempSheet + '_Reference') != -1):
            self.alertbox_open5()

        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        elif not (self.checkC.isChecked()) and not (self.checkD.isChecked()):
            self.alertbox_open7()

        else:
            if self.tempTE == '': self.tempTE = 0
            if self.tempPoint1 == '':
                self.tempPoint1 = '1000-01-01'
            if self.tempPoint2 == '':
                self.tempPoint2 = '9999-12-31'

            try:
                int(self.tempTE)
                try:
                    int(self.tempPoint1[0:4])
                    int(self.tempPoint1[5:7])
                    int(self.tempPoint1[8:10])
                    int(self.tempPoint2[0:4])
                    int(self.tempPoint2[5:7])
                    int(self.tempPoint2[8:10])

                    if len(str(self.tempPoint1)) != 10 or len(str(self.tempPoint2)) != 10:
                        self.alertbox_open4("시점은 'yyyy-mm-dd'의 형태로 입력해주시기 바랍니다.")
                    elif ((self.tempPoint1[5:7] < '01' or self.tempPoint1[8:10] > '31') and (
                            self.tempPoint2[5:7] >= '01' and self.tempPoint2[8:10] <= '31')):
                        self.alertbox_open4("해당 월일을 올바르게 입력해주시기 바랍니다.")
                    elif ((self.tempPoint2[5:7] < '01' or self.tempPoint2[8:10] > '31')
                          and (self.tempPoint1[5:7] >= '01' and self.tempPoint1[8:10] <= '31')):
                        self.alertbox_open4("해당 월일을 올바르게 입력해주시기 바랍니다.")
                    elif ((self.tempPoint1[5:7] < '01' or self.tempPoint1[8:10] > '31')
                          and (self.tempPoint2[5:7] < '01' and self.tempPoint2[8:10] > '31')):
                        self.alertbox_open4("해당 월일을 올바르게 입력해주시기 바랍니다.")
                    else:
                        if checked_account == 'AND JournalEntries.GLAccountNumber IN ()':
                            self.checked_account10 = ''
                        else:
                            self.checked_account10 = checked_account

                        if checked_preparer == 'AND JournalEntries.PreparerID IN ()':
                            self.checked_preparer10 = ''
                        else:
                            self.checked_preparer10 = checked_preparer
                        self.doAction()
                        self.th10 = Thread(target=self.extButtonClicked10)
                        self.th10.daemon = True
                        self.th10.start()

                except:
                    if self.tempPoint1[5:7] == '' or self.tempPoint1[8:10] == '' or self.tempPoint2[
                                                                                    5:7] == '' or self.tempPoint2[
                                                                                                  8:10] == '':
                        self.alertbox_open4("시점은 'yyyy-mm-dd'의 형태로 입력해주시기 바랍니다.")
                    elif self.tempPoint1[5:7] == '00' or self.tempPoint1[8:10] == '00' or self.tempPoint2[
                                                                                          5:7] == '00' or self.tempPoint2[
                                                                                                          8:10] == '00':
                        self.alertbox_open4("해당 월일을 올바르게 입력해주시기 바랍니다.")
                    else:
                        self.alertbox_open4("시점은 'yyyy-mm-dd'의 형태로 입력해주시기 바랍니다.")

            except ValueError:
                self.alertbox_open4("중요성금액 값을 숫자로만 입력해주시기 바랍니다.")

    def Thread14(self):
        self.baseKey = self.D14_Key.text().split(',')
        self.baseKey_clean = []
        for a in self.baseKey:
            a = a.strip()
            self.baseKey_clean.append(a)
        self.tempKey = str('%'.join(self.baseKey_clean))
        self.tempTE = self.D14_TE.text()
        self.tempSheet = self.D14_Sheet.text()

        if self.tempSheet == '':
            self.alertbox_open()
        # 시트명 중복 확인
        elif self.rbtn1.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Result') != -1:
            self.alertbox_open5()

        elif self.rbtn2.isChecked() and self.combo_sheet.findText(self.tempSheet + '_Journals') != -1:
            self.alertbox_open5()

        elif not (self.checkC.isChecked()) and not (self.checkD.isChecked()):
            self.alertbox_open7()

        else:
            if self.tempTE == '': self.tempTE = 0
            try:
                int(self.tempTE)

                if checked_account == 'AND JournalEntries.GLAccountNumber IN ()':
                    self.checked_account14 = ''
                else:
                    self.checked_account14 = checked_account

                self.doAction()
                self.th14 = Thread(target=self.extButtonClicked14)
                self.th14.daemon = True
                self.th14.start()
            except ValueError:
                try:
                    int(self.tempTE)
                except:
                    self.alertbox_open4('중요성금액 값을 숫자로만 입력해주시기 바랍니다.')

    def Sheet_ComboBox_Selected(self, text):

        model = DataFrameModel(self.scenario_dic[text])
        self.viewtable.setModel(model)

    def RemoveSheetButton_Clicked(self):

        ##예외 처리 - 삭제할 Sheet가 없는 경우
        if not self.scenario_dic:
            self.MessageBox_Open("삭제할 Sheet가 없습니다")
            return

        ##Sheet 정보 삭제
        del self.scenario_dic[self.combo_sheet.currentText()]
        ##Query 정보 삭제
        self.my_query.drop(labels=[self.combo_sheet.currentText()], axis=0, inplace=True)
        ##Sheet name 콤보박스에서 해당 sheet 삭제
        self.combo_sheet.removeItem(self.combo_sheet.currentIndex())
        gc.collect()

        if not self.scenario_dic:
            self.dataframe = pd.DataFrame({'No Sheet': []})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)
        else:
            model = DataFrameModel(self.scenario_dic[self.combo_sheet.currentText()])
            self.viewtable.setModel(model)

    def Save_Buttons_Group(self):
        ##GroupBox
        groupbox = QGroupBox("저장")
        font_groupbox = groupbox.font()
        font_groupbox.setBold(True)
        groupbox.setFont(font_groupbox)
        self.setStyleSheet('QGroupBox  {color: white;}')

        ##RemoveSheet 버튼
        RemoveSheet_button = QPushButton('Remove Sheet')
        RemoveSheet_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        RemoveSheet_button.setStyleSheet('color:white;background-image : url(./bar.png)')
        font_RemoveSheet = RemoveSheet_button.font()
        font_RemoveSheet.setBold(True)
        RemoveSheet_button.setFont(font_RemoveSheet)

        # label
        label_sheet = QLabel("Sheet names: ", self)
        font_sheet = label_sheet.font()
        font_sheet.setBold(True)
        label_sheet.setFont(font_sheet)
        label_sheet.setStyleSheet('color:white;')

        ##시나리오 Sheet를 표현할 콤보박스
        self.combo_sheet = QComboBox(self)

        ## Save 버튼
        export_file_button = QPushButton("Save", self)
        export_file_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        font_export_button = export_file_button.font()
        font_export_button.setBold(True)
        export_file_button.setFont(font_export_button)
        export_file_button.setStyleSheet('color:white;background-image : url(./bar.png)')

        #########
        #########버튼 클릭 or 콤보박스 선택시 발생하는 시그널 함수들
        RemoveSheet_button.clicked.connect(self.RemoveSheetButton_Clicked)
        export_file_button.clicked.connect(self.saveFile)
        self.combo_sheet.activated[str].connect(self.Sheet_ComboBox_Selected)

        ##layout 쌓기
        layout = QHBoxLayout()
        layout.addWidget(label_sheet, stretch=1)
        layout.addWidget(self.combo_sheet, stretch=4)
        layout.addWidget(RemoveSheet_button, stretch=1)
        layout.addWidget(export_file_button, stretch=1)
        groupbox.setLayout(layout)

        return groupbox

    def calendar6(self):
        self.dialog6.activateWindow()
        self.new_calendar.show()
        self.dialog6.activateWindow()

    def calendar7(self):
        self.dialog7.activateWindow()
        self.new_calendar.show()
        self.dialog7.activateWindow()

    def calendar10_1(self):
        self.dialog10.activateWindow()
        self.new_calendar1.show()

    def calendar10_2(self):
        self.dialog10.activateWindow()
        self.new_calendar2.show()

    def extButtonClicked4(self):
        cursor = self.cnxn.cursor()

        ### JE Line - Result
        if self.rbtn1.isChecked():

            sql_refer = """
                                SELECT				
                                JournalEntries.BusinessUnit			
                                , JournalEntries.JENumber			
                                , JournalEntries.JELineNumber			
                                , JournalEntries.EffectiveDate			
                                , JournalEntries.EntryDate			
                                , JournalEntries.Period			
                                , JournalEntries.GLAccountNumber			
                                , CoA.GLAccountName			
                                , JournalEntries.Debit			
                                , JournalEntries.Credit			
                                , CASE
                                       WHEN JournalEntries.Debit = 0 THEN 'Credit' ELSE 'Debit'
                                       END AS DebitCredit
                                , JournalEntries.Amount			
                                , JournalEntries.FunctionalCurrencyCode			
                                , JournalEntries.JEDescription			
                                , JournalEntries.JELineDescription			
                                , JournalEntries.PreparerID			
                                , JournalEntries.ApproverID			
                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries
							INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
							ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber
                            WHERE EXISTS				
                                (			
                                SELECT DISTINCT GLAccountNumber			
                                FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS Journal
								WHERE JournalEntries.GLAccountNumber = Journal.GLAccountNumber
                                GROUP BY GLAccountNumber			
                                HAVING COUNT(GLAccountNumber) <= {N}			
                                ) AND ABS(JournalEntries.Amount) > {TE}
                                {Account}
							ORDER BY JENumber,JELineNumber
                            """.format(field=self.selected_project_id, TE=self.temp_TE, N=self.temp_N,
                                       Account=self.checked_account4)

            ### JE Line - Refer
            sql_query = '''
                            SELECT				
                                JournalEntries.BusinessUnit			
                                , JournalEntries.JENumber			
                                , JournalEntries.JELineNumber			
                                , JournalEntries.EffectiveDate			
                                , JournalEntries.EntryDate			
                                , JournalEntries.Period			
                                , JournalEntries.GLAccountNumber			
                                , CoA.GLAccountName			
                                , JournalEntries.Debit			
                                , JournalEntries.Credit			
                                , CASE
                                       WHEN JournalEntries.Debit = 0 THEN 'Credit' ELSE 'Debit'
                                       END AS DebitCredit
                                , JournalEntries.Amount			
                                , JournalEntries.FunctionalCurrencyCode			
                                , JournalEntries.JEDescription			
                                , JournalEntries.JELineDescription			
                                , JournalEntries.PreparerID			
                                , JournalEntries.ApproverID			
                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries
                            INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
							ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber
                            WHERE EXISTS 				
                                (			
                                SELECT DISTINCT GLAccountNumber			
                                FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS Journal
                                WHERE JournalEntries.GLAccountNumber = Journal.GLAccountNumber
                                GROUP BY GLAccountNumber			
                                HAVING COUNT(GLAccountNumber) <= {N}			
                                ) AND ABS(JournalEntries.Amount) > {TE}
                                {Account}			
                            ORDER BY JENumber,JELineNumber				
                        '''.format(field=self.selected_project_id, TE=self.temp_TE, N=self.temp_N,
                                   Account=self.checked_account4)

            self.dataframe_refer = pd.read_sql(sql_refer, self.cnxn)

        ### JE - Journals
        elif self.rbtn2.isChecked():
            sql_query = '''
            SELECT
                             JournalEntries.BusinessUnit			
                            , JournalEntries.JENumber			
                            , JournalEntries.JELineNumber			
                            , JournalEntries.EffectiveDate			
                            , JournalEntries.EntryDate			
                            , JournalEntries.Period			
                            , JournalEntries.GLAccountNumber			
                            , CoA.GLAccountName			
                            , JournalEntries.Debit			
                            , JournalEntries.Credit			
                            , CASE
                                WHEN JournalEntries.Debit = 0 THEN 'Credit' ELSE 'Debit'
                                END AS DebitCredit
                            , JournalEntries.Amount			
                            , JournalEntries.FunctionalCurrencyCode			
                            , JournalEntries.JEDescription			
                            , JournalEntries.JELineDescription			
                            , JournalEntries.PreparerID			
                            , JournalEntries.ApproverID			
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries				
                        INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA			
                        ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber 
						WHERE EXISTS (				
                            SELECT DISTINCT JournalEntries.JENumber			
                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS Journal
							WHERE JournalEntries.JENumber = Journal.JENumber AND EXISTS 			
                                    (	
                                    SELECT DISTINCT Journal2.GLAccountNumber	
                                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS Journal2,	
                                        [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                                    WHERE Journal2.GLAccountNumber = CoA.GLAccountNumber AND Journal.GLAccountNumber = Journal2.GLAccountNumber	
                                    GROUP BY Journal2.GLAccountNumber	
                                    HAVING COUNT(Journal2.GLAccountNumber) <= {N}	
                                    ) AND ABS(Journal.Amount) > {TE}
                                    {Account}
                                   )
                                
                        ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber
                       
                '''.format(field=self.selected_project_id, TE=self.temp_TE, N=self.temp_N,
                           Account=self.checked_account4)

        self.dataframe = pd.read_sql(sql_query, self.cnxn)

        ### 마지막 시트 쿼리 내역 추가
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario04",
                                                             "---Filtered Result_2  Scenario04---\n" + sql_query]

            self.my_query.loc[self.tempSheet + "_Reference"] = [self.tempSheet + "_Reference", "Scenario04",
                                                                "---Filtered Result_1  Scenario04---\n" + sql_refer]

        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario04",
                                                               "---Filtered JE  Scenario04---\n" + sql_query]

        ### 차대 선택
        if not self.checkD.isChecked() and self.checkC.isChecked():
            self.dataframe = self.dataframe[self.dataframe['DebitCredit'] == 'Credit']
            self.dataframe.reset_index(drop=True, inplace=True)

        elif self.checkD.isChecked() and not self.checkC.isChecked():
            self.dataframe = self.dataframe[self.dataframe['DebitCredit'] == 'Debit']
            self.dataframe.reset_index(drop=True, inplace=True)

        ### 최대 추출 라인수
        if len(self.dataframe) > 1048576:
            self.communicate4.closeApp.emit()

        elif len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame({'No Data': ["[계정사용 빈도수: " + str(self.temp_N) + ","
                                                       + "중요성금액: " + str(self.temp_TE)
                                                       + '] 라인 수 ' + str(len(self.dataframe)) + '개입니다']})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)
            ### JE Line
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.scenario_dic[self.tempSheet + '_Reference'] = self.dataframe_refer
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.addItem(self.tempSheet + '_Reference')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 2)

            ### JE
            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            self.communicate4.closeApp.emit()

        else:
            ### JE Line
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.scenario_dic[self.tempSheet + '_Reference'] = self.dataframe_refer
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.addItem(self.tempSheet + '_Reference')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 2)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            ### JE
            elif self.rbtn2.isChecked():
                ### 시트 콤보박스에 저장
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            self.communicate4.closeApp.emit()

    def extButtonClicked5_SAP(self):

        ### 쿼리 연동
        cursor = self.cnxn.cursor()

        ### JE Line 선택시 - 추출 조건에 해당하는
        if self.rbtn1.isChecked():

            sql_query = """
                            SELECT 
                                JournalEntries.BusinessUnit
                                , JournalEntries.JENumber
                                , JournalEntries.JELineNumber
                                , JournalEntries.EffectiveDate
                                , JournalEntries.EntryDate
                                , JournalEntries.Period	
                                , JournalEntries.GLAccountNumber
                                , CoA.GLAccountName
                                , JournalEntries.Debit
                                , JournalEntries.Credit
                                , CASE
                                    WHEN JournalEntries.Debit = 0 THEN 'Credit' ELSE 'Debit'
                                    END AS DebitCredit
                                , JournalEntries.Amount
                                , JournalEntries.FunctionalCurrencyCode
                                , JournalEntries.JEDescription
                                , JournalEntries.JELineDescription
                                , JournalEntries.PreparerID
                                , JournalEntries.ApproverID
                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries  
                            INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
							ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber
                            WHERE JournalEntries.GLAccountNumber IN ({CODE})
                                    {Account}
                            ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber	
                        """.format(field=self.selected_project_id, CODE=self.real_Code,
                                   Account=self.checked_account5_SAP)

        elif self.rbtn2.isChecked():
            ####수정사항 시작####
            sql_query = '''
                                SELECT
                                    JournalEntries.BusinessUnit
                                    , JournalEntries.JENumber
                                    , JournalEntries.JELineNumber
                                    , JournalEntries.EffectiveDate
                                    , JournalEntries.EntryDate
                                    , JournalEntries.Period
                                    , JournalEntries.GLAccountNumber
                                    , CoA.GLAccountName
                                    , JournalEntries.Debit
                                    , JournalEntries.Credit
                                    , CASE
                                           WHEN JournalEntries.Debit = 0 THEN 'Credit' ELSE 'Debit'
                                           END AS DebitCredit
                                    , JournalEntries.Amount
                                    , JournalEntries.FunctionalCurrencyCode
                                    , JournalEntries.JEDescription
                                    , JournalEntries.JELineDescription
                                    , JournalEntries.PreparerID
                                    , JournalEntries.ApproverID
                                FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries  
                                INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                                ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber
                                WHERE EXISTS
                                    (
                                    SELECT DISTINCT JENumber
                                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS Journal
                                    WHERE JournalEntries.JENumber = Journal.JENumber
                                    AND GLAccountNumber IN ({CODE})
                                    {Account}
                                    )
                                ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber
                            '''.format(field=self.selected_project_id, CODE=self.real_Code,
                                       Account=self.checked_account5_SAP)
        ####수정사항 끝####
        self.dataframe = pd.read_sql(sql_query, self.cnxn)
        print(sql_query)

        ### 마지막 시트 쿼리 내역 추가
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet_SAP + "_Result"] = [self.tempSheet_SAP + "_Result",
                                                                 "Scenario05",
                                                                 "---Filtered Result  Scenario05---\n" + sql_query]

        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet_SAP + "_Journals"] = [self.tempSheet_SAP + "_Journals",
                                                                   "Scenario05",
                                                                   "---Filtered JE  Scenario05---\n" + sql_query]

        ### DebitCredit 열 생성
        if not self.checkD2.isChecked() and self.checkC2.isChecked():
            self.dataframe = self.dataframe[self.dataframe['DebitCredit'] == 'Credit']
            self.dataframe.reset_index(drop=True, inplace=True)

        elif self.checkD2.isChecked() and not self.checkC2.isChecked():
            self.dataframe = self.dataframe[self.dataframe['DebitCredit'] == 'Debit']
            self.dataframe.reset_index(drop=True, inplace=True)

        ### 예외처리 5 - 최대 라인 수 초과
        if len(self.dataframe) > 1048576:
            self.communicate5_SAP.closeApp.emit()

        ### 조건, 라인 수 추출
        elif len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame({'No Data': ["[계정코드: " + str(self.temp_AccCode) + "," +
                                                       "연도" + str(self.tempYear_SAP) + "] 라인수 " +
                                                       str(len(self.dataframe)) + "개 입니다"]})

            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet_SAP + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet_SAP + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet_SAP + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet_SAP + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            self.communicate5_SAP.closeApp.emit()

        else:
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet_SAP + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet_SAP + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet_SAP + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet_SAP + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            self.communicate5_SAP.closeApp.emit()

    def extButtonClicked5_Non_SAP(self):

        ### 쿼리 연동
        cursor = self.cnxn.cursor()

        ### JE Line
        if self.rbtn1.isChecked():

            sql_query = """
                            SELECT 
                                JournalEntries.BusinessUnit
                                , JournalEntries.JENumber
                                , JournalEntries.JELineNumber
                                , JournalEntries.EffectiveDate
                                , JournalEntries.EntryDate
                                , JournalEntries.Period    
                                , JournalEntries.GLAccountNumber
                                , CoA.GLAccountName
                                , JournalEntries.Debit
                                , JournalEntries.Credit
                                , CASE
                                    WHEN JournalEntries.Debit = 0 THEN "Credit" ELSE "Debit"
                                    END AS DebitCredit
                                , JournalEntries.Amount
                                , JournalEntries.FunctionalCurrencyCode
                                , JournalEntries.JEDescription
                                , JournalEntries.JELineDescription
                                , JournalEntries.PreparerID
                                , JournalEntries.ApproverID
                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries
                            INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                                ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber
                            WHERE JournalEntries.GLAccountNumber IN ({CODE})
                                    {Account}
                            ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber  
                        """.format(field=self.selected_project_id, CODE=self.temp_Code_Non_SAP,
                                   Account=self.checked_account5_Non)
        ### JE
        ####수정사항 시작####
        elif self.rbtn2.isChecked():

            sql_query = '''
                                SELECT
                                    JournalEntries.BusinessUnit
                                    , JournalEntries.JENumber
                                    , JournalEntries.JELineNumber
                                    , JournalEntries.EffectiveDate
                                    , JournalEntries.EntryDate
                                    , JournalEntries.Period
                                    , JournalEntries.GLAccountNumber
                                    , CoA.GLAccountName
                                    , JournalEntries.Debit
                                    , JournalEntries.Credit
                                    , CASE
                                           WHEN JournalEntries.Debit = 0 THEN 'Credit' ELSE 'Debit'
                                           END AS DebitCredit
                                    , JournalEntries.Amount
                                    , JournalEntries.FunctionalCurrencyCode
                                    , JournalEntries.JEDescription
                                    , JournalEntries.JELineDescription
                                    , JournalEntries.PreparerID
                                    , JournalEntries.ApproverID
                                FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries
                                INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                                ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber
                                WHERE JournalEntries.JENumber IN	
                                    (
                                    SELECT DISTINCT JENumber
                                    FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries]
                                    WHERE GLAccountNumber IN ({CODE})
                                    {Account}
                                    )
                                ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber
                        '''.format(field=self.selected_project_id, CODE=self.temp_Code_Non_SAP,
                                   Account=self.checked_account5_Non)
        ####수정사항 끝####
        self.dataframe = pd.read_sql(sql_query, self.cnxn)

        ### 마지막 시트 쿼리 내역 추가
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet_NonSAP + "_Result"] = [self.tempSheet_NonSAP + "_Result", "Scenario05",
                                                                    "---Filtered Result  Scenario05---\n" + sql_query]

        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet_NonSAP + "_Journals"] = [self.tempSheet_NonSAP + "_Journals",
                                                                      "Scenario05",
                                                                      "---Filtered JE  Scenario05---\n" + sql_query]

        ### DebitCredit 열 생성
        if not self.checkD1.isChecked() and self.checkC1.isChecked():
            self.dataframe = self.dataframe[self.dataframe['DebitCredit'] == 'Credit']
            self.dataframe.reset_index(drop=True, inplace=True)

        elif self.checkD1.isChecked() and not self.checkC1.isChecked():
            self.dataframe = self.dataframe[self.dataframe['DebitCredit'] == 'Debit']
            self.dataframe.reset_index(drop=True, inplace=True)

        ### 예외처리 5 - 최대 출력 라인 초과
        if len(self.dataframe) > 1048576:
            self.communicate5_Non_SAP.closeApp.emit()

        ### 예외처리 6 - 데이터 미추출
        elif len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame({'No Data': ['[연도: ' + str(self.tempYear_NonSAP) + ','
                                                       + '계정코드: ' + str(self.temp_Code_Non_SAP) + ','
                                                       + '] 라인수 ' + str(len(self.dataframe)) + '개 입니다']})

            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)

            ### JE Line
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet_NonSAP + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet_NonSAP + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            ### JE
            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet_NonSAP + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet_NoneSAP + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            self.communicate5_Non_SAP.closeApp.emit()

        else:
            ### JE Line
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet_NonSAP + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet_NonSAP + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            ### JE
            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet_NonSAP + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet_NonSAP + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            self.communicate5_Non_SAP.closeApp.emit()

    def extButtonClicked6(self):

        cursor = self.cnxn.cursor()

        if self.rbtn1.isChecked():
            sql = '''
                       SELECT											
                           JournalEntries.BusinessUnit											
                           , JournalEntries.JENumber											
                           , JournalEntries.JELineNumber											
                           , JournalEntries.EffectiveDate											
                           , JournalEntries.EntryDate											
                           , JournalEntries.Period											
                           , JournalEntries.GLAccountNumber											
                           , CoA.GLAccountName											
                           , JournalEntries.Debit											
                           , JournalEntries.Credit											
                           , CASE
                                WHEN JournalEntries.Debit = 0 THEN 'Credit' ELSE 'Debit'
                                END AS DebitCredit
                           , JournalEntries.Amount											
                           , JournalEntries.FunctionalCurrencyCode											
                           , JournalEntries.JEDescription											
                           , JournalEntries.JELineDescription											
                           , JournalEntries.PreparerID											
                           , JournalEntries.ApproverID											
                       FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries
                       INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                                ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber
                       WHERE (JournalEntries.EntryDate BETWEEN {first_date} AND {second_date})
                               AND ABS(JournalEntries.Amount) > {TE}
                               {Preparer}
                               {Account}
                       ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber
                    '''.format(field=self.selected_project_id, Account=self.checked_account6, TE=self.tempCost,
                               first_date=str(self.first), second_date=str(self.second),
                               Preparer=self.checked_preparer6)


        elif self.rbtn2.isChecked():
            sql = '''
                       SELECT											
                           JournalEntries.BusinessUnit											
                           , JournalEntries.JENumber											
                           , JournalEntries.JELineNumber											
                           , JournalEntries.EffectiveDate											
                           , JournalEntries.EntryDate											
                           , JournalEntries.Period											
                           , JournalEntries.GLAccountNumber											
                           , CoA.GLAccountName											
                           , JournalEntries.Debit											
                           , JournalEntries.Credit											
                           , CASE
                                WHEN JournalEntries.Debit = 0 THEN 'Credit' ELSE 'Debit'
                                END AS DebitCredit
                           , JournalEntries.Amount											
                           , JournalEntries.FunctionalCurrencyCode											
                           , JournalEntries.JEDescription											
                           , JournalEntries.JELineDescription											
                           , JournalEntries.PreparerID											
                           , JournalEntries.ApproverID											
                       FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries											
                       INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                       ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber
                       WHERE EXISTS		
                        (
                            SELECT DISTINCT JENumber
                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS Journal
                            WHERE JournalEntries.JENumber = Journal.JENumber
                            AND (EntryDate BETWEEN {first_date} AND {second_date})
                                    {Account}
                                    {Preparer}
                                    AND ABS(Amount) > {TE}
                        )
                        ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber		
                    '''.format(field=self.selected_project_id, Account=self.checked_account6, TE=self.tempCost,
                               first_date=str(self.first), second_date=str(self.second),
                               Preparer=self.checked_preparer6)

        self.dataframe = pd.read_sql(sql, self.cnxn)

        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario06",
                                                             "---Filtered Result  Scenario06---\n" + sql]

        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario06",
                                                               "---Filtered JE  Scenario06---\n" + sql]

        ### DebitCredit 열 생성
        if not self.checkD.isChecked() and self.checkC.isChecked():
            self.dataframe = self.dataframe[self.dataframe['DebitCredit'] == 'Credit']
            self.dataframe.reset_index(drop=True, inplace=True)

        elif self.checkD.isChecked() and not self.checkC.isChecked():
            self.dataframe = self.dataframe[self.dataframe['DebitCredit'] == 'Debit']
            self.dataframe.reset_index(drop=True, inplace=True)

        if len(self.dataframe) > 1048576:
            self.communicate6.closeApp.emit()

        elif len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame(
                {'No Data': ["[결산일: " + str(self.tempDate) + "," + "T일: " + str(int(self.tempTDate))
                             + "," + "중요성금액: " + str(self.tempCost)
                             + "] 라인수 " + str(len(self.dataframe)) + "개입니다"]})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            self.communicate6.closeApp.emit()

        else:
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            self.communicate6.closeApp.emit()

    def extButtonClicked7(self):

        cursor = self.cnxn.cursor()

        if self.rbtn3.isChecked():
            sql = '''
                       SELECT 											
                           JournalEntries.BusinessUnit											
                           , JournalEntries.JENumber											
                           , JournalEntries.JELineNumber											
                           , JournalEntries.EffectiveDate											
                           , JournalEntries.EntryDate											
                           , JournalEntries.Period											
                           , JournalEntries.GLAccountNumber											
                           , CoA.GLAccountName											
                           , JournalEntries.Debit											
                           , JournalEntries.Credit											
                           , CASE
                                WHEN JournalEntries.Debit = 0 THEN 'Credit' ELSE 'Debit'
                                END AS DebitCredit
                           , JournalEntries.Amount											
                           , JournalEntries.FunctionalCurrencyCode											
                           , JournalEntries.JEDescription											
                           , JournalEntries.JELineDescription											
                           , JournalEntries.Source											
                           , JournalEntries.PreparerID											
                           , JournalEntries.ApproverID											
                       FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries
                       INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                                ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber
                       WHERE ABS(JournalEntries.Amount) > {TE}
                            {Date}
                            {Account}
                            {Preparer}
                       ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber
                   '''.format(field=self.selected_project_id, TE=self.tempCost, Date=self.tempState,
                              Account=self.checked_account7, Preparer=self.checked_preparer7)

        elif self.rbtn4.isChecked():
            sql = '''
                       SELECT 											
                           JournalEntries.BusinessUnit											
                           , JournalEntries.JENumber											
                           , JournalEntries.JELineNumber											
                           , JournalEntries.EffectiveDate											
                           , JournalEntries.EntryDate											
                           , JournalEntries.Period											
                           , JournalEntries.GLAccountNumber											
                           , CoA.GLAccountName											
                           , JournalEntries.Debit											
                           , JournalEntries.Credit											
                           , CASE
                                WHEN JournalEntries.Debit = 0 THEN 'Credit' ELSE 'Debit'
                                END AS DebitCredit
                           , JournalEntries.Amount											
                           , JournalEntries.FunctionalCurrencyCode											
                           , JournalEntries.JEDescription											
                           , JournalEntries.JELineDescription											
                           , JournalEntries.Source											
                           , JournalEntries.PreparerID											
                           , JournalEntries.ApproverID											
                       FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries
                       INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                                ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber
                       
                       WHERE EXISTS(		
                           SELECT DISTINCT JENumber
                           FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS Journal
                           WHERE ABS(Journal.Amount) > {TE}
                                AND JournalEntries.JENumber = Journal.JENumber
                                {Account}
                                {Preparer}
                                {Date})
                       ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber
                   '''.format(field=self.selected_project_id, TE=self.tempCost, Date=self.tempState,
                              Account=self.checked_account7, Preparer=self.checked_preparer7)

        self.dataframe = pd.read_sql(sql, self.cnxn)

        ### 마지막 시트 쿼리 내역 추가
        if self.rbtn3.isChecked():
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario07",
                                                             "---Filtered Result  Scenario07---\n" + sql]

        elif self.rbtn4.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario07",
                                                               "---Filtered JE  Scenario07---\n" + sql]

        ### DebitCredit 열 생성
        if not self.checkD.isChecked() and self.checkC.isChecked():
            self.dataframe = self.dataframe[self.dataframe['DebitCredit'] == 'Credit']
            self.dataframe.reset_index(drop=True, inplace=True)

        elif self.checkD.isChecked() and not self.checkC.isChecked():
            self.dataframe = self.dataframe[self.dataframe['DebitCredit'] == 'Debit']
            self.dataframe.reset_index(drop=True, inplace=True)

        if len(self.dataframe) > 1048576:
            self.communicate7.closeApp.emit()

        elif len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame({'No Data': ["[EffectiveDate/EntryDate: " + str(self.tempState) + ","
                                                       + "중요성금액: " + str(self.tempCost)
                                                       + "] 라인수 " + str(len(self.dataframe)) + "개입니다"]})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)

            if self.rbtn3.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn4.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            self.communicate7.closeApp.emit()

        else:

            if self.rbtn3.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)


            elif self.rbtn4.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)
            self.communicate7.closeApp.emit()

    def extButtonClicked8(self):

        cursor = self.cnxn.cursor()

        if self.rbtn1.isChecked():

            sql = '''
                            SELECT 
                               JournalEntries.FileId	
                                , JournalEntries.BusinessUnit	
                                , JournalEntries.JENumber	
                                , JournalEntries.JELineNumber	
                                , JournalEntries.EffectiveDate	
                                , JournalEntries.EntryDate	
                                , JournalEntries.Period	
                                , JournalEntries.GLAccountNumber	
                                , CoA.GLAccountName	
                                , JournalEntries.Debit	
                                , JournalEntries.Credit	
                                , CASE
                                    WHEN JournalEntries.Debit = 0 THEN 'Credit' ELSE 'Debit'
                                    END AS DebitCredit
                                , JournalEntries.Amount	
                                , JournalEntries.FunctionalCurrencyCode	
                                , JournalEntries.JEDescription	
                                , JournalEntries.JELineDescription	
                                , JournalEntries.PreparerID	
                                , JournalEntries.ApproverID	
                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries
                            INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                            ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber
                            WHERE DATEDIFF(dd, JournalEntries.EntryDate ,JournalEntries.EffectiveDate) >= {realNDate}
                                {Preparer}
                                AND ABS(JournalEntries.Amount) > {TE}
                                {Account}
                            ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber
                        '''.format(field=self.selected_project_id, realNDate=self.realNDate, TE=self.tempCost,
                                   Preparer=self.checked_preparer8, Account=self.checked_account8)

        elif self.rbtn2.isChecked():

            sql = '''
                            SELECT 
                               JournalEntries.FileId	
                                , JournalEntries.BusinessUnit	
                                , JournalEntries.JENumber	
                                , JournalEntries.JELineNumber	
                                , JournalEntries.EffectiveDate	
                                , JournalEntries.EntryDate	
                                , JournalEntries.Period	
                                , JournalEntries.GLAccountNumber	
                                , CoA.GLAccountName	
                                , JournalEntries.Debit	
                                , JournalEntries.Credit	
                                , CASE
                                    WHEN JournalEntries.Debit = 0 THEN 'Credit' ELSE 'Debit'
                                    END AS DebitCredit
                                , JournalEntries.Amount	
                                , JournalEntries.FunctionalCurrencyCode	
                                , JournalEntries.JEDescription	
                                , JournalEntries.JELineDescription	
                                , JournalEntries.PreparerID	
                                , JournalEntries.ApproverID	
                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries
                            INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                            ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber
                            WHERE EXISTS  
                                (
                                SELECT DISTINCT JENumber
                                FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS Journal
                                WHERE DATEDIFF(dd, JournalEntries.EntryDate ,JournalEntries.EffectiveDate) >= {realNDate}
                                    {Preparer}
                                    AND ABS(JournalEntries.Amount) > {TE}
                                    {Account}
                                    AND JournalEntries.JENumber = Journal.JENumber
                                )
                            ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber
                            '''.format(field=self.selected_project_id, realNDate=self.realNDate, TE=self.tempCost,
                                       Preparer=self.checked_preparer8, Account=self.checked_account8)

        self.dataframe = pd.read_sql(sql, self.cnxn)

        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario08",
                                                             "---Filtered Result  Scenario08---\n" + sql]

        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario08",
                                                               "---Filtered JE  Scenario08---\n" + sql]

        ### DebitCredit 열 생성
        if not self.checkD.isChecked() and self.checkC.isChecked():
            self.dataframe = self.dataframe[self.dataframe['DebitCredit'] == 'Credit']
            self.dataframe.reset_index(drop=True, inplace=True)

        elif self.checkD.isChecked() and not self.checkC.isChecked():
            self.dataframe = self.dataframe[self.dataframe['DebitCredit'] == 'Debit']
            self.dataframe.reset_index(drop=True, inplace=True)

        if len(self.dataframe) > 1048576:
            self.communicate8.closeApp.emit()

        elif len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame(
                {'No Data': ["[Effective Date와 Entry Date 간 차이: " + str(int(self.realNDate))
                             + "," + "중요성금액: " + str(self.tempCost)
                             + "] 라인수 " + str(len(self.dataframe)) + "개입니다"]})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
            self.communicate8.closeApp.emit()

        else:

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            self.communicate8.closeApp.emit()

    def extButtonClicked9(self):

        cursor = self.cnxn.cursor()
        # sql문 수정
        if self.rbtn1.isChecked():
            sql = '''
                           SELECT				
                                  JournalEntries.BusinessUnit			
                                  , JournalEntries.JENumber			
                                  , JournalEntries.JELineNumber			
                                  , JournalEntries.EffectiveDate			
                                  , JournalEntries.EntryDate			
                                  , JournalEntries.Period			
                                  , JournalEntries.GLAccountNumber			
                                  , CoA.GLAccountName			
                                  , JournalEntries.Debit			
                                  , JournalEntries.Credit			
                                  , CASE
                                        WHEN JournalEntries.Debit = 0 THEN 'Credit' ELSE 'Debit'
                                        END AS DebitCredit
                                  , JournalEntries.Amount			
                                  , JournalEntries.FunctionalCurrencyCode			
                                  , JournalEntries.JEDescription			
                                  , JournalEntries.JELineDescription			
                                  , JournalEntries.PreparerID			
                                  , JournalEntries.ApproverID			
                           FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries
                           INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                           ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber
                           WHERE EXISTS
                                  (			
                                  SELECT DISTINCT PreparerID			
                                  FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS Journal
                                  WHERE JournalEntries.PreparerID = Journal.PreparerID
                                  GROUP BY PreparerID			
                                  HAVING COUNT(GLAccountNumber) <= {N}			
                                  )	AND ABS(JournalEntries.Amount) > {TE} {Account}		
                           ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber				
                        '''.format(field=self.selected_project_id, TE=self.tempTE, N=self.tempN,
                                   Account=self.checked_account9)

            sql_refer = '''
                           SELECT JournalEntries.PreparerID, COUNT(JournalEntries.PreparerID) AS User_Cnt, SUM(Debit) Sum_of_Debit, SUM(Credit) Sum_of_Credit				
                           FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries				
                           WHERE EXISTS				
                                  (			
                                  SELECT DISTINCT PreparerID			
                                  FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS Journal
                                  WHERE JournalEntries.PreparerID = Journal.PreparerID
                                  GROUP BY PreparerID			
                                  HAVING COUNT(GLAccountNumber) <= {N}			
                                  ) AND ABS(JournalEntries.Amount) > {TE} {Account}			
                           GROUP BY JournalEntries.PreparerID				
                        '''.format(field=self.selected_project_id, TE=self.tempTE, N=self.tempN,
                                   Account=self.checked_account9)
            self.dataframe_refer = pd.read_sql(sql_refer, self.cnxn)

        elif self.rbtn2.isChecked():
            sql = '''
            SELECT 				
                               JournalEntries.BusinessUnit			
                               , JournalEntries.JENumber			
                               , JournalEntries.JELineNumber			
                               , JournalEntries.EffectiveDate			
                               , JournalEntries.EntryDate			
                               , JournalEntries.Period			
                               , JournalEntries.GLAccountNumber			
                               , CoA.GLAccountName			
                               , JournalEntries.Debit			
                               , JournalEntries.Credit			
                               , CASE
                                    WHEN JournalEntries.Debit = 0 THEN 'Credit' ELSE 'Debit'
                                    END AS DebitCredit
                               , JournalEntries.Amount			
                               , JournalEntries.FunctionalCurrencyCode			
                               , JournalEntries.JEDescription			
                               , JournalEntries.JELineDescription			
                               , JournalEntries.PreparerID			
                               , JournalEntries.ApproverID			
                           FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries
                           INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                           ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber
                           WHERE JournalEntries.GLAccountNumber = CoA.GLAccountNumber AND 				
                               JournalEntries.JENumber IN 			
                                   (		
                                   SELECT DISTINCT JENumber		
                                   FROM  [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries		
                                   WHERE JournalEntries.PreparerID IN 		
                                       (	
                                       SELECT DISTINCT JournalEntries.PreparerID	
                                       FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries,	
                                           (
                                           SELECT JournalEntries.PreparerID, COUNT(JournalEntries.PreparerID) AS User_Cnt
                                           FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries
                                           GROUP BY JournalEntries.PreparerID
                                           HAVING COUNT(GLAccountNumber) <= {N}
                                           ) AS LVL1
                                       WHERE LVL1.PreparerID = JournalEntries.PreparerID	
                                       ) AND ABS(JournalEntries.Amount) > {TE}	
                                   ) {Account} 		
                           ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber
                        '''.format(field=self.selected_project_id, TE=self.tempTE, N=self.tempN,
                                   Account=self.checked_account9)

        self.dataframe = pd.read_sql(sql, self.cnxn)

        ### 마지막 시트 쿼리 내역 추가
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario09",
                                                             "---Filtered Result_2  Scenario09---\n" + sql]
            self.my_query.loc[self.tempSheet + "_Reference"] = [self.tempSheet + "_Reference", "Scenario09",
                                                                "---Filtered Result_1  Scenario09---\n" + sql_refer]

        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario09",
                                                               "---Filtered JE  Scenario09---\n" + sql]

        ### DebitCredit 열 생성
        if not self.checkD.isChecked() and self.checkC.isChecked():
            self.dataframe = self.dataframe[self.dataframe['DebitCredit'] == 'Credit']
            self.dataframe.reset_index(drop=True, inplace=True)

        elif self.checkD.isChecked() and not self.checkC.isChecked():
            self.dataframe = self.dataframe[self.dataframe['DebitCredit'] == 'Debit']
            self.dataframe.reset_index(drop=True, inplace=True)

        if len(self.dataframe) > 1048576:
            self.communicate9.closeApp.emit()

        elif len(self.dataframe) == 0:
            self.communicate9.closeApp.emit()

        else:
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.scenario_dic[self.tempSheet + '_Reference'] = self.dataframe_refer
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.addItem(self.tempSheet + '_Reference')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 2)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)
            self.communicate9.closeApp.emit()

    def extButtonClicked10(self):

        cursor = self.cnxn.cursor()

        # sql문 수정
        if self.rbtn1.isChecked():

            sql = '''
                                 SELECT			
                                       JournalEntries.BusinessUnit		
                                       , JournalEntries.JENumber		
                                       , JournalEntries.JELineNumber		
                                       , JournalEntries.EffectiveDate		
                                       , JournalEntries.EntryDate
                                       , JournalEntries.Period		
                                       , JournalEntries.GLAccountNumber		
                                       , CoA.GLAccountName		
                                       , JournalEntries.Debit		
                                       , JournalEntries.Credit		
                                       , CASE
                                                WHEN JournalEntries.Debit = 0 THEN 'Credit' ELSE 'Debit'
                                                END AS DebitCredit
                                       , JournalEntries.Amount		
                                       , JournalEntries.FunctionalCurrencyCode		
                                       , JournalEntries.JEDescription		
                                       , JournalEntries.JELineDescription		
                                       , JournalEntries.PreparerID		
                                       , JournalEntries.ApproverID		
                               FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries
                               INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                               ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber
                               WHERE ABS(JournalEntries.Amount) > {TE}
                                           {Preparer}
                                           AND JournalEntries.EntryDate BETWEEN '{Point1}' AND '{Point2}'			        	
                                           {Account}	
                               ORDER BY JENumber,JELineNumber			
                            '''.format(field=self.selected_project_id, TE=self.tempTE,
                                       Preparer=self.checked_preparer10,
                                       Account=self.checked_account10, Point1=self.tempPoint1, Point2=self.tempPoint2)

        elif self.rbtn2.isChecked():

            sql = '''
                                   SELECT 			
                                                JournalEntries.BusinessUnit		
                                                , JournalEntries.JENumber		
                                                , JournalEntries.JELineNumber		
                                                , JournalEntries.EffectiveDate		
                                                , JournalEntries.EntryDate
                                                , JournalEntries.Period		
                                                , JournalEntries.GLAccountNumber		
                                                , CoA.GLAccountName		
                                                , JournalEntries.Debit		
                                                , JournalEntries.Credit		
                                                , CASE
                                                        WHEN JournalEntries.Debit = 0 THEN 'Credit' ELSE 'Debit'
                                                        END AS DebitCredit
                                                , JournalEntries.Amount		
                                                , JournalEntries.FunctionalCurrencyCode		
                                                , JournalEntries.JEDescription		
                                                , JournalEntries.JELineDescription		
                                                , JournalEntries.PreparerID		
                                                , JournalEntries.ApproverID		
                               FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS JournalEntries
                               INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                               ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber
                               WHERE EXISTS
                                                    (	
                                                    SELECT DISTINCT JENumber	
                                                    FROM  [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS Journal	
                                                    WHERE JournalEntries.JENumber = Journal.JENumber
                                                        AND ABS(Journal.Amount) > {TE}	
                                                        AND JournalEntries.EntryDate BETWEEN '{Point1}' AND '{Point2}'
                                                        {Preparer}
                                                    ) {Account}	
                               ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber			
                            '''.format(field=self.selected_project_id, TE=self.tempTE,
                                       Preparer=self.checked_preparer10,
                                       Account=self.checked_account10, Point1=self.tempPoint1, Point2=self.tempPoint2)

        self.dataframe = pd.read_sql(sql, self.cnxn)

        ### 마지막 시트 쿼리 내역 추가
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario10",
                                                             "---Filtered Result  Scenario10---\n" + sql]

        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario10",
                                                               "---Filtered JE  Scenario10---\n" + sql]

        ### DebitCredit 열 생성
        if not self.checkD.isChecked() and self.checkC.isChecked():
            self.dataframe = self.dataframe[self.dataframe['DebitCredit'] == 'Credit']
            self.dataframe.reset_index(drop=True, inplace=True)

        elif self.checkD.isChecked() and not self.checkC.isChecked():
            self.dataframe = self.dataframe[self.dataframe['DebitCredit'] == 'Debit']
            self.dataframe.reset_index(drop=True, inplace=True)

        if len(self.dataframe) > 1048576:
            self.communicate10.closeApp.emit()

        elif len(self.dataframe) == 0:
            self.communicate10.closeApp.emit()

        else:

            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)
            self.communicate10.closeApp.emit()

    def extButtonClicked12(self):

        cursor = self.cnxn.cursor()
        sql = '''
                       SET NOCOUNT ON;
                       SELECT JENumber, JELineNumber, GLAccountNumber, Debit, Credit, Amount INTO #tmp
                       FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries]
                       WHERE JENumber IN (
                                          SELECT DISTINCT JENumber
                                          FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries]
                                          WHERE ABS(Amount) > {TE} AND Year = {YEAR}
                                          )
                                         SELECT *                                                                                                       														
                                         FROM                                                                                                    														
                                         (                                                                                                       														
                                               SELECT                                                                                                                                            													
                                                       LVL3.GLAccountNumber1 AS GL_Account_Number,                                                                                                                                											
                                                       MAX(LVL3.GLAccountName1) AS GL_ACcount_Name,                                                                                    											
                                                       MAX(LVL3.AccountType1) AS Account_Type,                                                                                                                              											
                                                       LVL3.DivideDC1 AS GL_Account_Position,                                                                                 											
                                                       CASE                                                                                                                      											
                                                       WHEN LVL3.GLAccountNumber1 = LVL3.GLAccountNumber2 and  LVL3.DivideDC1  = LVL3.DivideDC2 THEN '1.Analysis Account'                                                                                                                            											
                                                       WHEN LVL3.GLAccountNumber1 <> LVL3.GLAccountNumber2 and LVL3.DivideDC1 = LVL3.DivideDC2 THEN '3.Reference Account'                                                                                                                           											
                                                       ELSE '2.Correspondent Account'                                                                                                                   
                                                       END AS Posting_Type,                                                                                                                      
                                                       LVL3.GLAccountNumber2 AS Analysis_GL_Account_Number,                                                                                                                        
                                                       MAX(LVL3.GLAccountName2) AS Analysis_GL_ACcount_Name,                                                                                  
                                                       MAX(LVL3.AccountType2) AS Analysis_Account_Type,                                                                                      
                                                       LVL3.DivideDC2 AS Analysis_Position,                                                                                                            
                                                       SUM(LVL3.SumOfDebit2) AS Sum_Of_Debit_Amount,                                                                                                                                 
                                                       SUM(LVL3.SumOfCredit2) AS Sum_Of_Credit_Amount,                                                                                                                              
                                                       SUM(LVL3.Cnt2) AS JE_Line_Count                                                                                                                                    
                                               FROM                                                                                             
                                               (                                                                                                
                                                       SELECT *                                                                                         
                                                       FROM                                                                                     
                                                              (                                                                                
                                                                             SELECT                                                             
                                                                                    LVL1_1.JENumber1,                                                         
                                                                                    LVL1_1.GLAccountNumber1,                                                          
                                                                                    MAX(LVL1_1.CoA_GLAccountName1) AS GLAccountName1,                                                            			
                                                                                    MAX(LVL1_1.AccountType1) AS AccountType1,                                                      
                                                                                    SUM(LVL1_1.Debit1) AS SumOfDebit1,                                                       
                                                                                    SUM(LVL1_1.Credit1) AS SumOfCredit1,                                                      
                                                                                    DivideDC1,                                                         
                                                                                    COUNT(*) AS Cnt1                                                          
                                                                             FROM                                                               
                                                                             (                                                                  
                                                                                            SELECT                                               
                                                                                                   #tmp.JENumber AS JENumber1,                                          
                                                                                                   #tmp.GLAccountNumber AS GLAccountNumber1,                                          
                                                                                                   CoA.GLAccountNumber AS CoA_GLAccountNumber1,                                       
                                                                                                   CoA.GLAccountName AS CoA_GLAccountName1,                                      
                                                                                                   CoA.AccountType AS AccountType1,                                       
                                                                                                   #tmp.Debit AS Debit1,                                             
                                                                                                   #tmp.Credit AS Credit1,                                            
                                                                                                   #tmp.Amount AS Amount1,                                            
                                                                                                   CASE                                         
                                                                                                   WHEN #tmp.Debit = 0 THEN 'Credit' ELSE 'Debit'                                       
                                                                                                   END AS 'DivideDC1'                                            
                                                                                            FROM #tmp, [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] CoA                                                	
                                                                                            WHERE #tmp.GLAccountNumber = CoA.GLAccountNumber                                                
                                                                             ) LVL1_1                                                                  
                                                                             GROUP BY LVL1_1.JENumber1, LVL1_1.GLAccountNumber1, LVL1_1.DivideDC1                                                                					
                                                              ) LVL2_1,                                                                                
                                                              (                                                                                 
                                                                             SELECT                                                            
                                                                                    LVL1_2.JENumber2,                                                        
                                                                                    LVL1_2.GLAccountNumber2,                                                          
                                                                                    MAX(LVL1_2.CoA_GLAccountName2) AS GLAccountName2,                                                          
                                                                                    MAX(LVL1_2.AccountType2) AS AccountType2,                                                      
                                                                                    SUM(LVL1_2.Debit2) AS SumOfDebit2,                                                       
                                                                                    SUM(LVL1_2.Credit2) AS SumOfCredit2,                                                      
                                                                                    DivideDC2,                                                         
                                                                                    COUNT(*) AS Cnt2                                                          
                                                                             FROM                                                               
                                                                             (                                                                  
                                                                                            SELECT #tmp.JENumber AS JENumber2,                                                  
                                                                                                   #tmp.GLAccountNumber AS GLAccountNumber2,                                          
                                                                                                   CoA.GLAccountNumber AS CoA_GLAccountNumber2,                                       
                                                                                                   CoA.GLAccountName AS CoA_GLAccountName2,                                      
                                                                                                   CoA.AccountType AS AccountType2,                                       
                                                                                                   #tmp.Debit AS Debit2,                                             
                                                                                                   #tmp.Credit AS Credit2,                                            
                                                                                                   #tmp.Amount AS Amount2,                                            
                                                                                                   CASE                                         
                                                                                                   WHEN #tmp.Debit = 0 THEN 'Credit' ELSE 'Debit'                                       
                                                                                                   END AS 'DivideDC2'                                            
                                                                                            FROM #tmp, [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] CoA                                                
                                                                                            WHERE #tmp.GLAccountNumber = CoA.GLAccountNumber                                                
                                                                             ) LVL1_2                                                                 
                                                                             GROUP BY LVL1_2.JENumber2, LVL1_2.GLAccountNumber2, LVL1_2.DivideDC2                                                              
                                                              ) LVL2_2                                                                                 
                                                       WHERE LVL2_1.JENumber1 = LVL2_2.JENumber2                                                                                    
                                               ) LVL3                                                                                                  
                                               GROUP BY LVL3.GLAccountNumber1, LVL3.DivideDC1, LVL3.GLAccountNumber2, LVL3.DivideDC2                                                                                          													
                                        ) LVL4                                                                                                                                                                                                  														
                                        WHERE {CD}
                                              {Account}
                                              AND LVL4.Posting_Type = '2.Correspondent Account'
                                        ORDER BY LVL4.GL_Account_Number, LVL4.GL_Account_Position, LVL4.Posting_Type, LVL4.Analysis_GL_Account_Number     
                   '''.format(field=self.selected_project_id, CD=self.tempState12, Account=self.checked_account12,
                              TE=self.tempCost, YEAR=self.pname_year)

        sql2 = '''
                       SET NOCOUNT ON;
                       DROP TABLE #tmp
                       SELECT JENumber, JELineNumber, GLAccountNumber, Debit, Credit, Amount INTO #tmp
                       FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries]
                       WHERE JENumber IN (
                                          SELECT DISTINCT JENumber
                                          FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries]
                                          WHERE ABS(Amount) > {TE} AND Year = {YEAR}
                                          )
                                         SELECT *                                                                                                       														
                                         FROM                                                                                                    														
                                         (                                                                                                       														
                                               SELECT                                                                                                                                            													
                                                       LVL3.GLAccountNumber1 AS GL_Account_Number,                                                                                                                                											
                                                       MAX(LVL3.GLAccountName1) AS GL_ACcount_Name,                                                                                    											
                                                       MAX(LVL3.AccountType1) AS Account_Type,                                                                                                                              											
                                                       LVL3.DivideDC1 AS GL_Account_Position,                                                                                 											
                                                       CASE                                                                                                                      											
                                                       WHEN LVL3.GLAccountNumber1 = LVL3.GLAccountNumber2 and  LVL3.DivideDC1  = LVL3.DivideDC2 THEN '1.Analysis Account'                                                                                                                            											
                                                       WHEN LVL3.GLAccountNumber1 <> LVL3.GLAccountNumber2 and LVL3.DivideDC1 = LVL3.DivideDC2 THEN '3.Reference Account'                                                                                                                           											
                                                       ELSE '2.Correspondent Account'                                                                                                                   
                                                       END AS Posting_Type,                                                                                                                      
                                                       LVL3.GLAccountNumber2 AS Analysis_GL_Account_Number,                                                                                                                        
                                                       MAX(LVL3.GLAccountName2) AS Analysis_GL_ACcount_Name,                                                                                  
                                                       MAX(LVL3.AccountType2) AS Analysis_Account_Type,                                                                                      
                                                       LVL3.DivideDC2 AS Analysis_Position,                                                                                                            
                                                       SUM(LVL3.SumOfDebit2) AS Sum_Of_Debit_Amount,                                                                                                                                 
                                                       SUM(LVL3.SumOfCredit2) AS Sum_Of_Credit_Amount,                                                                                                                              
                                                       SUM(LVL3.Cnt2) AS JE_Line_Count                                                                                                                                    
                                               FROM                                                                                             
                                               (                                                                                                
                                                       SELECT *                                                                                         
                                                       FROM                                                                                     
                                                              (                                                                                
                                                                             SELECT                                                             
                                                                                    LVL1_1.JENumber1,                                                         
                                                                                    LVL1_1.GLAccountNumber1,                                                          
                                                                                    MAX(LVL1_1.CoA_GLAccountName1) AS GLAccountName1,                                                            			
                                                                                    MAX(LVL1_1.AccountType1) AS AccountType1,                                                      
                                                                                    SUM(LVL1_1.Debit1) AS SumOfDebit1,                                                       
                                                                                    SUM(LVL1_1.Credit1) AS SumOfCredit1,                                                      
                                                                                    DivideDC1,                                                         
                                                                                    COUNT(*) AS Cnt1                                                          
                                                                             FROM                                                               
                                                                             (                                                                  
                                                                                            SELECT                                               
                                                                                                   #tmp.JENumber AS JENumber1,                                          
                                                                                                   #tmp.GLAccountNumber AS GLAccountNumber1,                                          
                                                                                                   CoA.GLAccountNumber AS CoA_GLAccountNumber1,                                       
                                                                                                   CoA.GLAccountName AS CoA_GLAccountName1,                                      
                                                                                                   CoA.AccountType AS AccountType1,                                       
                                                                                                   #tmp.Debit AS Debit1,                                             
                                                                                                   #tmp.Credit AS Credit1,                                            
                                                                                                   #tmp.Amount AS Amount1,                                            
                                                                                                   CASE                                         
                                                                                                   WHEN #tmp.Debit = 0 THEN 'Credit' ELSE 'Debit'                                       
                                                                                                   END AS 'DivideDC1'                                            
                                                                                            FROM #tmp, [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] CoA                                                	
                                                                                            WHERE #tmp.GLAccountNumber = CoA.GLAccountNumber                                                
                                                                             ) LVL1_1                                                                  
                                                                             GROUP BY LVL1_1.JENumber1, LVL1_1.GLAccountNumber1, LVL1_1.DivideDC1                                                                					
                                                              ) LVL2_1,                                                                                
                                                              (                                                                                 
                                                                             SELECT                                                            
                                                                                    LVL1_2.JENumber2,                                                        
                                                                                    LVL1_2.GLAccountNumber2,                                                          
                                                                                    MAX(LVL1_2.CoA_GLAccountName2) AS GLAccountName2,                                                          
                                                                                    MAX(LVL1_2.AccountType2) AS AccountType2,                                                      
                                                                                    SUM(LVL1_2.Debit2) AS SumOfDebit2,                                                       
                                                                                    SUM(LVL1_2.Credit2) AS SumOfCredit2,                                                      
                                                                                    DivideDC2,                                                         
                                                                                    COUNT(*) AS Cnt2                                                          
                                                                             FROM                                                               
                                                                             (                                                                  
                                                                                            SELECT #tmp.JENumber AS JENumber2,                                                  
                                                                                                   #tmp.GLAccountNumber AS GLAccountNumber2,                                          
                                                                                                   CoA.GLAccountNumber AS CoA_GLAccountNumber2,                                       
                                                                                                   CoA.GLAccountName AS CoA_GLAccountName2,                                      
                                                                                                   CoA.AccountType AS AccountType2,                                       
                                                                                                   #tmp.Debit AS Debit2,                                             
                                                                                                   #tmp.Credit AS Credit2,                                            
                                                                                                   #tmp.Amount AS Amount2,                                            
                                                                                                   CASE                                         
                                                                                                   WHEN #tmp.Debit = 0 THEN 'Credit' ELSE 'Debit'                                       
                                                                                                   END AS 'DivideDC2'                                            
                                                                                            FROM #tmp, [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] CoA                                                
                                                                                            WHERE #tmp.GLAccountNumber = CoA.GLAccountNumber                                                
                                                                             ) LVL1_2                                                                 
                                                                             GROUP BY LVL1_2.JENumber2, LVL1_2.GLAccountNumber2, LVL1_2.DivideDC2                                                              
                                                              ) LVL2_2                                                                                 
                                                       WHERE LVL2_1.JENumber1 = LVL2_2.JENumber2                                                                                    
                                               ) LVL3                                                                                                  
                                               GROUP BY LVL3.GLAccountNumber1, LVL3.DivideDC1, LVL3.GLAccountNumber2, LVL3.DivideDC2                                                                                          													
                                        ) LVL4                                                                                                                                                                                                  														
                                        WHERE {CD}
                                              {Account}
                                              AND LVL4.Posting_Type = '2.Correspondent Account'
                                        ORDER BY LVL4.GL_Account_Number, LVL4.GL_Account_Position, LVL4.Posting_Type, LVL4.Analysis_GL_Account_Number     
                   '''.format(field=self.selected_project_id, CD=self.tempState12, Account=self.checked_account12,
                              TE=self.tempCost, YEAR=self.pname_year)

        if self.clickCount == 0:
            self.dataframe = pd.read_sql(sql, self.cnxn)
        else:
            self.dataframe = pd.read_sql(sql2, self.cnxn)
        self.clickCount += 1

        self.dataframe['비경상적계정 선택여부'] = ''

        self.my_query.loc[self.tempSheet + "_Reference"] = [self.tempSheet + "_Reference", "Scenario12",
                                                            "---Filtered Result  Scenario12---\n" + sql]

        if len(self.dataframe) > 1048576:
            self.communicate12.closeApp.emit()

        elif len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame({'No Data': ["[중요성금액: " + str(
                self.tempCost) + "] 라인수 " + str(len(self.dataframe) - 1) + "개입니다"]})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)
            self.scenario_dic[self.tempSheet + '_Reference'] = self.dataframe
            self.combo_sheet.addItem(self.tempSheet + '_Reference')
            self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
            self.communicate12.closeApp.emit()

        else:
            self.scenario_dic[self.tempSheet + '_Reference'] = self.dataframe
            self.combo_sheet.addItem(self.tempSheet + '_Reference')
            self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)
            self.communicate12.closeApp.emit()

    def extButtonClickedC(self):

        index = self.wbC[self.wbC.iloc[:, 12].notnull()].iloc[:, [0, 3, 5, 8]]
        cursorindex = []
        dflist = []
        cursortext = ''
        for i in range(len(index)):
            cursorindex.append("'" + str(index.iloc[i, 0]) + "'" + ',' +
                               "'" + index.iloc[i, 1] + "'" + ',' +
                               "'" + str(index.iloc[i, 2]) + "'" + ',' +
                               "'" + index.iloc[i, 3] + "'")

        for tempcursor in cursorindex:
            cursor = self.cnxn.cursor()
            cursortext = cursortext + tempcursor + '\n'

            if self.rbtn1.isChecked():

                # sql문 수정
                sql = '''
                        SET NOCOUNT ON
                        --****************************************************Filter Table***************************************************							
                        CREATE TABLE #filter							
                        (GLAccountNumber VARCHAR(100), Debit_Credit VARCHAR(100), AL_GLAccountNumber VARCHAR(100), AL_Debit_Credit VARCHAR(100))							
                        INSERT INTO #filter							
                        VALUES							
                        ({cursor})							
                        --****************************************************Insert ProjectID***************************************************							
                        SELECT JENumber,							
                            JELineNumber,						
                            EffectiveDate,						
                            EntryDate,						
                            Period,						
                            GLAccountNumber,						
                            Debit,						
                            Credit,						
                            Amount,						
                            FunctionalCurrencyCode,						
                            JEDescription,						
                            JELineDescription,						
                            PreparerID,						
                            ApproverID  INTO #JEData						
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JE							
                        SELECT * INTO #COAData							
                        FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts]						
                        --****************************************************Result Table***************************************************							
                        CREATE TABLE #result							
                        (JENumber NVARCHAR(100),							
                        JELineNumber BIGINT,							
                        EffectiveDate DATE,							
                        EntryDate DATE,							
                        Period NVARCHAR(25),							
                        GLAccountNumber NVARCHAR(100),							
                        Debit NUMERIC(21,6),							
                        Credit NUMERIC(21,6),							
                        Amount NUMERIC(21,6),							
                        FunctionalCurrencyCode NVARCHAR(50),							
                        JEDescription NVARCHAR(200),							
                        JELineDescription NVARCHAR(200),							
                        PreparerID NVARCHAR(100),							
                        ApproverID NVARCHAR(100)							
                        )							
                        --****************************************************Cursor Start***************************************************							
                        DECLARE cur CURSOR FOR 							
                        SELECT GLAccountNumber, Debit_Credit, AL_GLAccountNumber, AL_Debit_Credit FROM #filter							
                        DECLARE @GLAccountNumber VARCHAR(100)							
                        DECLARE @Debit_Credit VARCHAR(100)							
                        DECLARE @AL_GLAccountNumber VARCHAR(100)							
                        DECLARE @AL_Debit_Credit VARCHAR(100)							
                        OPEN cur							
                        Fetch Next From cur INTO @GLAccountNumber, @Debit_Credit, @AL_GLAccountNumber, @AL_Debit_Credit							
                        WHILE(@@FETCH_STATUS <> -1)							
                        BEGIN;							
                        IF (@Debit_Credit = 'Debit')							
                            IF (@AL_Debit_Credit='Debit') /* Debit/Debit */						
                                INSERT INTO #result (JENumber, JELineNumber, EffectiveDate, EntryDate, Period, GLAccountNumber,Debit,Credit,Amount, 					
                                FunctionalCurrencyCode, JEDescription, JELineDescription, PreparerID, ApproverID)					
                                SELECT JE1.JENumber, JE1.JELineNumber, JE1.EffectiveDate, JE1.EntryDate, JE1.Period, JE1.GLAccountNumber, 					
                                JE1.Debit,JE1.Credit,JE1.Amount, JE1.FunctionalCurrencyCode, JE1.JEDescription, JE1.JELineDescription, JE1.PreparerID, JE1.ApproverID FROM #JEData JE1					
                                WHERE JE1.JENumber IN (					
                                    SELECT DISTINCT(JE1_1.JENumber)				
                                    FROM #JEData JE1_1				
                                    WHERE JE1_1.GLAccountNumber = @GLAccountNumber AND JE1_1.Debit<>0				
                                    ) AND JE1.GLAccountNumber = @AL_GLAccountNumber AND JE1.Debit<>0				
                            ELSE /* Debit/Credit */						
                                INSERT INTO #result (JENumber, JELineNumber, EffectiveDate, EntryDate, Period, GLAccountNumber,Debit,Credit, Amount, 					
                                FunctionalCurrencyCode, JEDescription, JELineDescription, PreparerID, ApproverID)					
                                SELECT JE2.JENumber, JE2.JELineNumber, JE2.EffectiveDate, JE2.EntryDate, JE2.Period, JE2.GLAccountNumber, 					
                                JE2.Debit,JE2.Credit,JE2.Amount, JE2.FunctionalCurrencyCode, JE2.JEDescription, JE2.JELineDescription, JE2.PreparerID, JE2.ApproverID FROM #JEData JE2					
                                WHERE JE2.JENumber IN (					
                                    SELECT DISTINCT(JE2_1.JENumber)				
                                    FROM #JEData JE2_1				
                                    WHERE JE2_1.GLAccountNumber = @GLAccountNumber AND JE2_1.Debit<>0				
                                    ) AND JE2.GLAccountNumber = @AL_GLAccountNumber AND JE2.Credit<>0				
                        ELSE							
                            IF (@AL_Debit_Credit='Debit') /* Credit/Debit */						
                                INSERT INTO #result (JENumber, JELineNumber, EffectiveDate, EntryDate, Period, GLAccountNumber,Debit,Credit, Amount, 					
                                FunctionalCurrencyCode, JEDescription, JELineDescription, PreparerID, ApproverID)					
                                SELECT JE3.JENumber, JE3.JELineNumber, JE3.EffectiveDate, JE3.EntryDate, JE3.Period, JE3.GLAccountNumber, 					
                                JE3.Debit,JE3.Credit,JE3.Amount, JE3.FunctionalCurrencyCode, JE3.JEDescription, JE3.JELineDescription, JE3.PreparerID, JE3.ApproverID FROM #JEData JE3					
                                WHERE JE3.JENumber IN (					
                                    SELECT DISTINCT(JE3_1.JENumber)				
                                    FROM #JEData JE3_1				
                                    WHERE JE3_1.GLAccountNumber = @GLAccountNumber AND JE3_1.Credit<>0				
                                    ) AND JE3.GLAccountNumber = @AL_GLAccountNumber AND JE3.Debit<>0				
                            ELSE /* Credit/Credit */						
                                INSERT INTO #result (JENumber, JELineNumber, EffectiveDate, EntryDate, Period, GLAccountNumber,Debit,Credit, Amount, 					
                                FunctionalCurrencyCode, JEDescription, JELineDescription, PreparerID, ApproverID)					
                                SELECT JE4.JENumber, JE4.JELineNumber, JE4.EffectiveDate, JE4.EntryDate, JE4.Period, JE4.GLAccountNumber, 					
                                JE4.Debit,JE4.Credit,JE4.Amount, JE4.FunctionalCurrencyCode, JE4.JEDescription, JE4.JELineDescription, JE4.PreparerID, JE4.ApproverID FROM #JEData JE4					
                                WHERE JE4.JENumber IN (					
                                    SELECT DISTINCT(JE4_1.JENumber)				
                                    FROM #JEData JE4_1				
                                    WHERE JE4_1.GLAccountNumber = @GLAccountNumber AND JE4_1.Credit<>0				
                                    ) AND JE4.GLAccountNumber = @AL_GLAccountNumber AND JE4.Credit<>0				
                        Fetch Next From cur INTO @GLAccountNumber, @Debit_Credit, @AL_GLAccountNumber, @AL_Debit_Credit							
                        END;							
                        Close cur;							
                        Deallocate cur							
                        --****************************************************Filtered Result_1***************************************************							
                        SELECT JENumber,							
                            JELineNumber,						
                            EffectiveDate,						
                            EntryDate,						
                            Period,						
                            #result.GLAccountNumber,						
                            COA.GLAccountName,						
                            Debit,						
                            Credit,						
                            Amount,						
                            FunctionalCurrencyCode,						
                            JEDescription,						
                            JELineDescription,						
                            PreparerID,						
                            ApproverID						
                        FROM #result 							
                        LEFT JOIN #COAData COA							
                        ON #result.GLAccountNumber = COA.GLAccountNumber
                        DROP TABLE #filter, #JEData,#result,#COAData				
                           '''.format(field=self.selected_project_id, cursor=tempcursor)

            elif self.rbtn2.isChecked():
                sql = '''
                        SET NOCOUNT ON
                    -- Filtered JE				
                        --****************************************************Filter Table***************************************************			
                        CREATE TABLE #filter			
                        (GLAccountNumber VARCHAR(100), Debit_Credit VARCHAR(100), AL_GLAccountNumber VARCHAR(100), AL_Debit_Credit VARCHAR(100))			
                        INSERT INTO #filter			
                        VALUES			
                        ({cursor})			
                        --****************************************************Insert ProjectID***************************************************			
                        SELECT JENumber,			
                            JELineNumber,		
                            EffectiveDate,		
                            EntryDate,		
                            Period,		
                            GLAccountNumber,		
                            Debit,		
                            Credit,		
                            Amount,		
                            FunctionalCurrencyCode,		
                            JEDescription,		
                            JELineDescription,		
                            PreparerID,		
                            ApproverID  INTO #JEData		
                        FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JE			
                        SELECT * INTO #COAData			
                        FROM [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts]			
                        --****************************************************Result Table***************************************************			
                        CREATE TABLE #result			
                        (JENumber NVARCHAR(100),			
                        JELineNumber BIGINT,			
                        EffectiveDate DATE,			
                        EntryDate DATE,			
                        Period NVARCHAR(25),			
                        GLAccountNumber NVARCHAR(100),			
                        Debit NUMERIC(21,6),			
                        Credit NUMERIC(21,6),			
                        Amount NUMERIC(21,6),			
                        FunctionalCurrencyCode NVARCHAR(50),			
                        JEDescription NVARCHAR(200),			
                        JELineDescription NVARCHAR(200),			
                        PreparerID NVARCHAR(100),			
                        ApproverID NVARCHAR(100)			
                        )			
                        --****************************************************Cursor Start***************************************************			
                        DECLARE cur CURSOR FOR 			
                        SELECT GLAccountNumber, Debit_Credit, AL_GLAccountNumber, AL_Debit_Credit FROM #filter			
                        DECLARE @GLAccountNumber VARCHAR(100)			
                        DECLARE @Debit_Credit VARCHAR(100)			
                        DECLARE @AL_GLAccountNumber VARCHAR(100)			
                        DECLARE @AL_Debit_Credit VARCHAR(100)			
                        OPEN cur			
                        Fetch Next From cur INTO @GLAccountNumber, @Debit_Credit, @AL_GLAccountNumber, @AL_Debit_Credit			
                        WHILE(@@FETCH_STATUS <> -1)			
                        BEGIN;			
                        IF (@Debit_Credit = 'Debit')			
                            IF (@AL_Debit_Credit='Debit') /* Debit/Debit */		
                                INSERT INTO #result (JENumber, JELineNumber, EffectiveDate, EntryDate, Period, GLAccountNumber,Debit,Credit,Amount, 	
                                FunctionalCurrencyCode, JEDescription, JELineDescription, PreparerID, ApproverID)	
                                SELECT JE1.JENumber, JE1.JELineNumber, JE1.EffectiveDate, JE1.EntryDate, JE1.Period, JE1.GLAccountNumber, 	
                                JE1.Debit,JE1.Credit,JE1.Amount, JE1.FunctionalCurrencyCode, JE1.JEDescription, JE1.JELineDescription, JE1.PreparerID, JE1.ApproverID FROM #JEData JE1	
                                WHERE JE1.JENumber IN (	
                                    SELECT DISTINCT(JE1_1.JENumber)
                                    FROM #JEData JE1_1
                                    WHERE JE1_1.GLAccountNumber = @GLAccountNumber AND JE1_1.Debit<>0
                                    ) AND JE1.GLAccountNumber = @AL_GLAccountNumber AND JE1.Debit<>0
                            ELSE /* Debit/Credit */		
                                INSERT INTO #result (JENumber, JELineNumber, EffectiveDate, EntryDate, Period, GLAccountNumber,Debit,Credit, Amount, 	
                                FunctionalCurrencyCode, JEDescription, JELineDescription, PreparerID, ApproverID)	
                                SELECT JE2.JENumber, JE2.JELineNumber, JE2.EffectiveDate, JE2.EntryDate, JE2.Period, JE2.GLAccountNumber, 	
                                JE2.Debit,JE2.Credit,JE2.Amount, JE2.FunctionalCurrencyCode, JE2.JEDescription, JE2.JELineDescription, JE2.PreparerID, JE2.ApproverID FROM #JEData JE2	
                                WHERE JE2.JENumber IN (	
                                    SELECT DISTINCT(JE2_1.JENumber)
                                    FROM #JEData JE2_1
                                    WHERE JE2_1.GLAccountNumber = @GLAccountNumber AND JE2_1.Debit<>0
                                    ) AND JE2.GLAccountNumber = @AL_GLAccountNumber AND JE2.Credit<>0
                        ELSE			
                            IF (@AL_Debit_Credit='Debit') /* Credit/Debit */		
                                INSERT INTO #result (JENumber, JELineNumber, EffectiveDate, EntryDate, Period, GLAccountNumber,Debit,Credit, Amount, 	
                                FunctionalCurrencyCode, JEDescription, JELineDescription, PreparerID, ApproverID)	
                                SELECT JE3.JENumber, JE3.JELineNumber, JE3.EffectiveDate, JE3.EntryDate, JE3.Period, JE3.GLAccountNumber, 	
                                JE3.Debit,JE3.Credit,JE3.Amount, JE3.FunctionalCurrencyCode, JE3.JEDescription, JE3.JELineDescription, JE3.PreparerID, JE3.ApproverID FROM #JEData JE3	
                                WHERE JE3.JENumber IN (	
                                    SELECT DISTINCT(JE3_1.JENumber)
                                    FROM #JEData JE3_1
                                    WHERE JE3_1.GLAccountNumber = @GLAccountNumber AND JE3_1.Credit<>0
                                    ) AND JE3.GLAccountNumber = @AL_GLAccountNumber AND JE3.Debit<>0
                            ELSE /* Credit/Credit */		
                                INSERT INTO #result (JENumber, JELineNumber, EffectiveDate, EntryDate, Period, GLAccountNumber,Debit,Credit, Amount, 	
                                FunctionalCurrencyCode, JEDescription, JELineDescription, PreparerID, ApproverID)	
                                SELECT JE4.JENumber, JE4.JELineNumber, JE4.EffectiveDate, JE4.EntryDate, JE4.Period, JE4.GLAccountNumber, 	
                                JE4.Debit,JE4.Credit,JE4.Amount, JE4.FunctionalCurrencyCode, JE4.JEDescription, JE4.JELineDescription, JE4.PreparerID, JE4.ApproverID FROM #JEData JE4	
                                WHERE JE4.JENumber IN (	
                                    SELECT DISTINCT(JE4_1.JENumber)
                                    FROM #JEData JE4_1
                                    WHERE JE4_1.GLAccountNumber = @GLAccountNumber AND JE4_1.Credit<>0
                                    ) AND JE4.GLAccountNumber = @AL_GLAccountNumber AND JE4.Credit<>0
                        Fetch Next From cur INTO @GLAccountNumber, @Debit_Credit, @AL_GLAccountNumber, @AL_Debit_Credit			
                        END;			
                        Close cur;			
                        Deallocate cur			
                        --****************************************************Filtered 전표추출***************************************************			
                        SELECT JENumber,			
                            JELineNumber,		
                            EffectiveDate,		
                            EntryDate,		
                            Period,		
                            #JEData.GLAccountNumber,		
                            #COAData.GLAccountName,		
                            Debit,		
                            Credit,		
                            Amount,		
                            FunctionalCurrencyCode,		
                            JEDescription,		
                            JELineDescription,		
                            PreparerID,		
                            ApproverID		
                        FROM #JEData,#COAData			
                        WHERE #JEData.GLAccountNumber = #COAData.GLAccountNumber AND JENumber IN 			
                            (		
                            select distinct JENumber		
                            from #result,#COAData		
                            where #result.GLAccountNumber = #COAData.GLAccountNumber		
                            )		
                        ORDER BY JENumber,JELineNumber			
                        DROP TABLE #filter, #JEData,#result,#COAData	
                            '''.format(field=self.selected_project_id, cursor=tempcursor)

            readlist = pd.read_sql(sql, self.cnxn)
            dflist.append(readlist)

        self.dataframe = pd.concat(dflist, ignore_index=True)

        ### 마지막 시트 쿼리 내역 추가
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario12",
                                                             "---Filtered Result  Scenario12---\n" + sql]
        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario13",
                                                               "---Filtered JE  Scenario12---\n" + sql]

        if len(self.dataframe) == 0:
            self.communicateC.closeApp2.emit(cursortext)

        else:
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)
            self.communicateC.closeApp2.emit(cursortext)

    def extButtonClicked13(self):

        ### 쿼리 연동
        cursor = self.cnxn.cursor()

        ### JE Line
        if self.rbtn1.isChecked():

            sql_query = '''
                            SELECT
                                JournalEntries.BusinessUnit
                                , JournalEntries.JENumber
                                , JournalEntries.JELineNumber
                                , JournalEntries.EffectiveDate
                                , JournalEntries.EntryDate
                                , JournalEntries.Period
                                , JournalEntries.GLAccountNumber
                                , COA.GLAccountName
                                , JournalEntries.Debit
                                , JournalEntries.Credit
                                , CASE
                                    WHEN JournalEntries.Debit = 0 THEN 'Credit' ELSE 'Debit'
                                    END AS DebitCredit
                                , JournalEntries.Amount
                                , JournalEntries.FunctionalCurrencyCode
                                , JournalEntries.JEDescription
                                , JournalEntries.JELineDescription
                                , JournalEntries.PreparerID
                                , JournalEntries.ApproverID
                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries
                            INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                            ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber
                            WHERE RIGHT(FLOOR(Amount), 6) IN ({CONTI})
                                    {Account}
                                    AND ABS(JournalEntries.Amount) > {TE}
                            ORDER BY JENumber, JELineNumber
                    '''.format(field=self.selected_project_id, TE=self.temp_TE_13, CONTI=self.temp_Continuous,
                               Account=self.checked_account13)
        ### JE - Journals
        elif self.rbtn2.isChecked():

            sql_query = '''
                            SELECT
                                JournalEntries.BusinessUnit
                                , JournalEntries.JENumber
                                , JournalEntries.JELineNumber
                                , JournalEntries.EffectiveDate
                                , JournalEntries.EntryDate
                                , JournalEntries.Period
                                , JournalEntries.GLAccountNumber
                                , COA.GLAccountName
                                , JournalEntries.Debit
                                , JournalEntries.Credit
                                , CASE
                                    WHEN JournalEntries.Debit = 0 THEN 'Credit' ELSE 'Debit'
                                    END AS DebitCredit
                                , JournalEntries.Amount
                                , JournalEntries.FunctionalCurrencyCode
                                , JournalEntries.JEDescription
                                , JournalEntries.JELineDescription
                                , JournalEntries.PreparerID
                                , JournalEntries.ApproverID
                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries
                            INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                            ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber
                            WHERE EXISTS
                            (
                                SELECT DISTINCT JENumber
                                FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS Journal
                                WHERE RIGHT(FLOOR(Amount), 6) IN ({CONTI})
                                      AND JournalEntries.JENumber = Journal.JENumber
                                        {Account}
                                        AND ABS(JournalEntries.Amount) > {TE}
                            )
                            ORDER BY JENumber, JELineNumber
                        '''.format(field=self.selected_project_id, TE=self.temp_TE_13,
                                   CONTI=self.temp_Continuous,
                                   Account=self.checked_account13)

        self.dataframe = pd.read_sql(sql_query, self.cnxn)

        ### 마지막 시트 쿼리 내역 추가
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario13",
                                                             "---Filtered Result  Scenario13---\n" + sql_query]
        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario13",
                                                               "---Filtered JE  Scenario13---\n" + sql_query]

        ### DebitCredit 열 생성
        if not self.checkD.isChecked() and self.checkC.isChecked():
            self.dataframe = self.dataframe[self.dataframe['DebitCredit'] == 'Credit']
            self.dataframe.reset_index(drop=True, inplace=True)

        elif self.checkD.isChecked() and not self.checkC.isChecked():
            self.dataframe = self.dataframe[self.dataframe['DebitCredit'] == 'Debit']
            self.dataframe.reset_index(drop=True, inplace=True)

        ### 예외처리 3 - 최대 추출 라인수
        if len(self.dataframe) > 1048576:
            self.communicate13.closeApp.emit()

        elif len(self.dataframe) == 0:
            self.dataframe = pd.DataFrame({'No Data': ["[연속된 숫자: " + str(self.temp_Continuous) + ','
                                                       + '중요성금액: ' + str(self.temp_TE_13)
                                                       + '] 라인수 ' + str(len(self.dataframe)) + '개입니다']})
            model = DataFrameModel(self.dataframe)
            self.viewtable.setModel(model)

            ### JE Line
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            ### JE
            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)

            self.communicate13.closeApp.emit()

        else:
            ### JE Line
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            ### JE
            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            self.communicate13.closeApp.emit()

    def extButtonClicked14(self):

        cursor = self.cnxn.cursor()

        # sql 문 수정
        if self.rbtn1.isChecked():

            sql = '''
                   SELECT 		
                        JournalEntries.BusinessUnit	
                        , JournalEntries.JENumber	
                        , JournalEntries.JELineNumber	
                        , JournalEntries.EffectiveDate	
                        , JournalEntries.EntryDate	
                        , JournalEntries.Period	
                        , JournalEntries.GLAccountNumber	
                        , CoA.GLAccountName	
                        , JournalEntries.Debit	
                        , JournalEntries.Credit	
                        , CASE
                                    WHEN JournalEntries.Debit = 0 THEN 'Credit' ELSE 'Debit'
                                    END AS DebitCredit
                        , JournalEntries.Amount	
                        , JournalEntries.FunctionalCurrencyCode	
                        , JournalEntries.JEDescription	
                        , JournalEntries.JELineDescription	
                        , JournalEntries.PreparerID	
                        , JournalEntries.ApproverID	
                   FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries
                   INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                   ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber
                   WHERE (JournalEntries.JEDescription LIKE N'%{KEY}%' OR JournalEntries.JELineDescription LIKE N'%{KEY}%')	
                          AND ABS(JournalEntries.Amount) > {TE} {Account}
                   ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber		
                '''.format(field=self.selected_project_id, KEY=self.tempKey, TE=self.tempTE,
                           Account=self.checked_account14)

        elif self.rbtn2.isChecked():

            sql = '''
                   SELECT 			
                        JournalEntries.BusinessUnit		
                        , JournalEntries.JENumber		
                        , JournalEntries.JELineNumber		
                        , JournalEntries.EffectiveDate		
                        , JournalEntries.EntryDate		
                        , JournalEntries.Period		
                        , JournalEntries.GLAccountNumber		
                        , CoA.GLAccountName		
                        , JournalEntries.Debit		
                        , JournalEntries.Credit		
                        , CASE
                                    WHEN JournalEntries.Debit = 0 THEN 'Credit' ELSE 'Debit'
                                    END AS DebitCredit
                        , JournalEntries.Amount		
                        , JournalEntries.FunctionalCurrencyCode		
                        , JournalEntries.JEDescription		
                        , JournalEntries.JELineDescription		
                        , JournalEntries.PreparerID		
                        , JournalEntries.ApproverID		
                   FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] JournalEntries
                   INNER JOIN [{field}_Import_CY_01].[dbo].[pbcChartOfAccounts] AS CoA
                            ON JournalEntries.GLAccountNumber = CoA.GLAccountNumber
                   
                   WHERE EXISTS
                         (		
                            SELECT DISTINCT JournalEntries.JENumber		
                            FROM [{field}_Import_CY_01].[dbo].[pbcJournalEntries] AS Journal		
                            WHERE JournalEntries.JENumber = Journal.JENumber
                            AND (Journal.JEDescription LIKE N'%{KEY}%' OR Journal.JELineDescription LIKE N'%{KEY}%')		
                         )  AND ABS(JournalEntries.Amount) > {TE} {Account}
                   ORDER BY JournalEntries.JENumber, JournalEntries.JELineNumber			
                '''.format(field=self.selected_project_id, KEY=self.tempKey, TE=self.tempTE,
                           Account=self.checked_account14)

        self.dataframe = pd.read_sql(sql, self.cnxn)

        ### 마지막 시트 쿼리 내역 추가
        if self.rbtn1.isChecked():
            self.my_query.loc[self.tempSheet + "_Result"] = [self.tempSheet + "_Result", "Scenario14",
                                                             "---Filtered Result  Scenario14---\n" + sql]
        elif self.rbtn2.isChecked():
            self.my_query.loc[self.tempSheet + "_Journals"] = [self.tempSheet + "_Journals", "Scenario14",
                                                               "---Filtered JE  Scenario14---\n" + sql]

        ### DebitCredit 열 생성
        if not self.checkD.isChecked() and self.checkC.isChecked():
            self.dataframe = self.dataframe[self.dataframe['DebitCredit'] == 'Credit']
            self.dataframe.reset_index(drop=True, inplace=True)

        elif self.checkD.isChecked() and not self.checkC.isChecked():
            self.dataframe = self.dataframe[self.dataframe['DebitCredit'] == 'Debit']
            self.dataframe.reset_index(drop=True, inplace=True)

        if len(self.dataframe) > 1048576:
            self.communicate14.closeApp.emit()

        elif len(self.dataframe) == 0:
            self.communicate14.closeApp.emit()

        else:
            if self.rbtn1.isChecked():
                self.scenario_dic[self.tempSheet + '_Result'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Result')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)

            elif self.rbtn2.isChecked():
                self.scenario_dic[self.tempSheet + '_Journals'] = self.dataframe
                self.combo_sheet.addItem(self.tempSheet + '_Journals')
                self.combo_sheet.setCurrentIndex(self.combo_sheet.count() - 1)
                model = DataFrameModel(self.dataframe)
                self.viewtable.setModel(model)
            self.communicate14.closeApp.emit()

    @pyqtSlot(QModelIndex)
    def slot_clicked_item(self, QModelIndex):
        self.stk_w.setCurrentIndex(QModelIndex.row())

    def saveFile(self):
        if self.dataframe is None:
            self.MessageBox_Open("저장할 데이터가 없습니다")
            return
        if self.scenario_dic == {}:
            self.MessageBox_Open("저장할 Sheet가 없습니다")
            return
        else:
            fileName = QFileDialog.getSaveFileName(self, self.tr("Save Data files"), "./",
                                                   self.tr("xlsx(*.xlsx);; All Files(*.*)"))
            path = fileName[0]

            if path == '':
                pass

            else:

                if os.path.isfile(path):
                    changecount = 0
                    addcount = 0
                    wb = openpyxl.load_workbook(path)
                    wb.create_sheet('Scenario Updated>>>')
                    ws_names = wb.get_sheet_names()
                    for temp in list(self.scenario_dic.keys()):
                        if temp in ws_names:
                            changecount += 1
                            wb.remove(wb['' + temp + ''])
                        else:
                            addcount += 1
                    wb.save(path)
                    with pd.ExcelWriter('' + path + '', mode='a', engine='openpyxl') as writer:
                        for temp in self.scenario_dic:
                            self.scenario_dic['' + temp + ''].to_excel(writer, sheet_name='' + temp + '', index=False,
                                                                       freeze_panes=(1, 0))
                    self.MessageBox_Open("총 " + str(changecount) + "개 시트가 교체\n" + str(addcount) + "개 시트가 추가되었습니다")

                else:
                    with pd.ExcelWriter('' + path + '', mode='w', engine='openpyxl') as writer:
                        for temp in self.scenario_dic:
                            self.scenario_dic['' + temp + ''].to_excel(writer, sheet_name='' + temp + '', index=False,
                                                                       freeze_panes=(1, 0))
                        self.my_query.to_excel(writer, sheet_name='Query', index=False,
                                               freeze_panes=(1, 0))
                    self.MessageBox_Open("저장을 완료했습니다")


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MyApp()
    sys.exit(app.exec_())


# In[ ]:




